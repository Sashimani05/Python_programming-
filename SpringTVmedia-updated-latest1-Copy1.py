#!/usr/bin/env python
# coding: utf-8

# <h1>Table of Contents<span class="tocSkip"></span></h1>
# <div class="toc"><ul class="toc-item"></ul></div>

# In[1]:


import warnings
warnings.filterwarnings("ignore")


# In[2]:


import pandas as pd
import numpy as np
from openpyxl import Workbook
import re
import openpyxl.utils.dataframe
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


# In[3]:


#Display settings
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 3)


# # Check if the Sheet name is availabe

# In[4]:



workbook = load_workbook(filename='C:\\Users\\saraswathy.rajaman\\Downloads\\w86_Spring22_tvshows_short.xlsm')


# In[5]:


name_of_sheet='gerardo'


# In[6]:


def sheets_name_check(workbook,name_of_sheet):
    sheetnames=workbook.sheetnames
    sheetnames=list(sheetnames)
    #print(sheetnames)
    for i in sheetnames:
        if str(name_of_sheet) == i:
            
            print(name_of_sheet,"is found in the workbook",workbook)
            
            break
    else:
        print (name_of_sheet,"not found")


# In[7]:


sheets_name_check(workbook,name_of_sheet)


# # check the necessary headers values on a given worksheet 

# In[8]:


sheet=workbook['gerardo']


# In[9]:


row = sheet.max_row
column = sheet.max_column
  
print("Total Rows:", row)
print("Total Columns:", column)


# In[10]:


#print("\nValue of header")
HV={}
header_values=[]
for i in range(1, column + 1): 
    cell_obj = sheet.cell(row = 1, column = i)
    HV[i]=cell_obj.value  #Stores the column values 
    
    header_values.append(HV[i])
    print(cell_obj.value, end = '\n')


# # input the column header

# In[11]:


list_tbf=['Clean_Type','Network','S2022_Client',
'Show_Name','Section_Heading','List_Heading','Initial_Wave','F2021_Client']


# In[12]:


def get_key(val):
            for key, value in HV.items():# HV items are from the above which is the column names stored as list
                 if val== value:
                    print("found ",val,'in column',key)
                    return key
            
            print ("could not find ",val,'in any column')
            return "key doesn't exist"


# In[13]:


for i in  range(len(list_tbf)):
    val=list_tbf[i]
    get_key(val)
    #call the function above passing every values to check if it is avilable


# In[14]:


Flag={}
for i in list_tbf:
    Flag[i]= i in HV.values() #check if the values are available in the above HV where it stored the first row and allcoumn names


# In[15]:


#Flag.values()


# In[16]:


Flag_header=False in Flag.values()


# In[17]:


Flag_header
#if Flag_header==True
    #sys.exit("Column names are not as expected ")


# # convert  worksheet to dataframes 

# In[18]:



df = pd.read_excel('C:\\Users\\saraswathy.rajaman\\Downloads\\w86_Spring22_tvshows_short.xlsm', sheet_name='gerardo')
pm = pd.read_excel('C:\\Users\\saraswathy.rajaman\\Downloads\\w86_Spring22_tvshows_short.xlsm', sheet_name='PunchMap')


# In[19]:


sheet=workbook['gerardo']


# In[20]:


rows = sheet.max_row
columns = sheet.max_column


# In[21]:


#pm['PunchValue'] = pm['PunchValue'].str.lower()


# In[22]:


pm['PunchValue']=pm['PunchValue'].replace('X','x',regex=True)


# In[23]:


#pm['PunchValue']


# # check if the column header of the given sheet is empty

# if sheet["A1"].value=='Line_Type':
#     print ("First column name is right as",sheet["A1"].value)
# else:
#     print("First column value is not right ")
#     

# In[24]:


#convert the first row as list as the first row is always the column names in a dataframe
my_list = df.columns.values.tolist()


# In[25]:


#slice the string in the list and check if that is unnamed
mylist1= [w[:7] for w in my_list]


# In[26]:


i=0
for x in mylist1:
    if x=='Unnamed':
        print (i," cell in first row  empty")
        print('terminated further execution as the first row has empty cells :')
        Flag_empty=True
        break
        #sys.exit("Column names are not as expected ")
        #if Flag_header==True
        
    i+=1
print("No empty cells in first row")
Flag_Empty=False


# In[27]:


Flag_Empty


# # check the cleantype in two sheets are equal

# In[28]:


firstlist=list(pm['Clean_Type'].unique())

secondlist=list(df['Clean_Type'].unique())


# In[29]:


# function to check both lists if they are equal

def checkList(firstlist, secondlist):
    # sorting the lists
    firstlist.sort()
    secondlist.sort()
    # if both the lists are equal the print yes
    if(firstlist == secondlist):
        print("Both cleantype list are equal")
        Flag_cleantype=True
    else:
        print("Both cleantype list are not equal")
        Flag_cleantype=False

# passing both the lists to checklist function


# In[30]:


checkList(firstlist, secondlist)


# # Read the data from sql invoke the stored procedure

# import pyodbc 
# 
# conn = pyodbc.connect('Driver={SQL Server Native Client 11.0};'
#                       'Server=internalSQLdev.mridevops.com;'
#                       'Database=Codebook_Taxonomy;'
#                       'Trusted_Connection=yes;')
# Prev_Dict_StudyEntryID = '434'
# 
# Prev_Dict_VersionID = '13'
# 
# query = "EXEC [app_Codebook_Read] @VersionID = {0}, @StudyEntryID = {1}".format(Prev_Dict_VersionID, Prev_Dict_StudyEntryID)
# df_dict = pd.read_sql_query(query, conn)
# #df_dict.to_excel(r'C:\Users\saraswathy.rajaman\Documents\df_dict.xls', sheet_name = "Sheet1", header = True, index = False)
# 
# 
# print((df_dict.head(10)))

# In[31]:


#df_dict.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df-dict.txt', sep='\t', index=False,header=True,encoding='cp1252')


# In[32]:


df_dict=pd.read_excel(r'C:\Users\saraswathy.rajaman\Downloads\Winter-2021.xlsx' 
                          ,converters={'ORD': lambda x: f'{x:20}'})


    #data=pandas.read_csv(‘filename.txt’, names=[“Column1”, “Column2”])


# In[33]:


#df_dict['ORD']=df_dict['ORD'].apply(lambda x: '{:.0f}'.format(x))


# In[34]:


#type(df_dict1['ORD'])


# In[35]:


df['compare'] = (df['S2022_Client'] == df['F2021_Client'])


# In[36]:


#df_dict.head(10)


# In[37]:


df['col2pv'] = ''


# In[38]:


#Removing spl character
df['S2022_Client']=df['S2022_Client'].str.replace('*','')
df['F2021_Client']=df['F2021_Client'].str.replace('*','')


# In[39]:


#df.head(5)


# In[40]:


df['Show_name_index']=df.index


# In[41]:


#
#
#display(df.head(10))


# # Drop the supress with X rows and pick necessary column

# In[42]:


df.drop(df.index[df['DP_Status'] == 'X'], inplace = True)


# In[43]:


df=df[['Clean_Type','S2022_Client','Section_Heading','List_Heading', 'DP_Status', 'Show_Name','Show_name_index','Initial_Wave',
'F2021_Client','compare','col2pv']]


# In[44]:


#df.info()


# In[45]:


df.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_aftercleanup.csv',index=False,header=True,encoding='cp1252')


# In[46]:


#group data-df-(TVmedia file) based on cleantype into different dataframes
data={}
grouped = df.groupby('Clean_Type')
for group in grouped.groups.keys():
    #print(group)
    data[group] = grouped.get_group(group)


# In[47]:


#data['add_cabl']


# # Group Punchvalues

# In[48]:


#group punchvalues-pm-(Punchmap file) based on cleantype into different dataframes
PV={}
grouped = pm.groupby('Clean_Type')
for group in grouped.groups.keys():
    #print(group)
    PV[group] = grouped.get_group(group)


# In[49]:


#PV['Movie']


# In[50]:


# Function to find no of columns 1 or 2 pv in each cleantype in punch map dataframe


# In[51]:


def punchv(Punch):
    pm_ct=Punch['Clean_Type'].unique()
    #print (pm_ct,'-',Punch['Columns'].nunique())


# In[52]:



for i in grouped.groups.keys():
    Punch=PV[i]
    punchv(Punch)


# # add_cable

# In[53]:


Punch_variable=PV['add_cabl']['PunchValue']


# In[54]:


#data['add_cabl']['F2021_Client'].dtype


# In[55]:


datapv={}
add_cab=[]
for i in Punch_variable:
   
    datapv[i]=data['add_cabl'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    add_cab.append(datapv[i])


# In[56]:


add_cab=pd.concat(add_cab)


# In[57]:


df_merge_add_cab= pd.merge(add_cab, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[58]:


#df_merge_add_cab.shape


# In[59]:


#df_merge_add_cab.head(10)


# In[60]:


df_merge_add_cab['ORD'].isnull().value_counts()


# In[61]:


df_merge_add_cab['ORD'].isna().value_counts()


# def PV_assign(PV):
#     Punch_variable=PV['add_cabl']['PunchValue']
#     print (Punch_variable)

# PV_assign(PV)

# In[62]:


df_merge_add_cab['LastDigit_PV']=df_merge_add_cab['S2022_Client'].str.strip().str[-1]


# In[63]:


df_merge_add_cab['Detail2'].unique()


# In[64]:


df_merge_add_cab['Detail3'].unique()


# In[65]:


df_merge_add_cab=df_merge_add_cab.sort_values(['LastDigit_PV'], na_position='last',
               ascending=[True])
						  
df_merge_add_cab['Tmpl']=df_merge_add_cab['Tmpl'].fillna(method='ffill')
df_merge_add_cab['Super']=df_merge_add_cab['Super'].fillna(method='ffill')
df_merge_add_cab['Detail3']=df_merge_add_cab['Detail3'].fillna(method='ffill')


# In[66]:


df_merge_add_cab=df_merge_add_cab.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_add_cab['Category']=df_merge_add_cab['Category'].fillna(method='ffill')
df_merge_add_cab['QLevel']=df_merge_add_cab['QLevel'].fillna(method='ffill')
df_merge_add_cab['Detail2']=df_merge_add_cab['Detail2'].fillna(method='ffill')


# In[67]:


df_merge_add_cab['ORD']=df_merge_add_cab['ORD'].astype(str)


# In[68]:


df_merge_add_cab=df_merge_add_cab.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[ ]:





# In[69]:


df_merge_add_cab['CCCC']=df_merge_add_cab["ORD"].str.slice(9,13,1)


# In[70]:


df_merge_add_cab['CCCC']=df_merge_add_cab['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[71]:


df_merge_add_cab['CCCC']=df_merge_add_cab['CCCC'].fillna(method='ffill')


# In[72]:


df_merge_add_cab['DDDD']=df_merge_add_cab.groupby('Show_name_index').ngroup()


# In[73]:


df_merge_add_cab['DDDD']=df_merge_add_cab['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[74]:


df_merge_add_cab=df_merge_add_cab.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[75]:


df_merge_add_cab['ORD']=df_merge_add_cab['ORD'].astype(str)


# In[76]:


df_merge_add_cab['AAAA']=df_merge_add_cab["ORD"].str.slice(0,4,1)


# In[77]:


df_merge_add_cab['BBBB']=df_merge_add_cab["ORD"].str.slice(4,9,1)


# In[78]:


df_merge_add_cab=df_merge_add_cab.sort_values(['Super','Category'],
               ascending=[True,True],na_position='last')


# In[79]:


df_merge_add_cab['AAAA']=df_merge_add_cab['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[80]:


df_merge_add_cab['AAAA']=df_merge_add_cab['AAAA'].replace(r'nan',np.nan, regex=True)


# In[81]:


df_merge_add_cab['AAAA']=df_merge_add_cab['AAAA'].fillna(method='ffill')


# In[82]:


df_merge_add_cab['BBBB']=df_merge_add_cab['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[83]:


df_merge_add_cab['BBBB']=df_merge_add_cab['BBBB'].replace(r'nan',np.nan, regex=True)


# In[84]:


df_merge_add_cab['BBBB']=df_merge_add_cab['BBBB'].fillna(method='ffill')


# In[85]:



df_merge_add_cab['ORD_new'] =df_merge_add_cab['AAAA']+df_merge_add_cab['BBBB']+df_merge_add_cab['CCCC']+df_merge_add_cab['DDDD']


# In[86]:


df_merge_add_cab.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_add_cab.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # Cable 

# In[87]:


Punch_variable=PV['cable']['PunchValue']


# In[88]:


#Punch_variable


# In[89]:


datapv={}
cable=[]
for i in Punch_variable:
   
    datapv[i]=data['cable'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    cable.append(datapv[i])


# In[90]:


cable=pd.concat(cable)


# In[91]:


cable.S2022_Client = cable.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[92]:


df_merge_cable= pd.merge(cable, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[93]:


df_merge_cable.shape


# In[94]:


df_merge_cable['LastDigit_PV']=df_merge_cable['S2022_Client'].str.strip().str[-1]


# In[95]:


df2=df_merge_cable['LastDigit_PV'].groupby(df_merge_cable['Detail3']).unique().apply(pd.Series)
#df['subreddit'].groupby(df['author']).unique().apply(pd.Series)


# In[96]:


df2=pd.DataFrame(df2)
df2.reset_index(inplace=True)


# In[97]:


#df2.loc([['Full Attention', 'Most Attention', 'Some Attention', 'Watched last 30 days', 'Watched last 7 days']])


# In[98]:


#df2


# In[99]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[100]:


df_merge_cable['ORD']=df_merge_cable['ORD'].astype(str)


# In[101]:


df_merge_cable['CCCC']=df_merge_cable["ORD"].str.slice(9,13,1)


# In[102]:


df3=df_merge_cable['Detail3'].groupby(df_merge_cable['CCCC']).unique().apply(pd.Series)


# In[103]:


df3=pd.DataFrame(df3)


# In[104]:


df3.reset_index(inplace=True)


# In[105]:


df3.dropna(inplace=True)


# In[106]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[107]:


df3.dropna(inplace=True)


# In[108]:


df_merge_cable=pd.merge(df_merge_cable,df2,on='LastDigit_PV',how='left')


# In[109]:


df_merge_cable.shape


# In[110]:


df_merge_cable.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_cable_detail3.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[111]:


#df_merge_cable.drop("Detail3_x",axis='columns',inplace=True)


# In[112]:


df_merge_cable=df_merge_cable.rename(columns={"Detail3_y":"Detail3"})


# In[113]:


df_merge_cable=pd.merge(df_merge_cable,df3,on='Detail3',how='left')


# In[114]:


df_merge_cable=df_merge_cable.rename(columns={"CCCC_y":"CCCC"})


# In[115]:



df_merge_cable=df_merge_cable.sort_values(['LastDigit_PV'],ascending=[True],na_position='last')


# In[116]:


df_merge_cable['Tmpl']=df_merge_cable['Tmpl'].fillna(method='ffill')
df_merge_cable['Super']=df_merge_cable['Super'].fillna(method='ffill')
#df_merge_cable['Detail3']=df_merge_cable['Detail3'].fillna(method='ffill')
df_merge_cable['Tmpl']=df_merge_cable['Tmpl'].fillna(2)
df_merge_cable['Super']=df_merge_cable['Super'].fillna('Media - Cable')


# In[117]:


df_merge_cable=df_merge_cable.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_cable['Category']=df_merge_cable['Category'].fillna(method='ffill')
df_merge_cable['QLevel']=df_merge_cable['QLevel'].fillna(method='ffill')
#df_merge_cable['Detail2']=df_merge_cable['Detail2'].fillna(method='ffill')


# In[118]:


df_merge_cable['ORD']=df_merge_cable['ORD'].astype(str)


# In[119]:



df_merge_cable=df_merge_cable.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[120]:



df_merge_cable['CCCC']=df_merge_cable['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[121]:


#df_merge_cable['CCCC']=df_merge_cable['CCCC'].fillna(method='ffill')


# In[122]:


df_merge_cable['DDDD']=df_merge_cable.groupby('Show_name_index').ngroup()


# In[123]:


df_merge_cable['DDDD']=df_merge_cable['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[124]:


df_merge_cable=df_merge_cable.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[125]:


df_merge_cable['ORD']=df_merge_cable['ORD'].astype(str)


# In[126]:


df_merge_cable['AAAA']=df_merge_cable["ORD"].str.slice(0,4,1)


# In[127]:


df_merge_cable['BBBB']=df_merge_cable["ORD"].str.slice(4,9,1)


# In[128]:


df_merge_cable=df_merge_cable.sort_values(['Super','Category'],
               ascending=[True,True],na_position='last')


# In[129]:


df_merge_cable['AAAA']=df_merge_cable['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[130]:


df_merge_cable['AAAA']=df_merge_cable['AAAA'].replace(r'nan',np.nan, regex=True)


# In[131]:


df_merge_cable['AAAA']=df_merge_cable['AAAA'].fillna('1098')


# In[132]:


df_merge_cable['BBBB']=df_merge_cable['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[133]:


df_merge_cable['BBBB']=df_merge_cable['BBBB'].replace(r'nan',np.nan, regex=True)


# In[134]:


df_merge_cable['BBBB']=df_merge_cable['BBBB'].fillna('00167')


# In[135]:


df_merge_cable['ORD_new'] =df_merge_cable['AAAA']+df_merge_cable['BBBB']+df_merge_cable['CCCC']+df_merge_cable['DDDD']


# In[136]:


df_merge_cable['Show_Name'] = df_merge_cable.apply(lambda x: x['List_Heading']+': '+x['Show_Name'] , axis=1)


# In[137]:



df_merge_cable.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_cable.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[138]:


from sqlalchemy import create_engine


# In[139]:


DB = {'server':'internalSQLdev.mridevops.com','database':'Codebook_Taxonomy','driver':'driver=SQL Server Native Client 11.0','pyodb_d':'SQL Server Native Client 11.0'}
#engine=create_engine('mssql+pyodbc://'+ DB['server']+'/'+ DB['database']+'?'+ DB['driver'])


# In[140]:



engine = create_engine('mssql+pyodbc://' + DB['server'] + '/' + DB['database'] + '?' + DB['driver'], fast_executemany = True)


# In[141]:


import pyodbc


# In[142]:


#conn = pyodbc.connect('Driver={'+DB['pyodb_d']+'}; Server='+DB['server']+';Database='+DB['database']+'; Trusted_Connection=yes;')


# In[143]:


#conn.commit()


# with engine.begin() as connection:
#     df_merge_add_cab.to_sql(name="tmp_EditedRecords_addcable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
# #df.to_sql('db_table2', engine, if_exists='replace')
# df_merge_cable.to_sql(name="tmp_EditedRecords_cable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
# #

# # Movies

# In[144]:


Punch_variable=PV['Movie']['PunchValue']

#Punch_variable=PV['add_cabl']['PunchValue']


# In[ ]:





# In[145]:


datapv={}
Movie=[]
for i in Punch_variable:
   
    datapv[i]=data['Movie'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    Movie.append(datapv[i])


# In[146]:


#PV['Movie']['PunchValue']


# In[147]:


#Movie


# In[148]:


Movie=pd.concat(Movie)


# In[149]:


#Movie


# In[150]:


Movie.F2021_Client = Movie.F2021_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[151]:


df_merge_Movie= pd.merge(Movie, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[152]:


df_merge_Movie['LastDigit_PV']=df_merge_Movie['S2022_Client'].str.strip().str[-1]


# In[153]:


#df_merge_Movie
df_merge_Movie.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_Movie_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[154]:


df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '1', df_merge_Movie['Detail3'].fillna('Saw at movie theater'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '2', df_merge_Movie['Detail3'].fillna('Rented movie and viewed on DVD or Blu-ray'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '3', df_merge_Movie['Detail3'].fillna('Purchased movie and viewed on DVD or Blu-ray'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '4', df_merge_Movie['Detail3'].fillna('Viewed with Video On Demand or PPV'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3']= np.where(df_merge_Movie['LastDigit_PV'] == '5', df_merge_Movie['Detail3'].fillna('Downloaded or Streamed from the Internet'),df_merge_Movie['Detail3'])


# In[155]:


#df_merge_Movie
#df_merge_Movie
df_merge_Movie.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_Movie_beforefill_arranged.txt", sep='\t', index=False,header=True,encoding='cp1252')


# df_merge_Movie=df_merge_Movie.sort_values(['LastDigit_PV','ORD'], 
#                ascending=[True,True],na_position='last')

# In[156]:


df_merge_Movie=df_merge_Movie.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_Movie['Tmpl']=df_merge_Movie['Tmpl'].fillna(method='ffill')
df_merge_Movie['Super']=df_merge_Movie['Super'].fillna(method='ffill')
#f_merge_Movie['Detail3']=df_merge_Movie['Detail3'].fillna(method='ffill')


# In[157]:


df_merge_Movie=df_merge_Movie.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_Movie['Category']=df_merge_Movie['Category'].fillna(method='ffill')
df_merge_Movie['QLevel']=df_merge_Movie['QLevel'].fillna(method='ffill')
df_merge_Movie['Detail2']=df_merge_Movie['Detail2'].fillna(method='ffill')


# In[158]:


df_merge_Movie['ORD']=df_merge_Movie['ORD'].astype(str)


# In[159]:


df_merge_Movie=df_merge_Movie.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[160]:


df_merge_Movie['CCCC']=df_merge_Movie["ORD"].str.slice(9,13,1)


# In[161]:


df_merge_Movie['CCCC']=df_merge_Movie['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[162]:


#df_merge_Movie['CCCC']=df_merge_Movie['CCCC'].fillna(method='ffill')


# In[163]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Downloaded or Streamed from the Internet', df_merge_Movie['CCCC'].fillna('0004'),df_merge_Movie['CCCC'])


# In[164]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Purchased movie and viewed on DVD or Blu-ray', df_merge_Movie['CCCC'].fillna('0002'),df_merge_Movie['CCCC'])


# In[165]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Rented movie and viewed on DVD or Blu-ray', df_merge_Movie['CCCC'].fillna('0001'),df_merge_Movie['CCCC'])


# In[166]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Saw at movie theater', df_merge_Movie['CCCC'].fillna('0000'),df_merge_Movie['CCCC'])


# In[167]:



df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Viewed with Video On Demand or PPV', df_merge_Movie['CCCC'].fillna('0003'),df_merge_Movie['CCCC'])


# In[168]:


df_merge_Movie['DDDD']=df_merge_Movie.groupby('Show_name_index').ngroup()


# In[169]:



df_merge_Movie['DDDD']=df_merge_Movie['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[170]:



df_merge_Movie=df_merge_Movie.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[171]:


df_merge_Movie['ORD']=df_merge_Movie['ORD'].astype(str)


# In[172]:


df_merge_Movie['AAAA']=df_merge_Movie["ORD"].str.slice(0,4,1)


# In[173]:


df_merge_Movie['BBBB']=df_merge_Movie["ORD"].str.slice(4,9,1)


# In[174]:


df_merge_Movie=df_merge_Movie.sort_values(['Super','Category'],
               ascending=[True,True],na_position='last')


# In[175]:


df_merge_Movie['AAAA']=df_merge_Movie['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[176]:


df_merge_Movie['AAAA']=df_merge_Movie['AAAA'].replace(r'nan',np.nan, regex=True)


# In[177]:



df_merge_Movie['AAAA']=df_merge_Movie['AAAA'].fillna('1069')


# In[178]:


df_merge_Movie['BBBB']=df_merge_Movie['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[179]:


df_merge_Movie['BBBB']=df_merge_Movie['BBBB'].replace(r'nan',np.nan, regex=True)


# In[180]:


df_merge_Movie['BBBB']=df_merge_Movie['BBBB'].fillna('00157')


# In[181]:


df_merge_Movie['ORD_new'] =df_merge_Movie['AAAA']+df_merge_Movie['BBBB']+df_merge_Movie['CCCC']+df_merge_Movie['DDDD']


# In[182]:


df_merge_Movie.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_Movie.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[183]:


#df_merge_Movie.head(10)


# In[ ]:





# # SPTV1

# In[184]:


Punch_variable=PV['SPTV1']['PunchValue']


# In[185]:


datapv={}
SPTV1=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV1'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV1.append(datapv[i])


# In[186]:


#PV['SPTV1']['PunchValue']


# In[187]:


SPTV1=pd.concat(SPTV1)


# In[188]:


SPTV1.S2022_Client = SPTV1.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[189]:


df_merge_SPTV1= pd.merge(SPTV1, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[190]:


df_merge_SPTV1['LastDigit_PV']=df_merge_SPTV1['S2022_Client'].str.strip().str[-1]


# In[191]:


#df_merge_SPTV1.head(10)


# In[192]:


df_merge_SPTV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV1_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[193]:


df_merge_SPTV1['ORD']=df_merge_SPTV1['ORD'].astype(str)


# In[194]:


df_merge_SPTV1['AAAA']=df_merge_SPTV1["ORD"].str.slice(0,4,1)

df_merge_SPTV1['BBBB']=df_merge_SPTV1["ORD"].str.slice(4,9,1)


# In[195]:


df_merge_SPTV1['CCCC']=df_merge_SPTV1["ORD"].str.slice(9,13,1)


# In[196]:


df_merge_SPTV1=df_merge_SPTV1.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV1['Category']=df_merge_SPTV1['Category'].fillna(method='ffill')
df_merge_SPTV1['QLevel']=df_merge_SPTV1['QLevel'].fillna(method='ffill')
#df_merge_SPTV1['Detail2']=df_merge_SPTV1['Detail2'].fillna(method='ffill')


# In[197]:


df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'nan',np.nan, regex=True)


# In[198]:


df_merge_SPTV1['BBBB']=df_merge_SPTV1['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV1['BBBB']=df_merge_SPTV1['BBBB'].replace(r'nan',np.nan, regex=True)


# In[199]:


df2=df_merge_SPTV1['LastDigit_PV'].groupby(df_merge_SPTV1['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)


# In[200]:


type(df2)


# In[201]:


df2.reset_index(inplace=True)


# In[202]:



df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[203]:


df3=df_merge_SPTV1['Detail3'].groupby(df_merge_SPTV1['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[204]:


df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[205]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df2,on='LastDigit_PV',how='left')


# In[206]:


df_merge_SPTV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV1.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[207]:


df_merge_SPTV1.columns


# In[208]:


df_merge_SPTV1.drop("Detail3_x",axis='columns',inplace=True)


# In[209]:


df_merge_SPTV1=df_merge_SPTV1.rename(columns={"Detail3_y":"Detail3"})


# In[210]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df3,on='Detail3',how='left')


# In[211]:



df_merge_SPTV1=df_merge_SPTV1.rename(columns={"CCCC_y":"CCCC"})


# In[212]:


df_merge_SPTV1=df_merge_SPTV1.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV1['Tmpl']=df_merge_SPTV1['Tmpl'].fillna(method='ffill')
df_merge_SPTV1['Super']=df_merge_SPTV1['Super'].fillna(method='ffill')


# In[213]:



df_merge_SPTV1=df_merge_SPTV1.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV1['Category']=df_merge_SPTV1['Category'].fillna(method='ffill')
df_merge_SPTV1['QLevel']=df_merge_SPTV1['QLevel'].fillna(method='ffill')
#df_merge_SPTV1['Detail2']=df_merge_SPTV1['Detail2'].fillna(method='ffill')


# In[214]:


df_merge_SPTV1=df_merge_SPTV1.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[215]:


df_merge_SPTV1['CCCC']=df_merge_SPTV1['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[216]:


df_merge_SPTV1['DDDD']=df_merge_SPTV1.groupby('Show_name_index').ngroup()


# In[217]:


df_merge_SPTV1['DDDD']=df_merge_SPTV1['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[218]:



df_merge_SPTV1=df_merge_SPTV1.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[219]:


df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'nan',np.nan, regex=True)


# In[220]:


df4=df_merge_SPTV1['List_Heading'].groupby(df_merge_SPTV1['AAAA']).unique().apply(pd.Series)


# In[221]:


#df4


# In[222]:


df4.reset_index(inplace=True)


# In[223]:


df4=pd.DataFrame(df4)


# In[224]:


#df4.List_Heading


# In[225]:


df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[226]:


#df4['List_Heading']


# In[227]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df4,on='List_Heading',how='left')


# In[228]:


#df_merge_SPTV1.columns


# In[229]:


df_merge_SPTV1.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[230]:


df5=df_merge_SPTV1['List_Heading'].groupby(df_merge_SPTV1['BBBB']).unique().apply(pd.Series)


# In[231]:


#df5


# In[232]:


df5.reset_index(inplace=True)


# In[233]:


df5=pd.DataFrame(df5)


# In[234]:


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[235]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df5,on='List_Heading',how='left')


# In[236]:


df_merge_SPTV1.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[237]:


#df_merge_SPTV1.columns


# In[238]:


df_merge_SPTV1['ORD_new'] =df_merge_SPTV1['AAAA']+df_merge_SPTV1['BBBB']+df_merge_SPTV1['CCCC']+df_merge_SPTV1['DDDD']


# In[239]:


df_merge_SPTV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV1_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


#  # SPTV2

# In[240]:


Punch_variable=PV['SPTV2']['PunchValue']


# In[241]:


datapv={}
SPTV2=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV2'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV2.append(datapv[i])


# In[242]:


SPTV2=pd.concat(SPTV2)


# In[243]:



SPTV2.S2022_Client = SPTV2.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[244]:


df_merge_SPTV2= pd.merge(SPTV2, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[245]:


df_merge_SPTV2['LastDigit_PV']=df_merge_SPTV2['S2022_Client'].str.strip().str[-1]


# In[246]:


df_merge_SPTV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV2_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[247]:


df_merge_SPTV2['ORD']=df_merge_SPTV2['ORD'].astype(str)


# In[248]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2["ORD"].str.slice(0,4,1)


# In[249]:



df_merge_SPTV2['BBBB']=df_merge_SPTV2["ORD"].str.slice(4,9,1)


# In[250]:


df_merge_SPTV2['CCCC']=df_merge_SPTV2["ORD"].str.slice(9,13,1)


# In[251]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV2['Tmpl']=df_merge_SPTV2['Tmpl'].fillna(method='ffill')
df_merge_SPTV2['Super']=df_merge_SPTV2['Super'].fillna(method='ffill')


# In[252]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV2['Category']=df_merge_SPTV2['Category'].fillna(method='ffill')
df_merge_SPTV2['QLevel']=df_merge_SPTV2['QLevel'].fillna(method='ffill')
df_merge_SPTV2['Detail2']=df_merge_SPTV2['Detail2'].fillna(method='ffill')


# In[253]:



df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[254]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[255]:


df_merge_SPTV2['BBBB']=df_merge_SPTV2['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[256]:


df_merge_SPTV2['BBBB']=df_merge_SPTV2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[257]:


df2=df_merge_SPTV2['LastDigit_PV'].groupby(df_merge_SPTV2['Detail3']).unique().apply(pd.Series)


# In[258]:



df2=pd.DataFrame(df2)


# In[259]:


df2.reset_index(inplace=True)


# In[260]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[261]:


df3=df_merge_SPTV2['Detail3'].groupby(df_merge_SPTV2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[262]:


df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[263]:


df_merge_SPTV2=pd.merge(df_merge_SPTV2,df2,on='LastDigit_PV',how='left')


# In[264]:


df_merge_SPTV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[265]:


df_merge_SPTV2.drop("Detail3_x",axis='columns',inplace=True)


# In[266]:


df_merge_SPTV2=df_merge_SPTV2.rename(columns={"Detail3_y":"Detail3"})


# In[267]:


df_merge_SPTV2=pd.merge(df_merge_SPTV2,df3,on='Detail3',how='left')


# In[268]:


df_merge_SPTV2=df_merge_SPTV2.rename(columns={"CCCC_y":"CCCC"})


# In[269]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

#df_merge_SPTV2['CCCC']=df_merge_SPTV2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[270]:


df_merge_SPTV2['DDDD']=df_merge_SPTV2.groupby('Show_name_index').ngroup()

df_merge_SPTV2['DDDD']=df_merge_SPTV2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[271]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[272]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[273]:


df4=df_merge_SPTV2['List_Heading'].groupby(df_merge_SPTV2['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)


# In[274]:


df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[275]:


df_merge_SPTV2=pd.merge(df_merge_SPTV2,df4,on='List_Heading',how='left')


# In[276]:


df_merge_SPTV2.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[277]:


df5=df_merge_SPTV2['List_Heading'].groupby(df_merge_SPTV2['BBBB']).unique().apply(pd.Series)


# In[278]:


df5.reset_index(inplace=True)


# In[279]:


df5=pd.DataFrame(df5)


# In[280]:


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[281]:



df_merge_SPTV2=pd.merge(df_merge_SPTV2,df5,on='List_Heading',how='left')

df_merge_SPTV2.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[282]:


#df_merge_SPTV2.columns


# In[283]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].astype(str)


# In[284]:


df_merge_SPTV2['BBBB']=df_merge_SPTV2['BBBB'].astype(str)


# In[285]:


df_merge_SPTV2['CCCC']=df_merge_SPTV2['CCCC'].astype(str)


# In[286]:


df_merge_SPTV2['DDDD']=df_merge_SPTV2['DDDD'].astype(str)


# In[287]:


df_merge_SPTV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV2_1.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[288]:


#df_merge_SPTV2.info()


# In[289]:


df_merge_SPTV2['ORD_new'] =df_merge_SPTV2['AAAA']+df_merge_SPTV2['BBBB']+df_merge_SPTV2['CCCC']+df_merge_SPTV2['DDDD']


# # SPTV3

# In[290]:


Punch_variable=PV['SPTV3']['PunchValue']


# In[291]:


datapv={}
SPTV3=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV3'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV3.append(datapv[i])


# In[292]:


SPTV3=pd.concat(SPTV3)


# In[293]:


SPTV3.S2022_Client = SPTV3.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[294]:


df_merge_SPTV3= pd.merge(SPTV3, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[295]:


#df_merge_SPTV3


# In[296]:


df_merge_SPTV3['LastDigit_PV']=df_merge_SPTV3['S2022_Client'].str.strip().str[-1]


# In[297]:


df_merge_SPTV3['ORD']=df_merge_SPTV3['ORD'].astype(str)


# In[298]:



df_merge_SPTV3['AAAA']=df_merge_SPTV3["ORD"].str.slice(0,4,1)

df_merge_SPTV3['BBBB']=df_merge_SPTV3["ORD"].str.slice(4,9,1)

df_merge_SPTV3['CCCC']=df_merge_SPTV3["ORD"].str.slice(9,13,1)


# In[299]:


df_merge_SPTV3=df_merge_SPTV3.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV3['Category']=df_merge_SPTV3['Category'].fillna(method='ffill')
df_merge_SPTV3['QLevel']=df_merge_SPTV3['QLevel'].fillna(method='ffill')
#df_merge_SPTV3['Detail2']=df_merge_SPTV3['Detail2'].fillna(method='ffill')


# In[300]:


df_merge_SPTV3=df_merge_SPTV3.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV3['Tmpl']=df_merge_SPTV3['Tmpl'].fillna(method='ffill')
df_merge_SPTV3['Super']=df_merge_SPTV3['Super'].fillna(method='ffill')


# In[301]:


df_merge_SPTV3['AAAA']=df_merge_SPTV3['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV3['AAAA']=df_merge_SPTV3['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV3['BBBB']=df_merge_SPTV3['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV3['BBBB']=df_merge_SPTV3['BBBB'].replace(r'nan',np.nan, regex=True)


# In[302]:


df2=df_merge_SPTV3['LastDigit_PV'].groupby(df_merge_SPTV3['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[303]:


#df2


# In[304]:


df_merge_SPTV3=pd.merge(df_merge_SPTV3,df2,on='LastDigit_PV',how='left')


# In[305]:


#df_merge_SPTV3


# In[306]:


df_merge_SPTV3.columns


# In[307]:


df_merge_SPTV3.drop("Detail3_x",axis='columns',inplace=True)

df_merge_SPTV3=df_merge_SPTV3.rename(columns={"Detail3_y":"Detail3"})


# In[308]:


df3=df_merge_SPTV3['Detail3'].groupby(df_merge_SPTV3['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[309]:


df3.drop(0,inplace=True)


# In[310]:


df3.dropna(axis=1,inplace=True)


# In[311]:


df_merge_SPTV3=pd.merge(df_merge_SPTV3,df3,on='Detail3',how='left')


# In[312]:


#df3


# In[313]:


df_merge_SPTV3=df_merge_SPTV3.rename(columns={"CCCC_y":"CCCC"})


# In[314]:


#df_merge_SPTV3


# In[315]:


df_merge_SPTV3=df_merge_SPTV3.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV3['CCCC']=df_merge_SPTV3['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[316]:


df_merge_SPTV3['DDDD']=df_merge_SPTV3.groupby('Show_name_index').ngroup()

df_merge_SPTV3['DDDD']=df_merge_SPTV3['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[317]:


#df_merge_SPTV3


# In[318]:


df4=df_merge_SPTV3['List_Heading'].groupby(df_merge_SPTV3['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV3=pd.merge(df_merge_SPTV3,df4,on='List_Heading',how='left')


# In[ ]:





# In[319]:


df_merge_SPTV3.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[320]:


df5=df_merge_SPTV3['List_Heading'].groupby(df_merge_SPTV3['BBBB']).unique().apply(pd.Series)


# In[321]:


df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV3=pd.merge(df_merge_SPTV3,df5,on='List_Heading',how='left')

df_merge_SPTV3.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[322]:


df_merge_SPTV3['ORD_new'] =df_merge_SPTV3['AAAA']+df_merge_SPTV3['BBBB']+df_merge_SPTV3['CCCC']+df_merge_SPTV3['DDDD']


# In[323]:


df_merge_SPTV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV3_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[324]:


#df_merge_SPTV3


# # SPTV4

# In[325]:


Punch_variable=PV['SPTV4']['PunchValue']


# In[326]:


#Punch_variable


# In[327]:


datapv={}
SPTV4=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV4'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV4.append(datapv[i])


# In[328]:


SPTV4=pd.concat(SPTV4)


# In[329]:


SPTV4.S2022_Client = SPTV4.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[330]:


df_merge_SPTV4= pd.merge(SPTV4, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[331]:


df_merge_SPTV4['LastDigit_PV']=df_merge_SPTV4['S2022_Client'].str.strip().str[-1]


# In[332]:


df_merge_SPTV4['ORD']=df_merge_SPTV4['ORD'].astype(str)

df_merge_SPTV4['AAAA']=df_merge_SPTV4["ORD"].str.slice(0,4,1)

df_merge_SPTV4['BBBB']=df_merge_SPTV4["ORD"].str.slice(4,9,1)

df_merge_SPTV4['CCCC']=df_merge_SPTV4["ORD"].str.slice(9,13,1)


# In[333]:


df_merge_SPTV4=df_merge_SPTV4.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV4['Category']=df_merge_SPTV4['Category'].fillna(method='ffill')
df_merge_SPTV4['QLevel']=df_merge_SPTV4['QLevel'].fillna(method='ffill')
df_merge_SPTV4['Detail2']=df_merge_SPTV4['Detail2'].fillna(method='ffill')


# In[334]:


df_merge_SPTV4=df_merge_SPTV4.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV4['Tmpl']=df_merge_SPTV4['Tmpl'].fillna(method='ffill')
df_merge_SPTV4['Super']=df_merge_SPTV4['Super'].fillna(method='ffill')


# In[335]:



df_merge_SPTV4['AAAA']=df_merge_SPTV4['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV4['AAAA']=df_merge_SPTV4['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV4['BBBB']=df_merge_SPTV4['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV4['BBBB']=df_merge_SPTV4['BBBB'].replace(r'nan',np.nan, regex=True)


# In[336]:



df2=df_merge_SPTV4['LastDigit_PV'].groupby(df_merge_SPTV4['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[337]:


df3=df_merge_SPTV4['Detail3'].groupby(df_merge_SPTV4['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)

df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df2,on='LastDigit_PV',how='left')


# In[338]:


df_merge_SPTV4.drop("Detail3_x",axis='columns',inplace=True)

df_merge_SPTV4=df_merge_SPTV4.rename(columns={"Detail3_y":"Detail3"})

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df3,on='Detail3',how='left')


# In[339]:


df_merge_SPTV4=df_merge_SPTV4.rename(columns={"CCCC_y":"CCCC"})


# In[340]:


f_merge_SPTV4=df_merge_SPTV4.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV4['CCCC']=df_merge_SPTV4['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[341]:


df_merge_SPTV4['DDDD']=df_merge_SPTV4.groupby('Show_name_index').ngroup()

df_merge_SPTV4['DDDD']=df_merge_SPTV4['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[342]:


df4=df_merge_SPTV4['List_Heading'].groupby(df_merge_SPTV4['AAAA']).unique().apply(pd.Series)
df4=pd.DataFrame(df4)


df4.reset_index(inplace=True)


df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df4,on='List_Heading',how='left')


# In[343]:


df_merge_SPTV4.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[344]:


df5=df_merge_SPTV4['List_Heading'].groupby(df_merge_SPTV4['BBBB']).unique().apply(pd.Series)


df5=pd.DataFrame(df5)

df5.reset_index(inplace=True)


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df5,on='List_Heading',how='left')

df_merge_SPTV4.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[345]:


df_merge_SPTV4['ORD_new'] =df_merge_SPTV4['AAAA']+df_merge_SPTV4['BBBB']+df_merge_SPTV4['CCCC']+df_merge_SPTV4['DDDD']


# In[346]:


df_merge_SPTV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV4_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # SPTV5

# In[347]:


Punch_variable=PV['SPTV5']['PunchValue']


# In[348]:


Punch_variable


# In[349]:


datapv={}
SPTV5=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV5'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV5.append(datapv[i])
	


# In[350]:


SPTV5=pd.concat(SPTV5)


# In[351]:


SPTV5.S2022_Client = SPTV5.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[352]:


df_merge_SPTV5= pd.merge(SPTV5, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')

df_merge_SPTV5['LastDigit_PV']=df_merge_SPTV5['S2022_Client'].str.strip().str[-1]


# In[353]:


df_merge_SPTV5['ORD']=df_merge_SPTV5['ORD'].astype(str)

df_merge_SPTV5['AAAA']=df_merge_SPTV5["ORD"].str.slice(0,4,1)

df_merge_SPTV5['BBBB']=df_merge_SPTV5["ORD"].str.slice(4,9,1)

df_merge_SPTV5['CCCC']=df_merge_SPTV5["ORD"].str.slice(9,13,1)


# In[354]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV5['Category']=df_merge_SPTV5['Category'].fillna(method='ffill')
df_merge_SPTV5['QLevel']=df_merge_SPTV5['QLevel'].fillna(method='ffill')
df_merge_SPTV5['Detail2']=df_merge_SPTV5['Detail2'].fillna(method='ffill')


# In[355]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV5['Tmpl']=df_merge_SPTV5['Tmpl'].fillna(method='ffill')
df_merge_SPTV5['Super']=df_merge_SPTV5['Super'].fillna(method='ffill')


# In[356]:


df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV5['BBBB']=df_merge_SPTV5['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV5['BBBB']=df_merge_SPTV5['BBBB'].replace(r'nan',np.nan, regex=True)


# In[357]:


df2=df_merge_SPTV5['LastDigit_PV'].groupby(df_merge_SPTV5['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df2,on='LastDigit_PV',how='left')


# In[358]:


df_merge_SPTV5.drop("Detail3_x",axis='columns',inplace=True)

df_merge_SPTV5=df_merge_SPTV5.rename(columns={"Detail3_y":"Detail3"})


# In[359]:


#df_merge_SPTV5.columns


# In[360]:


df3=df_merge_SPTV5['Detail3'].groupby(df_merge_SPTV5['CCCC']).unique().apply(pd.Series)
df3=pd.DataFrame(df3)
df3.reset_index(inplace=True)

df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df3,on='Detail3',how='left')

df_merge_SPTV5=df_merge_SPTV5.rename(columns={"CCCC_y":"CCCC"})


# In[361]:


#df_merge_SPTV5.columns


# In[362]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV5['CCCC']=df_merge_SPTV5['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[363]:




df_merge_SPTV5['DDDD']=df_merge_SPTV5.groupby('Show_name_index').ngroup()

df_merge_SPTV5['DDDD']=df_merge_SPTV5['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[364]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'nan',np.nan, regex=True)


# In[365]:


df4=df_merge_SPTV5['List_Heading'].groupby(df_merge_SPTV5['AAAA']).unique().apply(pd.Series)

df4=pd.DataFrame(df4)


df4.reset_index(inplace=True)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df4,on='List_Heading',how='left')

df_merge_SPTV5.rename(columns={'AAAA_x':'AAAA'},inplace=True)


# In[366]:


df5=df_merge_SPTV5['List_Heading'].groupby(df_merge_SPTV5['BBBB']).unique().apply(pd.Series)

df5=pd.DataFrame(df5)


df5.reset_index(inplace=True)


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df5,on='List_Heading',how='left')

df_merge_SPTV5.rename(columns={'BBBB_x':'BBBB'},inplace=True)


# In[367]:



df_merge_SPTV5['ORD_new'] =df_merge_SPTV5['AAAA']+df_merge_SPTV5['BBBB']+df_merge_SPTV5['CCCC']+df_merge_SPTV5['DDDD']


# In[368]:


df_merge_SPTV5['Show_Name'] = df_merge_SPTV5.apply(lambda x: x['List_Heading']+': '+x['Show_Name'] , axis=1)


# In[369]:


df_merge_SPTV5.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV5_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # SPTV5.1

# In[370]:


Punch_variable=PV['SPTV5.1']['PunchValue']


# In[371]:


Punch_variable


# In[372]:


PV['SPTV5.1']['PunchValue']


# In[373]:


data['SPTV5.1'].shape


# In[374]:


datapv={}
SPTV51=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV5.1'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV51.append(datapv[i])


# In[375]:


SPTV51=pd.concat(SPTV51)


# In[376]:


SPTV51.S2022_Client = SPTV51.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[377]:


df_merge_SPTV51= pd.merge(SPTV51, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[378]:


df_merge_SPTV51['LastDigit_PV']=df_merge_SPTV51['S2022_Client'].str.strip().str[-1]


# In[379]:


#df_merge_SPTV51


# In[380]:


df_merge_SPTV51['ORD']=df_merge_SPTV51['ORD'].astype(str)

df_merge_SPTV51['AAAA']=df_merge_SPTV51["ORD"].str.slice(0,4,1)

df_merge_SPTV51['BBBB']=df_merge_SPTV51["ORD"].str.slice(4,9,1)

df_merge_SPTV51['CCCC']=df_merge_SPTV51["ORD"].str.slice(9,13,1)


# In[381]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV51['Category']=df_merge_SPTV51['Category'].fillna(method='ffill')
df_merge_SPTV51['QLevel']=df_merge_SPTV51['QLevel'].fillna(method='ffill')
df_merge_SPTV51['Detail2']=df_merge_SPTV51['Detail2'].fillna(method='ffill')


# In[382]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV51['Tmpl']=df_merge_SPTV51['Tmpl'].fillna(method='ffill')
df_merge_SPTV51['Super']=df_merge_SPTV51['Super'].fillna(method='ffill')


# In[383]:


#df_merge_SPTV51


# In[384]:


df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV51['BBBB']=df_merge_SPTV51['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_SPTV51['BBBB']=df_merge_SPTV51['BBBB'].replace(r'nan',np.nan, regex=True)


# In[385]:


df2=df_merge_SPTV51['LastDigit_PV'].groupby(df_merge_SPTV51['Detail3']).unique().apply(pd.Series)




# In[386]:


df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[387]:


df_merge_SPTV51=pd.merge(df_merge_SPTV51,df2,on='LastDigit_PV',how='left')


# In[388]:


#df2


# In[389]:


#df_merge_SPTV51


# In[390]:


#df_merge_SPTV51.drop("Detail3_x",axis='columns',inplace=True)
df_merge_SPTV51['Detail3_y']=df_merge_SPTV51['Detail3_y'].replace(np.nan,'UNKNOWN DETAIL',regex=True)

df_merge_SPTV51=df_merge_SPTV51.rename(columns={"Detail3_y":"Detail3"})


# In[391]:


df3=df_merge_SPTV51['Detail3'].groupby(df_merge_SPTV51['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[392]:


df3.dropna(axis=1,inplace=True)


# In[ ]:





# In[393]:


df3['CCCC']=df3['CCCC'].replace(r'^\s*$', np.nan, regex=True)

df3['CCCC']=df3['CCCC'].replace(r'nan',np.nan, regex=True)


# In[394]:


df3.dropna(inplace=True)


# In[395]:


#df3


# In[396]:


df_merge_SPTV51=pd.merge(df_merge_SPTV51,df3,on='Detail3',how='left')


# In[397]:


df_merge_SPTV51.shape


# In[398]:


df_merge_SPTV51['CCCC_y']=df_merge_SPTV51['CCCC_y'].replace(np.nan,'0001', regex=True)


# In[399]:


df_merge_SPTV51=df_merge_SPTV51.rename(columns={"CCCC_y":"CCCC"})


# In[400]:


#df_merge_SPTV51


# In[401]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV51['CCCC']=df_merge_SPTV51['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[402]:



df_merge_SPTV51['DDDD']=df_merge_SPTV51.groupby('Show_name_index').ngroup()

df_merge_SPTV51['DDDD']=df_merge_SPTV51['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[403]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'nan',np.nan, regex=True)


# In[404]:


df4=df_merge_SPTV51['List_Heading'].groupby(df_merge_SPTV51['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV51=pd.merge(df_merge_SPTV51,df4,on='List_Heading',how='left')

df_merge_SPTV51.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[405]:


df5=df_merge_SPTV51['List_Heading'].groupby(df_merge_SPTV51['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV51=pd.merge(df_merge_SPTV51,df5,on='List_Heading',how='left')

df_merge_SPTV51.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[406]:


df_merge_SPTV51['ORD_new'] =df_merge_SPTV51['AAAA']+df_merge_SPTV51['BBBB']+df_merge_SPTV51['CCCC']+df_merge_SPTV51['DDDD']


# In[407]:


df_merge_SPTV51.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV51_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[408]:


#df_merge_SPTV51


# # TV6

# In[409]:


Punch_variable=PV['TV6']['PunchValue']


# In[410]:


datapv={}
TV6=[]
for i in Punch_variable:
   
    datapv[i]=data['TV6'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV6.append(datapv[i])


# In[411]:



TV6=pd.concat(TV6)


# In[412]:


TV6.S2022_Client = TV6.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[413]:



df_merge_TV6= pd.merge(TV6, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[414]:


df_merge_TV6['LastDigit_PV']=df_merge_TV6['S2022_Client'].str.strip().str[-1]


# In[415]:


#df_merge_TV6.head(10)


# In[416]:



df_merge_TV6.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV6_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[417]:


df_merge_TV6['ORD']=df_merge_TV6['ORD'].astype(str)

df_merge_TV6['AAAA']=df_merge_TV6["ORD"].str.slice(0,4,1)

df_merge_TV6['BBBB']=df_merge_TV6["ORD"].str.slice(4,9,1)

df_merge_TV6['CCCC']=df_merge_TV6["ORD"].str.slice(9,13,1)


# In[418]:


df_merge_TV6=df_merge_TV6.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV6['Category']=df_merge_TV6['Category'].fillna(method='ffill')
df_merge_TV6['QLevel']=df_merge_TV6['QLevel'].fillna(method='ffill')
df_merge_TV6['Detail2']=df_merge_TV6['Detail2'].fillna(method='ffill')


# In[419]:


df_merge_TV6=df_merge_TV6.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV6['Tmpl']=df_merge_TV6['Tmpl'].fillna(method='ffill')
df_merge_TV6['Super']=df_merge_TV6['Super'].fillna(method='ffill')


# In[420]:


df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV6['BBBB']=df_merge_TV6['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV6['BBBB']=df_merge_TV6['BBBB'].replace(r'nan',np.nan, regex=True)


# In[421]:



df2=df_merge_TV6['LastDigit_PV'].groupby(df_merge_TV6['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df2,on='LastDigit_PV',how='left')


# In[422]:


#df2


# In[423]:


df_merge_TV6.drop("Detail3_x",axis='columns',inplace=True)

df_merge_TV6=df_merge_TV6.rename(columns={"Detail3_y":"Detail3"})


# In[424]:


df3=df_merge_TV6['Detail3'].groupby(df_merge_TV6['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[425]:


#df3.drop_duplicates(inplace=True)


# In[426]:


df3.drop(0,inplace=True)


# In[427]:


#df3


# In[428]:


df3.dropna(axis=1,inplace=True)


# In[429]:


#df3


# In[430]:



df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df3,on='Detail3',how='left')

df_merge_TV6=df_merge_TV6.rename(columns={"CCCC_y":"CCCC"})


# In[431]:


#df_merge_TV6.LastDigit_PV


# In[432]:


df_merge_TV6=df_merge_TV6.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV6['CCCC']=df_merge_TV6['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[433]:


df_merge_TV6['DDDD']=df_merge_TV6.groupby('Show_name_index').ngroup()

df_merge_TV6['DDDD']=df_merge_TV6['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[434]:


df_merge_TV6=df_merge_TV6.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'nan',np.nan, regex=True)


# In[435]:


df4=df_merge_TV6['List_Heading'].groupby(df_merge_TV6['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df4,on='List_Heading',how='left')

df_merge_TV6.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[436]:


#df4


# In[437]:


#df_merge_TV6


# In[438]:


df5=df_merge_TV6['List_Heading'].groupby(df_merge_TV6['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df5,on='List_Heading',how='left')

df_merge_TV6.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[439]:



df_merge_TV6['ORD_new'] =df_merge_TV6['AAAA']+df_merge_TV6['BBBB']+df_merge_TV6['CCCC']+df_merge_TV6['DDDD']


# In[440]:


df_merge_TV6.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV6_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV5

# In[441]:


Punch_variable=PV['TV5']['PunchValue']


# In[442]:


#Punch_variable


# In[443]:


datapv={}
TV5=[]
for i in Punch_variable:
   
    datapv[i]=data['TV5'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV5.append(datapv[i])


# In[444]:



TV5=pd.concat(TV5)


# In[445]:



TV5.S2022_Client = TV5.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[446]:


df_merge_TV5= pd.merge(TV5, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[447]:


df_merge_TV5['LastDigit_PV']=df_merge_TV5['S2022_Client'].str.strip().str[-1]


# In[448]:


#df_merge_TV5.head(10)


# In[449]:


df_merge_TV5.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV5_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[450]:


df_merge_TV5['ORD']=df_merge_TV5['ORD'].astype(str)

df_merge_TV5['AAAA']=df_merge_TV5["ORD"].str.slice(0,4,1)

df_merge_TV5['BBBB']=df_merge_TV5["ORD"].str.slice(4,9,1)

df_merge_TV5['CCCC']=df_merge_TV5["ORD"].str.slice(9,13,1)


# In[451]:



df_merge_TV5=df_merge_TV5.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV5['Category']=df_merge_TV5['Category'].fillna(method='ffill')
df_merge_TV5['QLevel']=df_merge_TV5['QLevel'].fillna(method='ffill')
df_merge_TV5['Detail2']=df_merge_TV5['Detail2'].fillna(method='ffill')


# In[452]:


df_merge_TV5=df_merge_TV5.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV5['Tmpl']=df_merge_TV5['Tmpl'].fillna(method='ffill')
df_merge_TV5['Super']=df_merge_TV5['Super'].fillna(method='ffill')


# In[453]:


df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV5['BBBB']=df_merge_TV5['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV5['BBBB']=df_merge_TV5['BBBB'].replace(r'nan',np.nan, regex=True)


# In[454]:


df2=df_merge_TV5['LastDigit_PV'].groupby(df_merge_TV5['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV5=pd.merge(df_merge_TV5,df2,on='LastDigit_PV',how='left')


# In[455]:


#df2


# In[456]:



df_merge_TV5.drop("Detail3_x",axis='columns',inplace=True)

df_merge_TV5=df_merge_TV5.rename(columns={"Detail3_y":"Detail3"})


# In[457]:


df3=df_merge_TV5['Detail3'].groupby(df_merge_TV5['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[458]:


#df3


# In[459]:


df3.drop(0,inplace=True)


# In[460]:


df3.dropna(axis=1,inplace=True)


# In[461]:


#df3


# In[462]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV5=pd.merge(df_merge_TV5,df3,on='Detail3',how='left')

df_merge_TV5=df_merge_TV5.rename(columns={"CCCC_y":"CCCC"})


# In[463]:


df_merge_TV5=df_merge_TV5.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV5['CCCC']=df_merge_TV5['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[464]:


df_merge_TV5['DDDD']=df_merge_TV5.groupby('Show_name_index').ngroup()

df_merge_TV5['DDDD']=df_merge_TV5['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[465]:



df_merge_TV5=df_merge_TV5.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'nan',np.nan, regex=True)


# In[466]:


df4=df_merge_TV5['List_Heading'].groupby(df_merge_TV5['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)



# In[467]:


#df4


# In[468]:


#df_merge_TV5=pd.merge(df_merge_TV5,df4,on='List_Heading',how='left')

#df_merge_TV5.rename(columns={'AAAA_y':'AAAA'},inplace=True)
df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].fillna('1237')


# In[469]:


df5=df_merge_TV5['List_Heading'].groupby(df_merge_TV5['BBBB']).unique().apply(pd.Series)

df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[470]:


#df5


# In[471]:


#df_merge_TV5=pd.merge(df_merge_TV5,df5,on='List_Heading',how='left')

#df_merge_TV5.rename(columns={'BBBB_y':'BBBB'},inplace=True)
df_merge_TV5['BBBB']=df_merge_TV5['BBBB'].fillna('00306')


# In[1214]:


s=pd.value_counts(df_merge_TV5['Show_Name'])
#print(s)

s1 = pd.Series({'nunique': len(s), 'unique values': s.index.tolist()})

# above line add the unique values and the no of shows count at the end 
s.append(s1)


# s=pd.value_counts(data['TV5']['Show_Name'])
# 
# s1 = pd.Series({'nunique': len(s), 'unique values': s.index.tolist()})
# s.append(s1)

# In[474]:



df_merge_TV5['ORD_new'] =df_merge_TV5['AAAA']+df_merge_TV5['BBBB']+df_merge_TV5['CCCC']+df_merge_TV5['DDDD']


# In[475]:


df_merge_TV5.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV5_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV1

# # seperate col1 and col2 Punch values

# In[1426]:


pm_col2=pm.query('Columns==2')


# In[1427]:


pm_col2=pd.DataFrame(pm_col2)


# In[1428]:


pm_col1=pm.query('Columns==1')


# In[1429]:


pm_col1=pd.DataFrame(pm_col1)


# In[1430]:


PV1={}
grouped1 = pm_col1.groupby('Clean_Type')
for group1 in grouped1.groups.keys():
    PV1[group1] = grouped1.get_group(group1)


# In[1431]:


PV2={}
grouped2 = pm_col2.groupby('Clean_Type')
for group2 in grouped2.groups.keys():
    PV2[group2] = grouped2.get_group(group2)


# In[ ]:





# In[1432]:


Punch_variable=PV1['TV1']['PunchValue']


# In[1433]:


#PV2['TV1']['PunchValue']


# In[1434]:


datapv={}
TV1=[]
for i in Punch_variable:
   
    datapv[i]=data['TV1'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV1.append(datapv[i])


# In[1435]:


TV1=pd.concat(TV1)


# In[1436]:


TV1.S2022_Client = TV1.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[1437]:


df_merge_TV1= pd.merge(TV1, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[1438]:


df_merge_TV1['LastDigit_PV']=df_merge_TV1['S2022_Client'].str.strip().str[-1]


# In[1439]:


#df_merge_TV1.head(10)


# In[1440]:


df_merge_TV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[1441]:


df_merge_TV1['ORD']=df_merge_TV1['ORD'].astype(str)

df_merge_TV1['AAAA']=df_merge_TV1["ORD"].str.slice(0,4,1)

df_merge_TV1['BBBB']=df_merge_TV1["ORD"].str.slice(4,9,1)

df_merge_TV1['CCCC']=df_merge_TV1["ORD"].str.slice(9,13,1)


# In[1442]:


df_merge_TV1=df_merge_TV1.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV1['Category']=df_merge_TV1['Category'].fillna(method='ffill')
df_merge_TV1['QLevel']=df_merge_TV1['QLevel'].fillna(method='ffill')
df_merge_TV1['Detail2']=df_merge_TV1['Detail2'].fillna(method='ffill')


# In[1443]:


df_merge_TV1=df_merge_TV1.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV1['Tmpl']=df_merge_TV1['Tmpl'].fillna(method='ffill')
df_merge_TV1['Super']=df_merge_TV1['Super'].fillna(method='ffill')


# In[1444]:



df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].replace(r'nan',np.nan, regex=True)


# In[1445]:


#df_merge_TV1


# In[1446]:


#df_merge_TV1=df_merge_TV1.rename(columns={"Detail3_y":"Detail3"})


# In[1447]:


df2_1=df_merge_TV1[['LastDigit_PV','Detail3','List_Heading']]


# df2=df_merge_TV1['LastDigit_PV'].groupby(df_merge_TV1['Detail3']).unique().apply(pd.Series)
# 
# df2=pd.DataFrame(df2)
# 
# type(df2)
# 
# df2.reset_index(inplace=True)
# 
# df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

# In[1448]:



df2_1.drop_duplicates(inplace=True)


# In[1449]:


df2_1.dropna(axis=0,inplace=True)


# In[1450]:


df2_1


# In[1451]:


df_merge_TV1=pd.merge(df_merge_TV1,df2_1,on=['LastDigit_PV','List_Heading'],how='left')


# In[1452]:


df_merge_TV1['Detail3_y'].isna().value_counts()


# In[1453]:


df_merge_TV1=df_merge_TV1.rename(columns={"Detail3_y":"Detail3"})


# In[1454]:



df_merge_TV1 = df_merge_TV1.drop_duplicates(subset='S2022_Client',keep='first')


# In[1456]:


#df_merge_TV1


# In[ ]:





# In[1457]:


#df2


# #Display settings
# pd.set_option('display.max_rows', None)
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width',None)
# pd.set_option('display.colheader_justify', 'center')
# pd.set_option('display.precision', 5)
# pd.set_option('display.max_colwidth', -1)

# In[1458]:



df3_1=df_merge_TV1[['Detail3','List_Heading','CCCC']]


# In[1459]:


df3_1.drop_duplicates(inplace=True)


# In[1460]:


df3_1


# In[1461]:


df3_1['CCCC']=df3_1['CCCC'].replace(r'^\s*$', np.nan, regex=True)

df3_1['CCCC']=df3_1['CCCC'].replace(r'nan',np.nan, regex=True)


# In[1462]:


df3_1.dropna(axis=0,inplace=True)


# In[1463]:


#df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV1=pd.merge(df_merge_TV1,df3_1,on=['List_Heading','Detail3'],how='left')

#df_merge_TV1=pd.merge(df_merge_TV1,df3,on='Detail3',how='left')

#df_merge_TV1['CCCC_y']=df_merge_TV1['CCCC_y'].fillna('0001')


df_merge_TV1=df_merge_TV1.rename(columns={"CCCC_y":"CCCC"})


# In[1464]:


#df_merge_TV1['CCCC'].isna().value_counts()


# In[1465]:


#df_merge_TV1['LastDigit_PV'].unique()


# In[1466]:


#df_merge_TV1['CCCC'].isna().value_counts()


# In[1467]:


df_merge_TV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_det3CCCfill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[1468]:


df_merge_TV1=df_merge_TV1.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')



df_merge_TV1['DDDD']=df_merge_TV1.groupby('Show_name_index').ngroup()

df_merge_TV1['DDDD']=df_merge_TV1['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[1469]:


#df_merge_TV1=df_merge_TV1.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'nan',np.nan, regex=True)

#df4=df_merge_TV1['List_Heading'].groupby(df_merge_TV1['AAAA']).unique().apply(pd.Series)


# In[1470]:



df4_1=df_merge_TV1[['AAAA','Category']]


# In[1471]:


df4_1.drop_duplicates(inplace=True)


# In[1472]:


#df4.reset_index(inplace=True)

df4_1=pd.DataFrame(df4_1)

#df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[1473]:


df4_1


# In[1474]:


df4_1.dropna(axis=0,inplace=True)


# In[1475]:


df_merge_TV1=pd.merge(df_merge_TV1,df4_1,on=['Category'],how='left')

df_merge_TV1.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[1477]:


df_merge_TV1


# In[1478]:


#df5=df_merge_TV1['List_Heading'].groupby(df_merge_TV1['BBBB']).unique().apply(pd.Series)
df5_1=df_merge_TV1[['BBBB','Category']]

df5_1.drop_duplicates(inplace=True)
#df5.reset_index(inplace=True)
df5_1.dropna(axis=0,inplace=True)
df5_1=pd.DataFrame(df5_1)

#df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[1479]:



df5_1


# In[1480]:


df_merge_TV1=pd.merge(df_merge_TV1,df5_1,on=['Category'],how='left')


# In[1481]:


df_merge_TV1.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[1482]:


df_merge_TV1


# In[1483]:


#df_merge_TV1.rename(columns={'BBBB_y':'BBBB'},inplace=True)
#df_merge_TV1=df_merge_TV1.sort_values(['Super','Category'],
               #ascending=[True,True],na_position='last')
#df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)
#df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'nan',np.nan, regex=True)
#df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].fillna(method='ffill')
##df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].replace(r'nan',np.nan, regex=True)
#df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].fillna(method='ffill')

df_merge_TV1['ORD_new'] =df_merge_TV1['AAAA']+df_merge_TV1['BBBB']+df_merge_TV1['CCCC']+df_merge_TV1['DDDD']


# In[1484]:


#df_merge_TV1['Detail3_x']=df_merge_TV1['Detail3_x'].fillna(df_merge_TV1['Detail3'])


# In[1485]:


df_merge_TV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[1486]:


df_merge_TV1


# # TV1 PV2

# In[1487]:


Punch_variable=PV2['TV1']['PunchValue']


# In[1488]:


#Punch_variable


# In[1489]:


data2=data['TV1'].copy()


# In[1490]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[1491]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[1492]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[1493]:


datapv={}
TV1=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV1.append(datapv[i])


# In[1494]:



TV1=pd.concat(TV1)


# In[1495]:


TV1.S2022_Client = TV1.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[1496]:


df_merge_TV1_col2= pd.merge(TV1, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[1497]:


#df_merge_TV1_col2['Detail3'].unique()


# In[1498]:


df_merge_TV1_col2['LastDigit_PV']=df_merge_TV1_col2['S2022_Client'].str.strip().str[-1]


# In[1499]:


df_merge_TV1_col2['ORD']=df_merge_TV1_col2['ORD'].astype(str)

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2["ORD"].str.slice(0,4,1)

df_merge_TV1_col2['BBBB']=df_merge_TV1_col2["ORD"].str.slice(4,9,1)

df_merge_TV1_col2['CCCC']=df_merge_TV1_col2["ORD"].str.slice(9,13,1)


# In[1501]:


#df_merge_TV1_col2


# In[1502]:


df2_1=df_merge_TV1_col2[['LastDigit_PV','Detail3','List_Heading']]


# In[1503]:


df2_1


# In[1504]:


df2_1.drop_duplicates(inplace=True)


# In[1505]:


df2_1.dropna(axis=0,inplace=True)


# In[1506]:


df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV1_col2['Category']=df_merge_TV1_col2['Category'].fillna(method='ffill')
df_merge_TV1_col2['QLevel']=df_merge_TV1_col2['QLevel'].fillna(method='ffill')
df_merge_TV1_col2['Detail2']=df_merge_TV1_col2['Detail2'].fillna(method='ffill')


# In[1379]:


#df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['col2pv','List_Heading','LastDigit_PV','Detail3'], 
               #ascending=[True,True,True,True],na_position='last')
						  

#df_merge_TV2['Super']=df_merge_TV2['Super'].fillna(method='ffill')
#df_merge_TV1_col2['Detail3']=df_merge_TV1_col2['Detail3'].fillna(method='ffill')


# In[1507]:


df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV1_col2['Tmpl']=df_merge_TV1_col2['Tmpl'].fillna(method='ffill')
df_merge_TV1_col2['Super']=df_merge_TV1_col2['Super'].fillna(method='ffill')


# In[1508]:


df_merge_TV1_col2


# In[1509]:


df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV1_col2['BBBB']=df_merge_TV1_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1_col2['BBBB']=df_merge_TV1_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[1510]:


#df2=df_merge_TV1_col2['LastDigit_PV'].groupby(df_merge_TV1_col2['Detail3']).unique().apply(pd.Series)
df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df2_1,on=['LastDigit_PV','List_Heading'],how='left')

#df2=pd.DataFrame(df2)

#type(df2)


# In[1511]:



#df2.reset_index(inplace=True)

#df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

#df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df2,on='LastDigit_PV',how='left')


# In[1512]:


df_merge_TV1_col2=df_merge_TV1_col2.rename(columns={"Detail3_y":"Detail3"})


# In[1513]:


df_merge_TV1_col2.Detail3.isna().value_counts()


# In[1514]:


df_merge_TV1_col2


# In[1515]:


df_merge_TV1_col2 = df_merge_TV1_col2.drop_duplicates(subset='S2022_Client',keep='first')


# In[1516]:


df3_1=df_merge_TV1_col2[['Detail3','List_Heading','CCCC']]


# In[1517]:


df3_1.drop_duplicates(inplace=True)


# In[1518]:


#df3=df_merge_TV1_col2['Detail3'].groupby(df_merge_TV1_col2['CCCC']).unique().apply(pd.Series)
#df3.reset_index(inplace=True)


# In[1519]:


#df3.drop(0,inplace=True)


# In[1520]:



df3_1['CCCC']=df3_1['CCCC'].replace(r'^\s*$', np.nan, regex=True)

df3_1['CCCC']=df3_1['CCCC'].replace(r'nan',np.nan, regex=True)


# In[1521]:


df3_1


# In[1522]:


df3_1.dropna(axis=0,inplace=True)


# In[1523]:


#df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df3_1,on=['Detail3','List_Heading'],how='left')

#df_merge_TV1_col2['CCCC_y']=df_merge_TV1_col2['CCCC_y'].replace(np.nan,'0001', regex=True)

df_merge_TV1_col2=df_merge_TV1_col2.rename(columns={"CCCC_y":"CCCC"})


# In[1524]:


df_merge_TV1_col2.CCCC.isna().value_counts()


# In[1525]:


#df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

#df_merge_TV1_col2['CCCC']=df_merge_TV1_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV1_col2['DDDD']=df_merge_TV1_col2.groupby('Show_name_index').ngroup()

df_merge_TV1_col2['DDDD']=df_merge_TV1_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[1526]:


#df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'nan',np.nan, regex=True)

#df4=df_merge_TV1_col2['List_Heading'].groupby(df_merge_TV1_col2['AAAA']).unique().apply(pd.Series)


# In[1527]:



df4_1=df_merge_TV1_col2[['Category','AAAA']]


# In[1528]:


df4_1.drop_duplicates(inplace=True)


# In[1529]:


df4_1


# In[ ]:





# In[1530]:


df4_1.dropna(axis=0,inplace=True)


# In[1531]:


#df4.reset_index(inplace=True)

df4_1=pd.DataFrame(df4_1)

#df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[1532]:


df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df4_1,on='Category',how='left')

df_merge_TV1_col2.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[1534]:


type(df_merge_TV1_col2)


# In[1535]:


#df5=df_merge_TV1_col2['List_Heading'].groupby(df_merge_TV1_col2['BBBB']).unique().apply(pd.Series)

df5_1=df_merge_TV1_col2[['Category','BBBB']]

#df5.reset_index(inplace=True)

df5_1=pd.DataFrame(df5_1)

#df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[1536]:


df5_1.drop_duplicates(inplace=True)


# In[1539]:


df5_1


# In[1540]:


df5_1.dropna(axis=0,inplace=True)


# In[1541]:


df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df5_1,on='Category',how='left')


# In[1542]:


df_merge_TV1_col2


# In[1543]:


df_merge_TV1_col2.rename(columns={'BBBB_y':'BBBB'},inplace=True)
#df_merge_TV1_col2["AAAA"].fillna(df_merge_TV1_col2["AAAA_x"], inplace=True)
#df_merge_TV1_col2["BBBB"].fillna(df_merge_TV1_col2["BBBB_x"], inplace=True)

df_merge_TV1_col2['ORD_new'] =df_merge_TV1_col2['AAAA']+df_merge_TV1_col2['BBBB']+df_merge_TV1_col2['CCCC']+df_merge_TV1_col2['DDDD']


# In[1544]:



df_merge_TV1_col2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_col2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[1546]:


type(df_merge_TV1_col2)


# In[1547]:


df_merge_TV1=[df_merge_TV1,df_merge_TV1_col2]


# In[1548]:


#df_merge_TV1=pd.DataFrame(df_merge_TV1)


# In[1549]:


df_merge_TV1=pd.concat(df_merge_TV1)


# In[1420]:


#df_merge_TV1.isna()


# In[1421]:


#df_merge_TV1['Detail3_x']=df_merge_TV1['Detail3_x'].fillna(df_merge_TV1['Detail3'])


# In[1422]:


#df_merge_TV1.drop("Detail3",axis='columns',inplace=True)


# In[1423]:


#df_merge_TV1=df_merge_TV1.rename(columns={"Detail3_x":"Detail3"})


# In[1550]:



df_merge_TV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_all.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV2 

# In[574]:


Punch_variable=PV1['TV2']['PunchValue']


# In[575]:


datapv={}
TV2=[]
for i in Punch_variable:
   
    datapv[i]=data['TV2'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV2.append(datapv[i])


# In[576]:


TV2=pd.concat(TV2)


# In[577]:


TV2.S2022_Client = TV2.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[578]:


df_merge_TV2= pd.merge(TV2, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[579]:


df_merge_TV2['LastDigit_PV']=df_merge_TV2['S2022_Client'].str.strip().str[-1]


# In[580]:


#df_merge_TV2


# In[581]:


#df_merge_TV2


# In[582]:



df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_b4fill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[583]:


df_merge_TV2_net=df_merge_TV2[(df_merge_TV2['Show_Name'].str.startswith('CBS Dream Team net') | df_merge_TV2['Show_Name'].str.startswith('Litton Weekend Adventure net'))]
#s.str.startswith('a', na=False)


# In[584]:



df_merge_TV2_net.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_net.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[585]:


df_merge_TV2_net=df_merge_TV2_net.sort_values(['List_Heading','ORD'], 
               ascending=[True,False],na_position='last')
df_merge_TV2_net['Category']=df_merge_TV2_net['Category'].fillna(method='ffill')
df_merge_TV2_net['QLevel']=df_merge_TV2_net['QLevel'].fillna(method='ffill')
df_merge_TV2_net['Tmpl']=df_merge_TV2_net['Tmpl'].fillna(method='ffill')
df_merge_TV2_net['Super']=df_merge_TV2_net['Super'].fillna(method='ffill')


# In[586]:


#df_merge_TV2_net


# In[587]:


df_merge_TV2_net
df_merge_TV2_net['ORD']=df_merge_TV2_net['ORD'].astype(str)

df_merge_TV2_net['AAAA']=df_merge_TV2_net["ORD"].str.slice(0,4,1)

df_merge_TV2_net['BBBB']=df_merge_TV2_net["ORD"].str.slice(4,9,1)

df_merge_TV2_net['CCCC']=df_merge_TV2_net["ORD"].str.slice(9,13,1)


# In[588]:



df_merge_TV2_net['AAAA']=df_merge_TV2_net['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2_net['AAAA']=df_merge_TV2_net['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV2_net['BBBB']=df_merge_TV2_net['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2_net['BBBB']=df_merge_TV2_net['BBBB'].replace(r'nan',np.nan, regex=True)
df_merge_TV2_net['CCCC']=df_merge_TV2_net['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2_net['CCCC']=df_merge_TV2_net['CCCC'].replace(r'nan',np.nan, regex=True)


# In[589]:


df_merge_TV2_net=df_merge_TV2_net.sort_values(['List_Heading','ORD'], 
               ascending=[True,True],na_position='last')
df_merge_TV2_net['AAAA']=df_merge_TV2_net['AAAA'].fillna(method='ffill')
df_merge_TV2_net['BBBB']=df_merge_TV2_net['BBBB'].fillna(method='ffill')
df_merge_TV2_net['CCCC']=df_merge_TV2_net['CCCC'].fillna(method='ffill')


# In[590]:


df_merge_TV2_net['DDDD']=df_merge_TV2_net.groupby('Show_name_index').ngroup()

df_merge_TV2_net['DDDD']=df_merge_TV2_net['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[591]:


df_merge_TV2_net.head(5)


# In[592]:


df_merge_TV2_net['ORD_new'] =df_merge_TV2_net['AAAA']+df_merge_TV2_net['BBBB']+df_merge_TV2_net['CCCC']+df_merge_TV2_net['DDDD']


# In[593]:


df_merge_TV2 = pd.merge(df_merge_TV2,df_merge_TV2_net, how = 'outer',on = 'S2022_Client',indicator=True).loc[lambda x : x['_merge']=='left_only']


# In[594]:


df_merge_TV2.columns


# In[595]:


df_merge_TV2=df_merge_TV2[['Clean_Type_x', 'S2022_Client', 'Section_Heading_x', 'List_Heading_x', 'DP_Status_x', 'Show_Name_x', 'Show_name_index_x', 'Initial_Wave_x', 'F2021_Client_x', 'compare_x', 'col2pv_x', 'SID_x', 'SDID_x', 'Category_x', 'QID_x', 'Super_x', 'Tmpl_x', 'Time Period_x', 'Detail1_x', 'Detail2_x', 'Detail3_x', 'Detail4_x', 'UCode_x', 'Definition_x', 'CCP_x', 'ORD_x', 'Full_Label_x', 'QLevel_x', 'Wave_x', 'StatisticID_x', 'CatSynID_x', 'NoteID_x', 'QuestionID_x', 'QUESTID_x', 'AnswerID_x', 'StudyEntryID_x', 'VersionID_x', 'StudyAnswerID_x', 'Status_x', 'statusid_x', 'LastDigit_PV_x']]


# In[596]:


df_merge_TV2=df_merge_TV2.rename(columns={'Clean_Type_x':'Clean_Type', 
                                          'Section_Heading_x':'Section_Heading', 
                                          'List_Heading_x':'List_Heading', 
                                          'DP_Status_x':'DP_Status',
                                          'Show_Name_x':'Show_Name',
                                          'Show_name_index_x':'Show_name_index', 
                                          'Initial_Wave_x':'Initial_Wave',
                                          'F2021_Client_x':'F2021_Client',
                                          'compare_x':'compare', 
                                          'col2pv_x':'col2pv',
                                          'SID_x':'SID', 
                                          'SDID_x':'SDID',
                                          'Category_x':'Category',
                                          'QID_x':'QID',
                                          'Super_x':'Super', 
                                          'Tmpl_x':'Tmpl', 
                                          'Time Period_x':'Time Period', 
                                          'Detail1_x':'Detail1', 
                                          'Detail2_x':'Detail2', 
                                          'Detail3_x':'Detail3',
                                          'Detail4_x':'Detail4',
                                          'UCode_x':'UCode', 
                                          'Definition_x':'Definition',
                                          'CCP_x':'CCP', 
                                          'ORD_x':'ORD', 
                                          'Full_Label_x':'Full_Label',
                                          'QLevel_x':'QLevel',
                                          'Wave_x':'Wave',
                                          'StatisticID_x':'StatisticID',
                                          'CatSynID_x':'CatSynID',
                                          'NoteID_x':'NoteID',
                                          'QuestionID_x':'QuestionID',
                                          'QUESTID_x':'QUESTID',
                                          'AnswerID_x':'AnswerID',
                                          'StudyEntryID_x':'StudyEntryID',
                                          'VersionID_x':'VersionID',
                                          'StudyAnswerID_x':'StudyAnswerID',
                                          'Status_x':'Status',
                                          'statusid_x':'statusid',
                                          'LastDigit_PV_x':'LastDigit_PV'})


# In[597]:



df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_nonet.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[598]:


#df_merge_TV2
df2_cat=df_merge_TV2[['Category','List_Heading']]


# In[599]:


df2_cat.drop_duplicates(inplace=True)


# In[600]:


df2_cat.dropna(axis=0,inplace=True)


# In[601]:


df2_cat


# In[602]:


df_merge_TV2=pd.merge(df_merge_TV2,df2_cat,on=['List_Heading'],how='left')


# In[603]:


#df_merge_TV2.query('Show_Name=="United States of AI"')


# In[604]:


df_merge_TV2=df_merge_TV2.rename(columns={"Category_y":"Category"})


# In[605]:



df_merge_TV2=df_merge_TV2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
#df_merge_TV2['Category']=df_merge_TV2['Category'].fillna(method='ffill')
df_merge_TV2['QLevel']=df_merge_TV2['QLevel'].fillna(method='ffill')
#df_merge_TV2['Detail2']=df_merge_TV2['Detail2'].fillna(method='ffill')


# In[606]:


g=df_merge_TV2.groupby('List_Heading')


# In[607]:


i=0
n=0
df_merge_TV2_LH={}
for List_Heading, g_df in g:
    #print (List_Heading)
    df_merge_TV2_LH[i]=pd.DataFrame(g_df)
    i=i+1
    n=n+1


# In[608]:


Listheading=df_merge_TV2['List_Heading'].unique()


# In[609]:


n=0
for values in Listheading:
    df_merge_TV2_LH[n]=df_merge_TV2_LH[n].sort_values(['LastDigit_PV'], 
               ascending=[True
                        ])
    df_merge_TV2_LH[n]['Detail2']=df_merge_TV2_LH[n]['Detail2'].fillna(method='ffill')
    n=n+1
#for each values List heading FFill the Detail2 values in each  DF 


# In[610]:


df_merge_TV2_Frames=pd.DataFrame()
df_merge_TV2_Frames = df_merge_TV2_Frames.append([df_merge_TV2_LH[i] for i in range(n)])


# In[611]:


df_merge_TV2_Frames.shape


# In[612]:


df_merge_TV2=df_merge_TV2_Frames.copy()


# In[613]:


df_merge_TV2['ORD']=df_merge_TV2['ORD'].astype(str)

df_merge_TV2['AAAA']=df_merge_TV2["ORD"].str.slice(0,4,1)

df_merge_TV2['BBBB']=df_merge_TV2["ORD"].str.slice(4,9,1)

df_merge_TV2['CCCC']=df_merge_TV2["ORD"].str.slice(9,13,1)


# In[614]:


df_merge_TV2=df_merge_TV2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV2['Tmpl']=df_merge_TV2['Tmpl'].fillna(method='ffill')
df_merge_TV2['Super']=df_merge_TV2['Super'].fillna(method='ffill')


# In[615]:


df_merge_TV2


# In[616]:



df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV2['BBBB']=df_merge_TV2['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2['BBBB']=df_merge_TV2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[617]:


df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_1", sep='\t', index=False,header=True,encoding='cp1252')


# In[618]:


df2=df_merge_TV2['LastDigit_PV'].groupby(df_merge_TV2['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)


# In[619]:


df2


# In[620]:


df2.rename(columns={0:"LastDigit_PV"},inplace=True)


# In[621]:


df2_1=df_merge_TV2[['LastDigit_PV','Detail3','List_Heading']]


# In[622]:


#df2_1=df2_1.dropna(axis=0,inplace=True)


# In[623]:


#df2_1


# In[624]:


df2_1.dropna(inplace=True)


# In[625]:


df2_1.drop_duplicates().reset_index(drop=True)


# In[626]:


#df2_1.drop_duplicates().reset_index(drop=True)


# In[627]:


df_merge_TV2_copy=df_merge_TV2.copy()


# In[628]:


df_merge_TV2_copy_1=pd.merge(df_merge_TV2_copy,df2_1,on=['LastDigit_PV','List_Heading'],how='left')


# In[629]:


#df_merge_TV2_copy_1.head(100)


# In[630]:


#df2


# In[631]:





#df_merge_TV2=pd.merge(df_merge_TV2,df2,on='LastDigit_PV',how='left')


# In[632]:


df_merge_TV2=df_merge_TV2_copy_1.copy()


# In[633]:


#df_merge_TV2


# In[634]:


#df_merge_TV2.drop("Detail3_x",axis='columns',inplace=True)

df_merge_TV2=df_merge_TV2.rename(columns={"Detail3_y":"Detail3"})


# In[635]:


df_merge_TV2 = df_merge_TV2.drop_duplicates(subset='S2022_Client',keep='first')


# In[636]:


#df_merge_TV2_copy_1
#df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_copy_1.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[637]:


df3_1=df_merge_TV2[['Detail3','List_Heading','CCCC']]


# In[638]:


df3_1=df3_1.drop_duplicates().reset_index(drop=True)


# In[639]:


df3_1['CCCC']=df3_1['CCCC'].replace(r'^\s*$', np.nan, regex=True)

df3_1['CCCC']=df3_1['CCCC'].replace(r'nan',np.nan, regex=True)


# In[640]:


#df3_1


# In[641]:


df3_1.dropna(axis=0,inplace=True)


# In[642]:


df3_1


# In[643]:


df_merge_TV2=pd.merge(df_merge_TV2,df3_1,on=['Detail3','List_Heading'],how='left')


# In[644]:


#df3=df_merge_TV2['Detail3'].groupby(df_merge_TV2['CCCC']).unique().apply(pd.Series)
#df3.reset_index(inplace=True)

#df3.dropna()
#df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)



# In[645]:


#df_merge_TV2


# In[646]:


#duplicateRows = df_merge_TV2[df_merge_TV2.duplicated(['S2022_Client'])]


# In[647]:


#duplicateRows.count()


# In[648]:


#df3.drop(0,inplace=True)


# In[649]:


#df3.dropna(axis=1,inplace=True)


# In[650]:


#df3


# In[651]:


#df_merge_TV2=pd.merge(df_merge_TV2,df3,on='Detail3',how='left')

df_merge_TV2['CCCC_y']=df_merge_TV2['CCCC_y'].fillna(df_merge_TV2['CCCC_x'])

df_merge_TV2=df_merge_TV2.rename(columns={"CCCC_y":"CCCC"})


# In[652]:


df_merge_TV2=df_merge_TV2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV2['CCCC']=df_merge_TV2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2['DDDD']=df_merge_TV2.groupby('Show_name_index').ngroup()

df_merge_TV2['DDDD']=df_merge_TV2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))

df_merge_TV2=df_merge_TV2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[653]:


df4_1=df_merge_TV2[['List_Heading','AAAA']]


# In[654]:


df4_1=df4_1.drop_duplicates()


# In[655]:


df4_1.dropna(axis=0,inplace=True)


# In[656]:


df4_1


# df4=df_merge_TV2['List_Heading'].groupby(df_merge_TV2['AAAA']).unique().apply(pd.Series)
# 
# df4.reset_index(inplace=True)
# 
# df4=pd.DataFrame(df4)
# 
# df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)
# 
# 

# df4

# df4list=df4.values.tolist()

# print(df4list)

# df4.dropna(axis=1,inplace=True)

# In[657]:


df_merge_TV2=pd.merge(df_merge_TV2,df4_1,on=['List_Heading'],how='left')


# In[658]:


#df_merge_TV2=pd.merge(df_merge_TV2,df4,on='List_Heading',how='left')


# In[659]:


df_merge_TV2.columns


# In[660]:


#df_merge_TV2['AAAA_y'].isna().value_counts()


# In[661]:


df_merge_TV2.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# df5=df_merge_TV2['List_Heading'].groupby(df_merge_TV2['BBBB']).unique().apply(pd.Series)
# 
# 
# 
# df5.reset_index(inplace=True)
# 
# df5=pd.DataFrame(df5)
# 
# df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)
# 

# In[662]:


df5_1=df_merge_TV2[['List_Heading','BBBB']]


# In[663]:


#df5.dropna(axis=1,inplace=True)
df5_1.drop_duplicates(inplace=True)


# In[664]:


df5_1.dropna(axis=0,inplace=True)


# In[665]:


df_merge_TV2=pd.merge(df_merge_TV2,df5_1,on=['List_Heading'],how='left')



# In[666]:


df_merge_TV2['BBBB_y'].isna().value_counts()


# In[667]:


#df_merge_TV2['BBBB_y']=df_merge_TV2['BBBB_y'].fillna(df_merge_TV2['BBBB_x'])


# In[668]:


df_merge_TV2.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[669]:


#df_merge_TV2 = df_merge_TV2.drop_duplicates(subset='S2022_Client',keep='first')


# In[670]:


df_merge_TV2['ORD_new'] =df_merge_TV2['AAAA']+df_merge_TV2['BBBB']+df_merge_TV2['CCCC']+df_merge_TV2['DDDD']


# In[671]:


df_merge_TV2=[df_merge_TV2,df_merge_TV2_net]


# In[672]:


df_merge_TV2=pd.concat(df_merge_TV2)


# In[673]:



df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV2 col2

# In[674]:



Punch_variable=PV2['TV2']['PunchValue']
#Punch_variable=Punch_variable.lower()
data2=data['TV2'].copy()


# In[675]:


#Punch_variable


# In[676]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[677]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[678]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[679]:


datapv={}
TV2=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV2.append(datapv[i])


# In[680]:


TV2=pd.concat(TV2)


# In[681]:



TV2.S2022_Client = TV2.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[682]:


df_merge_TV2_col2= pd.merge(TV2, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[683]:


df_merge_TV2_col2['Tmpl'].unique()


# In[684]:


df_merge_TV2_col2['LastDigit_PV']=df_merge_TV2_col2['S2022_Client'].str.strip().str[-1]


# In[685]:


df_merge_TV2_col2['ORD']=df_merge_TV2_col2['ORD'].astype(str)

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2["ORD"].str.slice(0,4,1)

df_merge_TV2_col2['BBBB']=df_merge_TV2_col2["ORD"].str.slice(4,9,1)

df_merge_TV2_col2['CCCC']=df_merge_TV2_col2["ORD"].str.slice(9,13,1)


# In[686]:


df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV2_col2['Category']=df_merge_TV2_col2['Category'].fillna(method='ffill')
df_merge_TV2_col2['QLevel']=df_merge_TV2_col2['QLevel'].fillna(method='ffill')
#df_merge_TV2_col2['Detail2']=df_merge_TV2_col2['Detail2'].fillna(method='ffill')


# In[687]:



Listheading=df_merge_TV2_col2['List_Heading'].unique()


# In[688]:


g=df_merge_TV2_col2.groupby('List_Heading')


# In[689]:


i=0
n=0
df_merge_TV2_LH={}
for List_Heading, g_df in g:
    #print (List_Heading)
    df_merge_TV2_LH[i]=pd.DataFrame(g_df)
    i=i+1
    n=n+1


# In[690]:


n=0

for values in Listheading:
    df_merge_TV2_LH[n]=df_merge_TV2_LH[n].sort_values(['LastDigit_PV'], 
               ascending=[True
                        ])
    df_merge_TV2_LH[n]['Detail2']=df_merge_TV2_LH[n]['Detail2'].fillna(method='ffill')
    n=n+1
#for each values List heading FFill the Detail2 values in each  DF 


# In[691]:



df_merge_TV2_Frames=pd.DataFrame()
df_merge_TV2_Frames = df_merge_TV2_Frames.append([df_merge_TV2_LH[i] for i in range(n)])


# In[692]:



df_merge_TV2_col2=df_merge_TV2_Frames.copy()


# In[693]:


df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV2_col2['Tmpl']=df_merge_TV2_col2['Tmpl'].fillna(method='ffill')
df_merge_TV2_col2['Super']=df_merge_TV2_col2['Super'].fillna(method='ffill')


# In[694]:


df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV2_col2['BBBB']=df_merge_TV2_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2_col2['BBBB']=df_merge_TV2_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[695]:


df2=df_merge_TV2_col2['LastDigit_PV'].groupby(df_merge_TV2_col2['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

#type(df2)


# In[696]:


df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df2,on='LastDigit_PV',how='left')

df_merge_TV2_col2=df_merge_TV2_col2.rename(columns={"Detail3_y":"Detail3"})


# In[697]:


df_merge_TV2_col2 = df_merge_TV2_col2.drop_duplicates(subset='S2022_Client',keep='first')

df3=df_merge_TV2_col2['Detail3'].groupby(df_merge_TV2_col2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[698]:


df3

df3.dropna(axis=1,inplace=True)

df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)



# In[699]:


#df3


# In[700]:


df3.drop(0,inplace=True)


# In[701]:


df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df3,on='Detail3',how='left')


# In[702]:


df_merge_TV2_col2=df_merge_TV2_col2.rename(columns={"CCCC_y":"CCCC"})

df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV2_col2['CCCC']=df_merge_TV2_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2_col2['DDDD']=df_merge_TV2_col2.groupby('Show_name_index').ngroup()

df_merge_TV2_col2['DDDD']=df_merge_TV2_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[703]:


df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[704]:


df4=df_merge_TV2_col2['List_Heading'].groupby(df_merge_TV2_col2['AAAA']).unique().apply(pd.Series)

df4.dropna(axis=1,inplace=True)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)


# In[705]:


#df4


# In[706]:


df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df4,on='List_Heading',how='left')


# In[707]:



df_merge_TV2_col2.rename(columns={'AAAA_y':'AAAA'},inplace=True)

df5=df_merge_TV2_col2['List_Heading'].groupby(df_merge_TV2_col2['BBBB']).unique().apply(pd.Series)


# In[708]:


df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df5.dropna(axis=1,inplace=True)

df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df5,on='List_Heading',how='left')


# In[709]:


df_merge_TV2_col2.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[710]:


df_merge_TV2_col2["AAAA"].fillna(df_merge_TV2_col2["AAAA_x"], inplace=True)
df_merge_TV2_col2["BBBB"].fillna(df_merge_TV2_col2["BBBB_x"], inplace=True)


# In[711]:


df_merge_TV2_col2['ORD_new'] =df_merge_TV2_col2['AAAA']+df_merge_TV2_col2['BBBB']+df_merge_TV2_col2['CCCC']+df_merge_TV2_col2['DDDD']


# In[712]:


df_merge_TV2=[df_merge_TV2,df_merge_TV2_col2]


# In[713]:


type(df_merge_TV2)


# In[714]:


#df_merge_TV2=pd.DataFrame(df_merge_TV2)


# In[715]:


df_merge_TV2=pd.concat(df_merge_TV2)


# In[716]:


df_merge_TV2['Detail3_x']=df_merge_TV2['Detail3_x'].fillna(df_merge_TV2['Detail3'])


# In[717]:


df_merge_TV2.drop("Detail3",axis='columns',inplace=True)


# In[718]:


df_merge_TV2=df_merge_TV2.rename(columns={"Detail3_x":"Detail3"})


# In[719]:


df_merge_TV2 = df_merge_TV2.drop_duplicates(subset='S2022_Client',keep='first')


# In[720]:


df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV3

# In[721]:


Punch_variable=PV1['TV3']['PunchValue']


# In[722]:


datapv={}
TV3=[]
for i in Punch_variable:
   
    datapv[i]=data['TV3'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV3.append(datapv[i])


# In[723]:



TV3=pd.concat(TV3)


# In[724]:


TV3.S2022_Client = TV3.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[725]:


df_merge_TV3= pd.merge(TV3, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')

df_merge_TV3['LastDigit_PV']=df_merge_TV3['S2022_Client'].str.strip().str[-1]


# In[726]:



df_merge_TV3.head(10)


# In[727]:


df_merge_TV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[728]:


df_merge_TV3['ORD']=df_merge_TV3['ORD'].astype(str)

df_merge_TV3['AAAA']=df_merge_TV3["ORD"].str.slice(0,4,1)

df_merge_TV3['BBBB']=df_merge_TV3["ORD"].str.slice(4,9,1)

df_merge_TV3['CCCC']=df_merge_TV3["ORD"].str.slice(9,13,1)


# In[729]:



df_merge_TV3=df_merge_TV3.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV3['Category']=df_merge_TV3['Category'].fillna(method='ffill')
df_merge_TV3['QLevel']=df_merge_TV3['QLevel'].fillna(method='ffill')
df_merge_TV3['Detail2']=df_merge_TV3['Detail2'].fillna(method='ffill')


# In[730]:


df_merge_TV3=df_merge_TV3.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV3['Tmpl']=df_merge_TV3['Tmpl'].fillna(method='ffill')
df_merge_TV3['Super']=df_merge_TV3['Super'].fillna(method='ffill')


# In[731]:


df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].replace(r'nan',np.nan, regex=True)


# In[732]:



df2=df_merge_TV3['LastDigit_PV'].groupby(df_merge_TV3['Detail3']).unique().apply(pd.Series)


# In[733]:


df2=pd.DataFrame(df2)


# In[734]:


df2


# In[735]:


df2.reset_index(inplace=True)


# In[736]:



df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[737]:


df_merge_TV3=pd.merge(df_merge_TV3,df2,on='LastDigit_PV',how='left')


# In[738]:


df_merge_TV3=df_merge_TV3.rename(columns={"Detail3_y":"Detail3"})


# In[739]:


df3=df_merge_TV3['Detail3'].groupby(df_merge_TV3['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[740]:



#df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[741]:


df3


# In[742]:


df_merge_TV3=pd.merge(df_merge_TV3,df3,on='Detail3',how='left')


# In[743]:


df_merge_TV3=df_merge_TV3.rename(columns={"CCCC_y":"CCCC"})


# In[744]:


df_merge_TV3=df_merge_TV3.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV3['CCCC']=df_merge_TV3['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[745]:



df_merge_TV3['DDDD']=df_merge_TV3.groupby('Show_name_index').ngroup()

df_merge_TV3['DDDD']=df_merge_TV3['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[746]:


df_merge_TV3=df_merge_TV3.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'nan',np.nan, regex=True)


# In[747]:


df4=df_merge_TV3['List_Heading'].groupby(df_merge_TV3['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)


# In[748]:


df4


# In[749]:


df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].fillna('1246')


# In[750]:


df5=df_merge_TV3['List_Heading'].groupby(df_merge_TV3['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)


# In[751]:


df5


# In[752]:


df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].fillna('00315')


# In[ ]:





# In[753]:


df_merge_TV3['ORD_new'] =df_merge_TV3['AAAA']+df_merge_TV3['BBBB']+df_merge_TV3['CCCC']+df_merge_TV3['DDDD']


# In[754]:


df_merge_TV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV3 Col2

# In[755]:


Punch_variable=PV2['TV3']['PunchValue']
data2=data['TV3'].copy()


# In[756]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[757]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[758]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[759]:


datapv={}
TV3=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV3.append(datapv[i])


# In[760]:


TV3=pd.concat(TV3)


# In[761]:


TV3.S2022_Client = TV3.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[762]:


df_merge_TV3_col2= pd.merge(TV3, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[763]:



df_merge_TV3_col2['Detail3'].unique()


# In[764]:


df_merge_TV3_col2['LastDigit_PV']=df_merge_TV3_col2['S2022_Client'].str.strip().str[-1]


# In[765]:


df_merge_TV3_col2['ORD']=df_merge_TV3_col2['ORD'].astype(str)


# In[766]:



df_merge_TV3_col2['AAAA']=df_merge_TV3_col2["ORD"].str.slice(0,4,1)

df_merge_TV3_col2['BBBB']=df_merge_TV3_col2["ORD"].str.slice(4,9,1)

df_merge_TV3_col2['CCCC']=df_merge_TV3_col2["ORD"].str.slice(9,13,1)


# In[767]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV3_col2['Category']=df_merge_TV3_col2['Category'].fillna(method='ffill')
df_merge_TV3_col2['QLevel']=df_merge_TV3_col2['QLevel'].fillna(method='ffill')
df_merge_TV3_col2['Detail2']=df_merge_TV3_col2['Detail2'].fillna(method='ffill')


# In[768]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV3_col2['Tmpl']=df_merge_TV3_col2['Tmpl'].fillna(method='ffill')
df_merge_TV3_col2['Super']=df_merge_TV3_col2['Super'].fillna(method='ffill')


# In[769]:


df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[770]:


df_merge_TV3_col2['BBBB']=df_merge_TV3_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3_col2['BBBB']=df_merge_TV3_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[771]:


df2=df_merge_TV3_col2['LastDigit_PV'].groupby(df_merge_TV3_col2['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)


# In[772]:


df2


# In[773]:



df2.reset_index(inplace=True)


# In[774]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV3_col2=pd.merge(df_merge_TV3_col2,df2,on='LastDigit_PV',how='left')


# In[775]:



df_merge_TV3_col2=df_merge_TV3_col2.rename(columns={"Detail3_y":"Detail3"})


# In[776]:


#df_merge_TV3_col2 = df_merge_TV3_col2.drop_duplicates(subset='S2022_Client',keep='first')


# In[777]:


df3=df_merge_TV3_col2['Detail3'].groupby(df_merge_TV3_col2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[778]:


df3


# In[779]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[780]:



df_merge_TV3_col2=pd.merge(df_merge_TV3_col2,df3,on='Detail3',how='left')


# In[781]:


df_merge_TV3_col2=df_merge_TV3_col2.rename(columns={"CCCC_y":"CCCC"})


# In[782]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[783]:


df_merge_TV3_col2['CCCC']=df_merge_TV3_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[784]:


df_merge_TV3_col2['DDDD']=df_merge_TV3_col2.groupby('Show_name_index').ngroup()

df_merge_TV3_col2['DDDD']=df_merge_TV3_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[785]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[786]:


df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[787]:


df4=df_merge_TV3_col2['List_Heading'].groupby(df_merge_TV3_col2['AAAA']).unique().apply(pd.Series)


# In[788]:


df4


# In[789]:


df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].fillna('1247')


# In[790]:


df5=df_merge_TV3_col2['List_Heading'].groupby(df_merge_TV3_col2['BBBB']).unique().apply(pd.Series)


# In[791]:


df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].fillna('00316')


# In[792]:


df_merge_TV3_col2['ORD_new'] =df_merge_TV3_col2['AAAA']+df_merge_TV3_col2['BBBB']+df_merge_TV3_col2['CCCC']+df_merge_TV3_col2['DDDD']


# In[793]:


df_merge_TV3_col2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3_col2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[794]:


df_merge_TV3=[df_merge_TV3,df_merge_TV3_col2]

df_merge_TV3=pd.concat(df_merge_TV3)


# In[795]:


df_merge_TV3['Detail3_x']=df_merge_TV3['Detail3_x'].fillna(df_merge_TV3['Detail3'])


# In[796]:


df_merge_TV3.drop("Detail3",axis='columns',inplace=True)


# In[797]:


df_merge_TV3=df_merge_TV3.rename(columns={"Detail3_x":"Detail3"})


# In[798]:


df_merge_TV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[799]:


df_merge_TV3.head(100)


# # TV4 

# In[800]:


Punch_variable=PV1['TV4']['PunchValue']


# In[801]:


datapv={}
TV4=[]
for i in Punch_variable:
   
    datapv[i]=data['TV4'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV4.append(datapv[i])


# In[802]:



TV4=pd.concat(TV4)


# In[803]:



TV4.S2022_Client = TV4.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[804]:



df_merge_TV4= pd.merge(TV4, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[805]:


df_merge_TV4['LastDigit_PV']=df_merge_TV4['S2022_Client'].str.strip().str[-1]


# In[806]:


df_merge_TV4.head(10)


# In[807]:


df_merge_TV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[808]:


df_merge_TV4['ORD']=df_merge_TV4['ORD'].astype(str)

df_merge_TV4['AAAA']=df_merge_TV4["ORD"].str.slice(0,4,1)

df_merge_TV4['BBBB']=df_merge_TV4["ORD"].str.slice(4,9,1)

df_merge_TV4['CCCC']=df_merge_TV4["ORD"].str.slice(9,13,1)


# In[809]:


df_merge_TV4=df_merge_TV4.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV4['Category']=df_merge_TV4['Category'].fillna(method='ffill')
df_merge_TV4['QLevel']=df_merge_TV4['QLevel'].fillna(method='ffill')
df_merge_TV4['Detail2']=df_merge_TV4['Detail2'].fillna(method='ffill')


# In[810]:


df_merge_TV4=df_merge_TV4.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV4['Tmpl']=df_merge_TV4['Tmpl'].fillna(method='ffill')
df_merge_TV4['Super']=df_merge_TV4['Super'].fillna(method='ffill')


# In[811]:


df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV4['BBBB']=df_merge_TV4['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV4['BBBB']=df_merge_TV4['BBBB'].replace(r'nan',np.nan, regex=True)


# In[812]:


df2=df_merge_TV4['LastDigit_PV'].groupby(df_merge_TV4['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)


# In[813]:


df2


# In[814]:


df2.reset_index(inplace=True)


# In[815]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[816]:


df_merge_TV4=pd.merge(df_merge_TV4,df2,on='LastDigit_PV',how='left')


# In[817]:



df_merge_TV4=df_merge_TV4.rename(columns={"Detail3_y":"Detail3"})


# In[818]:


df3=df_merge_TV4['Detail3'].groupby(df_merge_TV4['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[819]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[820]:


#df3


# In[821]:


#df3.dropna(axis=1,inplace=True)


# In[822]:



df_merge_TV4=pd.merge(df_merge_TV4,df3,on='Detail3',how='left')

df_merge_TV4=df_merge_TV4.rename(columns={"CCCC_y":"CCCC"})


# In[823]:


df_merge_TV4=df_merge_TV4.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV4['CCCC']=df_merge_TV4['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[824]:


df_merge_TV4['DDDD']=df_merge_TV4.groupby('Show_name_index').ngroup()

df_merge_TV4['DDDD']=df_merge_TV4['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[825]:


df_merge_TV4=df_merge_TV4.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'nan',np.nan, regex=True)


# In[826]:


df4=df_merge_TV4['List_Heading'].groupby(df_merge_TV4['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)


# In[827]:


#df4


# In[828]:


df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].fillna('1235')


# In[829]:


df5=df_merge_TV4['List_Heading'].groupby(df_merge_TV4['BBBB']).unique().apply(pd.Series)


# In[830]:


#df5


# In[831]:


df_merge_TV4['BBBB']=df_merge_TV4['BBBB'].fillna('00304')


# In[832]:


df_merge_TV4['ORD_new'] =df_merge_TV4['AAAA']+df_merge_TV4['BBBB']+df_merge_TV4['CCCC']+df_merge_TV4['DDDD']


# In[833]:


df_merge_TV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV4 col2

# In[834]:


Punch_variable=PV2['TV4']['PunchValue']
data2=data['TV4'].copy()


# In[835]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[836]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[837]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[838]:


datapv={}
TV4=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV4.append(datapv[i])


# In[839]:


TV4=pd.concat(TV4)


# In[840]:


TV4.S2022_Client = TV4.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[841]:


df_merge_TV4_col2= pd.merge(TV4, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[842]:


#df_merge_TV4_col2['Detail3'].unique()


# In[843]:


df_merge_TV4_col2['LastDigit_PV']=df_merge_TV4_col2['S2022_Client'].str.strip().str[-1]


# In[844]:


df_merge_TV4_col2['ORD']=df_merge_TV4_col2['ORD'].astype(str)

df_merge_TV4_col2['AAAA']=df_merge_TV4_col2["ORD"].str.slice(0,4,1)

df_merge_TV4_col2['BBBB']=df_merge_TV4_col2["ORD"].str.slice(4,9,1)

df_merge_TV4_col2['CCCC']=df_merge_TV4_col2["ORD"].str.slice(9,13,1)


# In[845]:


df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV4_col2['Category']=df_merge_TV4_col2['Category'].fillna(method='ffill')
df_merge_TV4_col2['QLevel']=df_merge_TV4_col2['QLevel'].fillna(method='ffill')
df_merge_TV4_col2['Detail2']=df_merge_TV4_col2['Detail2'].fillna(method='ffill')


# In[846]:



df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV4_col2['Tmpl']=df_merge_TV4_col2['Tmpl'].fillna(method='ffill')
df_merge_TV4_col2['Super']=df_merge_TV4_col2['Super'].fillna(method='ffill')


# In[847]:


df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV4_col2['BBBB']=df_merge_TV4_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4_col2['BBBB']=df_merge_TV4_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[848]:


df2=df_merge_TV4_col2['LastDigit_PV'].groupby(df_merge_TV4_col2['Detail3']).unique().apply(pd.Series)


# In[849]:


#df2


# In[850]:



df2=pd.DataFrame(df2)


# In[851]:



df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[852]:



df_merge_TV4_col2=pd.merge(df_merge_TV4_col2,df2,on='LastDigit_PV',how='left')

df_merge_TV4_col2=df_merge_TV4_col2.rename(columns={"Detail3_y":"Detail3"})


# In[853]:


df3=df_merge_TV4_col2['Detail3'].groupby(df_merge_TV4_col2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[854]:


#df3


# In[855]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV4_col2=pd.merge(df_merge_TV4_col2,df3,on='Detail3',how='left')


# In[856]:



df_merge_TV4_col2=df_merge_TV4_col2.rename(columns={"CCCC_y":"CCCC"})

df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[857]:



df_merge_TV4_col2['CCCC']=df_merge_TV4_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV4_col2['DDDD']=df_merge_TV4_col2.groupby('Show_name_index').ngroup()


# In[858]:


df_merge_TV4_col2['DDDD']=df_merge_TV4_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[859]:


df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[860]:



df4=df_merge_TV4_col2['List_Heading'].groupby(df_merge_TV4_col2['AAAA']).unique().apply(pd.Series)

#df4.dropna(axis=1,inplace=True)

df4.reset_index(inplace=True)


# In[861]:


#df4


# In[862]:


df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].fillna('1236')


# In[863]:


df5=df_merge_TV4_col2['List_Heading'].groupby(df_merge_TV4_col2['BBBB']).unique().apply(pd.Series)


# In[864]:


#df5


# In[865]:


df_merge_TV4_col2['BBBB']=df_merge_TV4_col2['BBBB'].fillna('00305')


# In[866]:



df_merge_TV4_col2['ORD_new'] =df_merge_TV4_col2['AAAA']+df_merge_TV4_col2['BBBB']+df_merge_TV4_col2['CCCC']+df_merge_TV4_col2['DDDD']


# In[867]:


df_merge_TV4_col2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4_col2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[868]:



df_merge_TV4=[df_merge_TV4,df_merge_TV4_col2]


# In[869]:


df_merge_TV4=pd.concat(df_merge_TV4)


# In[870]:



df_merge_TV4['Detail3_x']=df_merge_TV4['Detail3_x'].fillna(df_merge_TV4['Detail3'])


# In[871]:


df_merge_TV4.drop("Detail3",axis='columns',inplace=True)


# In[872]:


df_merge_TV4=df_merge_TV4.rename(columns={"Detail3_x":"Detail3"})


# In[873]:


df_merge_TV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # concat DF

# In[874]:


df_all=[df_merge_TV1,
        df_merge_TV2,
        df_merge_TV3,
        df_merge_TV4,
        df_merge_TV5,
        df_merge_TV6,
        df_merge_SPTV1,
        df_merge_SPTV2,
        df_merge_SPTV3,
        df_merge_SPTV4,
        df_merge_SPTV5,
        df_merge_SPTV51,
        df_merge_Movie,
        df_merge_cable,      
        df_merge_add_cab]


# In[875]:


df_all=pd.concat(df_all)


# In[876]:


#df_all.columns


# In[877]:


df_all.drop(['StatisticID', 'CatSynID','NoteID','statusid'], axis=1, inplace=True)


# In[878]:


df_all['EditedBy']='codebookcreator'
df_all['EditedDate']=pd.to_datetime('today')
df_all['StudyEntryID']=451
df_all['VersionID']=17
df_all['SID']=1915


# In[879]:


df_all['Status']='Add'


# In[880]:


df_all['Definition'] = df_all.apply(lambda x: 0  if x['compare']==False else x['Definition'], axis=1)


# In[881]:


df_all['Definition']=df_all['Definition'].fillna('0')


# In[882]:


df_all['UCode']=df_all['UCode'].fillna('U0')
df_all['QuestionID']=df_all['QuestionID'].fillna(0)
df_all['QUESTID']=df_all['QUESTID'].fillna(0)
df_all['SDID']=df_all['SDID'].fillna(0)


# In[883]:


df_all['Initial_wave']=df_all['Initial_Wave'].replace(r'nan',np.nan, regex=True)
df_all['Initial_wave']=df_all['Initial_Wave'].fillna(0)


# In[884]:


df_all['StudyAnswerID']=0
df_all['Full_Label']=''


# In[885]:


df_all['AnswerID']=df_all['AnswerID'].fillna(0)


# In[886]:


df_all['Imported']=''
df_all['Min']=''
df_all['Max']=''


# In[887]:


df_all=df_all.sort_values(['Category','Detail1','Detail3'],ascending=[True,True,True])


# In[888]:


#df_all = df_all.astype( {"QLevel":'int32', "QUESTID":'int32', "AnswerID":'int32',"QuestionID":'int32',"SID":'int64', "SDID":'int32', "VersionID":'int32', "Wave":'int32', "Min":'float',"Max":'float', "StudyEntryID":'int64',"Imported":'bool'} )


# In[889]:


#import difflib as dl


# In[890]:


#list1=df_all['Show_Name']


# In[891]:


#list2=df_all['Detail1']


# In[892]:


#list1=list(list1)
#list2=list(list2)


# In[893]:


#dl.context_diff(list1,list2)


# In[894]:


#for diff in dl.context_diff(list1,list2):
   # print(diff)


# In[895]:


df_all['Show_Name'] = df_all.apply(lambda x: '#'+ x['Show_Name']  if x['DP_Status']=='#' else x['Show_Name'], axis=1)
#df_TV_Movie['Shows_Name'] = df_TV_Movie.apply(lambda x: '#'+x['Shows_Name']  if x['OneWave_Suppress']=='#' else x['Shows_Name'], axis=1)


# In[896]:


df_all['Diff_Detail1'] = df_all.apply(lambda x: 'Same'  if x['Show_Name']==x['Detail1'] else 'Not_Same', axis=1)


# In[897]:


df_all_with_detail1_diff=df_all[['Show_Name','Detail1','Diff_Detail1','Clean_Type','Category','DP_Status']]


# In[898]:


type(df_all_with_detail1_diff)


# In[899]:


df_all_with_detail1_diff.drop_duplicates(inplace=True)
#dataframe.where(dataframe.ID=='1')


# In[900]:


df_all_with_detail1_diff


# In[901]:


df_all_with_detail1_diff.shape


# 

# In[902]:


#df_all.shape


# In[903]:


df_all.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_all.csv",index=False,header=True,encoding='cp1252')


# In[904]:


df_all_with_detail1_diff.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_all_with_detail1_diff.csv",index=False,header=True,encoding='cp1252')


# In[905]:


#df_all.columns


# In[ ]:





# In[906]:


df_all.drop(['Clean_Type', 'Detail1','F2021_Client','compare','col2pv','LastDigit_PV','QID','CCP','Wave','ORD'], axis=1, inplace=True)


# In[907]:


df_all.rename(columns={'S2022_Client':'CCP','Show_Name':'Detail1','VersionID':'Version','Initial_Wave':'Wave','ORD_new':'ORD'},inplace=True)


# In[908]:


df_all=df_all[["StudyEntryID","SID","Version","Category","Super","Tmpl","Time Period","Detail1","Detail2",
"Detail3","Detail4","UCode","Definition","CCP","ORD","Wave","Status","Full_Label","QLevel","QUESTID","AnswerID","EditedBy","EditedDate","SDID",
"StudyAnswerID","QuestionID","Imported","Min","Max"]]


# In[909]:


df_all['Max'] =df_all['Max'].apply(pd.to_numeric)
df_all['Min'] =df_all['Min'].apply(pd.to_numeric)


# In[910]:


df_all['Wave']=df_all['Wave'].fillna(0)
df_all['Wave']=df_all['Wave'].replace(r'W', '', regex=True)


# In[911]:


df_all.to_csv('C:\\Users\\saraswathy.rajaman\\Documents\\Spring-2022_needupdates.csv',index=False,header=True,encoding='cp1252')


# In[912]:


#df_all.isnull().sum ()


# In[913]:


df_all = df_all.astype( {"QLevel":'int32',
                         "QUESTID":'int32',
                         "AnswerID":'int32',
                         "QuestionID":'int64', 
                         "SDID":'int32', 
                         "Version":'int32', 
                         "Min":'float',
                         "Max":'float', 
                         "StudyEntryID":'int64',
                         "Imported":'bool',
                         "Tmpl":'int64',
                         "Wave":'int32',
                         "StudyAnswerID":'int32'} )


# In[914]:


with engine.begin() as connection:
    df_merge_add_cab.to_sql(name="tmp_EditedRecords_addcable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df.to_sql('db_table2', engine, if_exists='replace')
#df_merge_cable.to_sql(name="tmp_EditedRecords_cable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_Movie.to_sql(name="tmp_EditedRecords_Movie_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_SPTV1.to_sql(name="tmp_EditedRecords_SPTV1_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_SPTV2.to_sql(name="tmp_EditedRecords_SPTV2_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_SPTV3.to_sql(name="tmp_EditedRecords_SPTV3_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_SPTV4.to_sql(name="tmp_EditedRecords_SPTV4_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_SPTV5.to_sql(name="tmp_EditedRecords_SPTV5_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_SPTV51.to_sql(name="tmp_EditedRecords_SPTV51_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_TV6.to_sql(name="tmp_EditedRecords_TV6_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_TV5.to_sql(name="tmp_EditedRecords_TV5_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_TV1.to_sql(name="tmp_EditedRecords_TV1_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_TV2.to_sql(name="tmp_EditedRecords_TV2_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_TV3.to_sql(name="tmp_EditedRecords_TV3_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df_merge_TV4.to_sql(name="tmp_EditedRecords_TV4_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_all.to_sql(name="tmp_EditedRecords_dfall_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_all.to_sql(name="tmp_EditedRecords_Hold",con=engine,schema="dbo",if_exists='append', chunksize=1000,index=False)
#"tmp_EditedRecords_Hold"
#df_all_with_detail1_diff.to_sql(name="tmp_EditedRecords_dfall_detail1_diff_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)


# In[ ]:


df_all.shape


# In[ ]:




