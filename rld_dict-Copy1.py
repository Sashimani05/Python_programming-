#!/usr/bin/env python
# coding: utf-8

# <h1>Table of Contents<span class="tocSkip"></span></h1>
# <div class="toc"><ul class="toc-item"><li><span><a href="#Import-necessary-Functions" data-toc-modified-id="Import-necessary-Functions-1"><span class="toc-item-num">1&nbsp;&nbsp;</span>Import necessary Functions</a></span></li><li><span><a href="#Get-the-input-file-path" data-toc-modified-id="Get-the-input-file-path-2"><span class="toc-item-num">2&nbsp;&nbsp;</span>Get the input file path</a></span></li><li><span><a href="#Get-the-Sheet-Name" data-toc-modified-id="Get-the-Sheet-Name-3"><span class="toc-item-num">3&nbsp;&nbsp;</span>Get the Sheet Name</a></span></li><li><span><a href="#Data-validation" data-toc-modified-id="Data-validation-4"><span class="toc-item-num">4&nbsp;&nbsp;</span>Data validation</a></span><ul class="toc-item"><li><span><a href="#Check-if-the-sheet-is-available-in-the-given-excel-workbook" data-toc-modified-id="Check-if-the-sheet-is-available-in-the-given-excel-workbook-4.1"><span class="toc-item-num">4.1&nbsp;&nbsp;</span>Check if the sheet is available in the given excel workbook</a></span></li><li><span><a href="#Read-that-sheet-as-Dataframe" data-toc-modified-id="Read-that-sheet-as-Dataframe-4.2"><span class="toc-item-num">4.2&nbsp;&nbsp;</span>Read that sheet as Dataframe</a></span></li><li><span><a href="#Check-if-the-dataframe-is-empty" data-toc-modified-id="Check-if-the-dataframe-is-empty-4.3"><span class="toc-item-num">4.3&nbsp;&nbsp;</span>Check if the dataframe is empty</a></span></li><li><span><a href="#clean-the--whitespace-in-column-name" data-toc-modified-id="clean-the--whitespace-in-column-name-4.4"><span class="toc-item-num">4.4&nbsp;&nbsp;</span>clean the  whitespace in column name</a></span></li><li><span><a href="#Drop-duplicates-in-dataframe" data-toc-modified-id="Drop-duplicates-in-dataframe-4.5"><span class="toc-item-num">4.5&nbsp;&nbsp;</span>Drop duplicates in dataframe</a></span></li><li><span><a href="#pad-the-column-with-zero-in-start-o-make-in-as-2-char--if-it-is-not" data-toc-modified-id="pad-the-column-with-zero-in-start-o-make-in-as-2-char--if-it-is-not-4.6"><span class="toc-item-num">4.6&nbsp;&nbsp;</span>pad the column with zero in start o make in as 2 char  if it is not</a></span></li><li><span><a href="#Concat-card-and-col-to-get-cardcol-values" data-toc-modified-id="Concat-card-and-col-to-get-cardcol-values-4.7"><span class="toc-item-num">4.7&nbsp;&nbsp;</span>Concat card and col to get cardcol values</a></span></li><li><span><a href="#Remove-white-space-from-the-data-frame" data-toc-modified-id="Remove-white-space-from-the-data-frame-4.8"><span class="toc-item-num">4.8&nbsp;&nbsp;</span>Remove white space from the data frame</a></span></li><li><span><a href="#Filter-non-numeric-values-and-save-in-a-seprate-dataframe" data-toc-modified-id="Filter-non-numeric-values-and-save-in-a-seprate-dataframe-4.9"><span class="toc-item-num">4.9&nbsp;&nbsp;</span>Filter non numeric values and save in a seprate dataframe</a></span></li><li><span><a href="#convert-all-non-numeric-value-in-cardcol-to-nan-values-and-drop-those-rows" data-toc-modified-id="convert-all-non-numeric-value-in-cardcol-to-nan-values-and-drop-those-rows-4.10"><span class="toc-item-num">4.10&nbsp;&nbsp;</span>convert all non numeric value in cardcol to nan values and drop those rows</a></span></li></ul></li><li><span><a href="#Read-the-RLD-file" data-toc-modified-id="Read-the-RLD-file-5"><span class="toc-item-num">5&nbsp;&nbsp;</span>Read the RLD file</a></span><ul class="toc-item"><li><span><a href="#convert-RLD-to-dataframe" data-toc-modified-id="convert-RLD-to-dataframe-5.1"><span class="toc-item-num">5.1&nbsp;&nbsp;</span>convert RLD to dataframe</a></span></li><li><span><a href="#Clean-the-columns-from-white-space" data-toc-modified-id="Clean-the-columns-from-white-space-5.2"><span class="toc-item-num">5.2&nbsp;&nbsp;</span>Clean the columns from white space</a></span></li><li><span><a href="#Clean-df-from-whitespace-on-all-the-string-values" data-toc-modified-id="Clean-df-from-whitespace-on-all-the-string-values-5.3"><span class="toc-item-num">5.3&nbsp;&nbsp;</span>Clean df from whitespace on all the string values</a></span></li><li><span><a href="#Drop-duplicates" data-toc-modified-id="Drop-duplicates-5.4"><span class="toc-item-num">5.4&nbsp;&nbsp;</span>Drop duplicates</a></span></li></ul></li><li><span><a href="#Convert-unique-function-to-list-values" data-toc-modified-id="Convert-unique-function-to-list-values-6"><span class="toc-item-num">6&nbsp;&nbsp;</span>Convert unique function to list values</a></span></li><li><span><a href="#Group-each-function-and-convert-them-to-dataframe" data-toc-modified-id="Group-each-function-and-convert-them-to-dataframe-7"><span class="toc-item-num">7&nbsp;&nbsp;</span>Group each function and convert them to dataframe</a></span></li><li><span><a href="#Calculate-the-end-range" data-toc-modified-id="Calculate-the-end-range-8"><span class="toc-item-num">8&nbsp;&nbsp;</span>Calculate the end range</a></span></li><li><span><a href="#calculate-the-list-range" data-toc-modified-id="calculate-the-list-range-9"><span class="toc-item-num">9&nbsp;&nbsp;</span>calculate the list range</a></span></li><li><span><a href="#Inner-merge-to-find-matching-values-b/w-rld-and-dictionary" data-toc-modified-id="Inner-merge-to-find-matching-values-b/w-rld-and-dictionary-10"><span class="toc-item-num">10&nbsp;&nbsp;</span>Inner merge to find matching values b/w rld and dictionary</a></span></li><li><span><a href="#Right-merge-to-find-the-list-that-dint-math-from-dictionary-only" data-toc-modified-id="Right-merge-to-find-the-list-that-dint-math-from-dictionary-only-11"><span class="toc-item-num">11&nbsp;&nbsp;</span>Right merge to find the list that dint math from dictionary only</a></span></li><li><span><a href="#RVOL-calculations" data-toc-modified-id="RVOL-calculations-12"><span class="toc-item-num">12&nbsp;&nbsp;</span>RVOL calculations</a></span></li><li><span><a href="#Range-Calculation" data-toc-modified-id="Range-Calculation-13"><span class="toc-item-num">13&nbsp;&nbsp;</span>Range Calculation</a></span><ul class="toc-item"><li><span><a href="#check-if-range-is-available-and-execute-the-calculation" data-toc-modified-id="check-if-range-is-available-and-execute-the-calculation-13.1"><span class="toc-item-num">13.1&nbsp;&nbsp;</span>check if range is available and execute the calculation</a></span></li></ul></li><li><span><a href="#CCP-Calculation" data-toc-modified-id="CCP-Calculation-14"><span class="toc-item-num">14&nbsp;&nbsp;</span>CCP Calculation</a></span></li><li><span><a href="#Prob-Calculation" data-toc-modified-id="Prob-Calculation-15"><span class="toc-item-num">15&nbsp;&nbsp;</span>Prob Calculation</a></span></li><li><span><a href="#DVOL-calculations" data-toc-modified-id="DVOL-calculations-16"><span class="toc-item-num">16&nbsp;&nbsp;</span>DVOL calculations</a></span><ul class="toc-item"><li><span><a href="#calculate-the-mapping-punch-in-rld-file" data-toc-modified-id="calculate-the-mapping-punch-in-rld-file-16.1"><span class="toc-item-num">16.1&nbsp;&nbsp;</span>calculate the mapping punch in rld file</a></span></li></ul></li></ul></div>

# ## Import necessary Functions 

# In[1]:


import pandas as pd
import numpy as np
import sys
from openpyxl import load_workbook
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)


# ## Get the input file path

# In[2]:


csv_path = input("Enter the csv path: ")

workbook = load_workbook(filename=csv_path)

print(f'Worksheet names: {workbook.sheetnames}')


# ## Get the Sheet Name 

# In[3]:


sname = input("Enter the sheetname: ")


# ## Data validation

# ### Check if the sheet is available in the given excel workbook

# In[4]:


if sname  in workbook.sheetnames:
    print (f'Sheet name',sname,'found')
else:
    import sys
    exit()


# ### Read that sheet as Dataframe

# In[5]:


#get_ipython().run_line_magic('timeit', '')
data = pd.read_excel( csv_path ,sheet_name=sname)


# ### Check if the dataframe is empty 

# In[6]:


def read_file():
    df_read = data
    if(df_read.empty):
        print ('file is empty')
        exit()
    else:
        print ('file is not empty')
        return df_read


# In[7]:


data=read_file()


# ### clean the  whitespace in column name 

# In[8]:


data.columns = data.columns.str.replace(' ', '')


# In[9]:


data=data.astype(str)


# ### Drop duplicates in dataframe

# In[10]:


data.drop_duplicates(inplace=True)


# ### pad the column with zero in start o make in as 2 char  if it is not

# In[11]:


data[['Column']] = data[['Column']].apply(lambda x: x.str.zfill(2))


# ### Concat card and col to get cardcol values

# In[12]:


data['CardCol']=data['Card']+data['Column']


# ### Remove white space from the data frame

# In[13]:


data = data.applymap(lambda x: x.strip() if isinstance(x, str) else x)


# ### Filter non numeric values and save in a seprate dataframe

# In[14]:


df_CC_non_num=data[data['CardCol'].apply(lambda x: x.isnumeric()==False)]


# In[15]:


df_CC_non_num=df_CC_non_num[['StudyName','StudyID','CCP','Full_Label','PersistentID']]


# In[16]:


df_CC_non_num.to_csv(r'.\df_CC_non_num.csv',index=False,header=True)


# ### convert all non numeric value in cardcol to nan values and drop those rows

# In[17]:


data['CardCol'] = data['CardCol'].apply(pd.to_numeric, errors='coerce')


# In[18]:


data.dropna(subset=["CardCol"], inplace=True) 


# ## Read the RLD file

# ### convert RLD to dataframe

# In[19]:


import pandas as pd
import zipfile

zip_file=zipfile.ZipFile((r"C:\Users\saraswathy.rajaman\Documents\Output.zip"))
dfs = []
for i in range(len(zip_file.namelist())):
    df=pd.read_csv(zip_file.open(zipfile.ZipFile.namelist(zip_file)[i]),header='infer', sep='|',encoding='UTF-8', on_bad_lines='skip')
    dfs.append(df)
    i=i+1
df_RLD = pd.concat(dfs, ignore_index=True)


# ### Clean the columns from white space

# In[20]:


df_RLD.columns = df_RLD.columns.str.replace(' ', '')


# ### Clean df from whitespace on all the string values

# In[21]:


df_RLD = df_RLD.applymap(lambda x: x.strip() if isinstance(x, str) else x)


# ### Drop duplicates

# In[ ]:


#if duplicated deducted in RLD write in a sepeprate file and exit 

#boolean = df_RLD.duplicated().any()
#if boolean==True:
    #print ("We found duplicate records in the RLD File.Please check the file")
    #RLD_Duplicated=df_RLD.duplicated(subset=None,keep=First)
    #RLD_Duplicated.to_csv("r'C:\Users\saraswathy.rajaman\Documents\RLD_Duplicates.csv',index=False,header=True)           
    #exit()             
    
    


# In[22]:


df_RLD.drop_duplicates(inplace=True)


# In[23]:


df_RLD=df_RLD.astype(str)


# ## Convert unique function to list values

# In[25]:


functions=data['Function'].unique().tolist()


# dfs = {Function: data.loc[data.Function == Function, :] for Function in functions}

# ## Group each function and convert them to dataframe 

# In[26]:


df={}
grouped = data.groupby('Function')
for group in grouped.groups.keys():
    #print(group)
    df[group] = grouped.get_group(group)


# ## Calculate the end range
# ## calculate the list range

# In[27]:


def Endrange(df):
    
    #data.shape
    if(df.empty):
        print ('file is empty')
        exit()
    else:
        df['Col2Span']=df['Col2Span'].astype(int)
        df['CardCol']=df['CardCol'].astype(int)
        
        df['Endrange']=df['CardCol']+df['Col2Span']-1
        def ccp_range(row):
            return range(row.CardCol, row.Endrange+1)
        df['list_range'] = df.apply(ccp_range, axis=1)
        df=df.explode('list_range', ignore_index=True)
        #df['list_range'] = df.apply(ccp_range, axis=1)
        
        #df=df.explode('list_range', ignore_index=True)
    return df   


# ## Inner merge to find matching values b/w rld and dictionary

# In[28]:


def inner_merge(rld,rld_dict):
    
    df_inner= pd.merge(rld,rld_dict, left_on=['Cardcol'], right_on=['list_range'],how='inner',indicator=True)
    df_inner=df_inner.where(df_inner['_merge']=='both')
    df_inner.dropna(subset=['Punch'],inplace=True)
    return (df_inner)
    


# ## Right merge to find the list that dint math from dictionary only

# In[29]:



def right_merge(rld,rld_dict):
    df_leftout=pd.merge(rld,rld_dict, left_on=['Cardcol'], right_on=['list_range'],how='right',indicator=True)
    df_leftout=df_leftout.where(df_leftout['_merge']=='right_only')
    return (df_leftout)


# In[30]:


all_left = pd.DataFrame()
all_inner = pd.DataFrame()
#newDF = newDF.append(oldDF, ignore_index = True)


# ## RVOL calculations

# In[31]:


if 'rvol' in functions:
    rvol=df['rvol']
    rvol=Endrange(rvol)
    rvol_inner=inner_merge(df_RLD,rvol)
    #rvol_outer=outer_merge(df_RLD,rvol)
    if len(rvol_inner)>0:
        rvol_inner = rvol_inner.sort_values(['Respondent_ID','CCP','list_range'], ascending=(True, True,True))
        rvol_inner['Punch'] = rvol_inner.groupby(['Respondent_ID','CCP'])['Punch'].transform(lambda x: ''.join(x))
        rvol_inner['Punch']=rvol_inner['Punch'].astype(int)
        rvol_inner['ImpliedDecimals']=rvol_inner['ImpliedDecimals'].astype(int)
        rvol_inner['Punch_with_implieddecimals']=rvol_inner['Punch']/np.power(10,rvol_innerl['ImpliedDecimals'])
        rvol_inner=rvol_inner[['StudyID','Respondent_ID','Punch_with_implieddecimals','PersistentID','CCP','Full_Label']]
        rvol_inner.drop_duplicates(inplace=True)
        rvol_inner.rename(columns = {'Punch_with_implieddecimals':'Value'}, inplace = True)
        all_inner = pd.concat([all_inner,rvol_inner], ignore_index = True)
    rvol_leftout=right_merge(df_RLD,rvol)
    rvol_leftout=rvol_leftout[['StudyName','StudyID','CCP','Full_Label','PersistentID',]]
    rvol_leftout.drop_duplicates(inplace=True)
    all_left = pd.concat([all_left,rvol_leftout], ignore_index = True)
    
    #rvol.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_rvol.csv',index=False,header=True)
    #rvol_inner.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_inner_rvol.csv',index=False,header=True)
    #rvol_leftout.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_rvol_leftout.csv',index=False,header=True)
    #rvol_outer.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_rvol_outer.csv',index=False,header=True)
    


# ## Range Calculation

# ### check if range is available and execute the calculation

# In[32]:


if 'range' in functions:
    df_range=df['range']
    df_range=Endrange(df_range)
    range_inner=inner_merge(df_RLD,df_range)
    if len(range_inner)>0:
        range_inner = range_inner.sort_values(['Respondent_ID','CCP','list_range'], ascending=(True, True,True))
        range_inner['Punch'] = range_inner.groupby(['Respondent_ID','CCP'])['Punch'].transform(lambda x: ''.join(x))
        range_inner['Punch']=range_inner['Punch'].astype(int)
        
        #range_inner['ImpliedDecimals']=range_inner['ImpliedDecimals'].astype(int)
        #range_inner['Punch_with_implieddecimals']=range_inner['Punch']/np.power(10,range_innerl['ImpliedDecimals'])
        
        range_inner=range_inner[['StudyID','Respondent_ID','Punch','PersistentID','CCP','Full_Label']]
        range_inner.drop_duplicates(inplace=True)
        range_inner.rename(columns = {'Punch':'Value'}, inplace = True)
        all_inner =  pd.concat([all_inner,range_inner], ignore_index = True)
    range_leftout=right_merge(df_RLD,df_range)
    range_leftout=range_leftout[['StudyName','StudyID','CCP','Full_Label','PersistentID',]]
    range_leftout.drop_duplicates(inplace=True)
    all_left = pd.concat([all_left,range_leftout], ignore_index = True)
    
    #df_range.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_range.csv',index=False,header=True)
    #range_inner.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_range_rvol.csv',index=False,header=True)
    #range_leftout.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_range_leftout.csv',index=False,header=True)


# ## CCP Calculation

# In[ ]:


if 'ccp' in functions:
    ccp=df['ccp']
    #ccp=Endrange(ccp)
    ccp_inner=pd.merge(df_RLD.astype(str),ccp.astype(str), left_on=['Cardcol'], right_on=['CardCol'],how='inner',indicator=True)
     
    if len(ccp_inner)>0:
        ccp_inner = ccp_inner.sort_values(['Respondent_ID','CCP'], ascending=(True, True))
        ccp_inner=ccp_inner[['StudyID','Respondent_ID','Punch','PersistentID','CCP','Full_Label']]
        ccp_inner.drop_duplicates(inplace=True)
        ccp_inner.rename(columns = {'Punch':'Value'}, inplace = True)
        all_inner = pd.concat([all_inner,ccp_inner], ignore_index = True)
    df_RLD=df_RLD.astype(str)
    ccp=ccp.astype(str)
    ccp_leftout=pd.merge(df_RLD,ccp, left_on=['Cardcol'], right_on=['CardCol'],how='right',indicator=True)
    ccp_leftout=ccp_leftout.where(ccp_leftout['_merge']=='right_only')
    ccp_leftout=ccp_leftout[['StudyName','StudyID','CCP','Full_Label','PersistentID',]]
    ccp_leftout.drop_duplicates(inplace=True)
    all_left =  pd.concat([all_left,ccp_leftout], ignore_index = True)
    
    #ccp.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_ccp.csv',index=False,header=True)
    #ccp_inner.to_csv(r'C:\Users\saraswathy.rajaman\Documents\inner_ccp.csv',index=False,header=True)
    #ccp_leftout.to_csv(r'C:\Users\saraswathy.rajaman\Documents\ccp_leftout.csv',index=False,header=True)


# ## Prob Calculation

# In[34]:


if 'prob' in functions:
    prob=df['prob']
    prob=Endrange(prob)
    prob_inner=inner_merge(df_RLD,prob)
    #prob_outer=outer_merge(df_RLD,prob)
    if len(prob_inner)>0:
        prob_inner = prob_inner.sort_values(['Respondent_ID','CCP','list_prob'], ascending=(True, True,True))
        prob_inner['Punch'] = prob_inner.groupby(['Respondent_ID','CCP'])['Punch'].transform(lambda x: ''.join(x))
        prob_inner['Punch']=prob_inner['Punch'].astype(int)
        prob_inner=prob_inner[['StudyID','Respondent_ID','Punch','PersistentID','CCP','Full_Label']]
        prob_inner.drop_duplicates(inplace=True)
        prob_inner.rename(columns = {'Punch':'Value'}, inplace = True)
        all_inner =  pd.concat([all_inner,prob_inner], ignore_index = True)
    prob_leftout=right_merge(df_RLD,prob)
    prob_leftout=prob_leftout[['StudyName','StudyID','CCP','Full_Label','PersistentID',]]
    prob_leftout.drop_duplicates(inplace=True)
    all_left =  pd.concat([all_left,prob_leftout], ignore_index = True)
    
    #prob.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_prob.csv',index=False,header=True)
    #prob_inner.to_csv(r'C:\Users\saraswathy.rajaman\Documents\prob_inner.csv',index=False,header=True)
    #prob_leftout.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_prob_leftout.csv',index=False,header=True)
    #prob_outer.to_csv(r'C:\Users\saraswathy.rajaman\Documents\prob_outer.csv',index=False,header=True)


# ##  DVOL calculations

# ### calculate the mapping punch in rld file

# In[ ]:


Punch_map = {'Y':'11', 
             'X':'10',
             '0':'9',
             '1':'8',
             '2':'7',
             '3':'6',
             '4':'5',
             '5':'4',
             '6':'3',
             '7':'2',
             '8':'1',
             '9':'0'}


# In[ ]:


df_RLD['Punch_new'] = df_RLD['Punch'].map(Punch_map)


# In[ ]:


df_RLD['Punch_new']=df_RLD['Punch_new'].astype(int)


# In[ ]:


df_RLD['pow']=np.power(2,df_RLD['Punch_new'])


# In[ ]:


if 'dvol' in functions:
    dvol=df['dvol']
    
    dvol_inner=pd.merge(df_RLD.astype(str),dvol.astype(str), left_on=['Cardcol'], right_on=['CardCol'],how='inner',indicator=True)
    if len(dvol_inner)>0:
        dvol_inner = dvol_inner.sort_values(['Respondent_ID','CCP'], ascending=(True, True))
       
        dvol_inner['pow']=dvol_inner['pow'].astype(int)
        dvol_inner['Cardcol']=dvol_inner['Cardcol'].astype(int)
        dvol_inner = dvol_inner.groupby(['StudyID','Respondent_ID','CCP','PersistentID','Full_Label'],as_index=False).agg({'Cardcol':'count','pow':'sum'})
        dvol_inner=dvol_inner[['StudyID','Respondent_ID','pow','CCP','PersistentID','Full_Label']]
        dvol_inner.drop_duplicates(inplace=True)
        dvol_inner.rename(columns = {'pow':'Value'}, inplace = True)
        all_inner =  pd.concat([all_inner,dvol_inner],ignore_index=True)
       
    dvol_leftout=pd.merge(df_RLD.astype(str),dvol.astype(str), left_on=['Cardcol'], right_on=['CardCol'],how='right',indicator=True)
    dvol_leftout=dvol_leftout[['StudyName','StudyID','CCP','Full_Label','PersistentID',]]
    dvol_leftout.drop_duplicates(inplace=True)
    all_left =pd.concat([all_left,dvol_leftout],ignore_index=True)
    
    #dvol.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_dvol.csv',index=False,header=True)
    #dvol_inner.to_csv(r'C:\Users\saraswathy.rajaman\Documents\inner_dvol.csv',index=False,header=True)
    #dvol_leftout.to_csv(r'C:\Users\saraswathy.rajaman\Documents\dvol_leftout.csv',index=False,header=True)


# In[ ]:


all_left.to_csv(r'C:\Users\saraswathy.rajaman\Documents\all_left.csv',index=False,header=True)


# In[ ]:


all_inner.to_csv(r'C:\Users\saraswathy.rajaman\Documents\all_inner.csv',index=False,header=True)


# 
