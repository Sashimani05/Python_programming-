#!/usr/bin/env python
# coding: utf-8

# <h1>Table of Contents<span class="tocSkip"></span></h1>
# <div class="toc"><ul class="toc-item"></ul></div>

# In[1]:


import warnings
warnings.filterwarnings("ignore")


# In[2]:


import pandas as pd
import numpy as np
from openpyxl import Workbook
import re
import openpyxl.utils.dataframe
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


# In[3]:


#Display settings
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
pd.set_option('display.colheader_justify', 'center')
pd.set_option('display.precision', 3)


# # Check if the Sheet name is availabe

# In[4]:



workbook = load_workbook(filename='C:\\Users\\saraswathy.rajaman\\Downloads\\w86_Spring22_tvshows_short.xlsm')


# In[5]:


name_of_sheet='gerardo'


# In[6]:


def sheets_name_check(workbook,name_of_sheet):
    sheetnames=workbook.sheetnames
    sheetnames=list(sheetnames)
    #print(sheetnames)
    for i in sheetnames:
        if str(name_of_sheet) == i:
            
            print(name_of_sheet,"is found in the workbook",workbook)
            
            break
    else:
        print (name_of_sheet,"not found")


# In[7]:


sheets_name_check(workbook,name_of_sheet)


# # check the necessary headers values on a given worksheet 

# In[8]:


sheet=workbook['gerardo']


# In[9]:


row = sheet.max_row
column = sheet.max_column
  
#print("Total Rows:", row)
#print("Total Columns:", column)


# In[10]:


#print("\nValue of header")
HV={}
header_values=[]
for i in range(1, column + 1): 
    cell_obj = sheet.cell(row = 1, column = i)
    HV[i]=cell_obj.value
    
    header_values.append(HV[i])
    #print(cell_obj.value, end = '\n')


# # input the column header

# In[11]:


list_tbf=['Clean_Type','Network','S2022_Client',
'Show_Name','Section_Heading','List_Heading','Initial_Wave','F2021_Client']


# In[12]:


def get_key(val):
            for key, value in HV.items():
                 if val== value:
                    print("found ",val,'in column',key)
                    return key
            
            print ("could not find ",val,'in any column')
            return "key doesn't exist"


# In[13]:


for i in  range(len(list_tbf)):
    val=list_tbf[i]
    get_key(val)


# In[14]:


Flag={}
for i in list_tbf:
    Flag[i]= i in HV.values() 


# In[15]:


#Flag.values()


# In[16]:


Flag_header=False in Flag.values()


# # convert  worksheet to dataframes 

# In[17]:



df = pd.read_excel('C:\\Users\\saraswathy.rajaman\\Downloads\\w86_Spring22_tvshows_short.xlsm', sheet_name='gerardo')
pm = pd.read_excel('C:\\Users\\saraswathy.rajaman\\Downloads\\w86_Spring22_tvshows_short.xlsm', sheet_name='PunchMap')


# In[18]:


sheet=workbook['gerardo']


# In[19]:


rows = sheet.max_row
columns = sheet.max_column


# In[20]:


#pm['PunchValue'] = pm['PunchValue'].str.lower()


# In[21]:


pm['PunchValue']=pm['PunchValue'].replace('X','x',regex=True)


# In[22]:


#pm['PunchValue']


# # check if the column header of the given sheet is empty

# if sheet["A1"].value=='Line_Type':
#     print ("First column name is right as",sheet["A1"].value)
# else:
#     print("First column value is not right ")
#     

# In[23]:


#convert the first row as list as the first row is always the column names in a dataframe
my_list = df.columns.values.tolist()


# In[24]:


#slice the string in the list and check if that is unnamed
mylist1= [w[:7] for w in my_list]


# In[25]:


i=0
for x in mylist1:
    if x=='Unnamed':
        print (i," cell in first row  empty")
        print('terminated further execution as the first row has empty cells :')
        Flag_empty=True
        break
    i+=1
print("No empty cells in first row")
Flag_Empty=False


# In[26]:


Flag_Empty


# # check the cleantype in two sheets are equal

# In[27]:


firstlist=list(pm['Clean_Type'].unique())

secondlist=list(df['Clean_Type'].unique())


# In[28]:


# function to check both lists if they are equal

def checkList(firstlist, secondlist):
    # sorting the lists
    firstlist.sort()
    secondlist.sort()
    # if both the lists are equal the print yes
    if(firstlist == secondlist):
        print("Both cleantype list are equal")
        Flag_cleantype=True
    else:
        print("Both cleantype list are not equal")
        Flag_cleantype=False

# passing both the lists to checklist function


# In[29]:


checkList(firstlist, secondlist)


# # Read the data from sql invoke the stored procedure

# import pyodbc 
# 
# conn = pyodbc.connect('Driver={SQL Server Native Client 11.0};'
#                       'Server=internalSQLdev.mridevops.com;'
#                       'Database=Codebook_Taxonomy;'
#                       'Trusted_Connection=yes;')
# Prev_Dict_StudyEntryID = '434'
# 
# Prev_Dict_VersionID = '13'
# 
# query = "EXEC [app_Codebook_Read] @VersionID = {0}, @StudyEntryID = {1}".format(Prev_Dict_VersionID, Prev_Dict_StudyEntryID)
# df_dict = pd.read_sql_query(query, conn)
# #df_dict.to_excel(r'C:\Users\saraswathy.rajaman\Documents\df_dict.xls', sheet_name = "Sheet1", header = True, index = False)
# 
# 
# print((df_dict.head(10)))

# In[30]:


#df_dict.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df-dict.txt', sep='\t', index=False,header=True,encoding='cp1252')


# In[31]:


df_dict=pd.read_excel(r'C:\Users\saraswathy.rajaman\Downloads\Winter-2021.xlsx' 
                          ,converters={'ORD': lambda x: f'{x:20}'})


    #data=pandas.read_csv(‘filename.txt’, names=[“Column1”, “Column2”])


# In[32]:


#df_dict['ORD']=df_dict['ORD'].apply(lambda x: '{:.0f}'.format(x))


# In[33]:


#type(df_dict1['ORD'])


# In[34]:


df['compare'] = (df['S2022_Client'] == df['F2021_Client'])


# In[35]:


#df_dict.head(10)


# In[36]:


df['col2pv'] = ''


# In[37]:


#Removing spl character
df['S2022_Client']=df['S2022_Client'].str.replace('*','')
df['F2021_Client']=df['F2021_Client'].str.replace('*','')


# In[38]:


#df.head(5)


# In[39]:


df['Show_name_index']=df.index


# In[40]:


#
#
#display(df.head(10))


# # Drop the supress with X rows and pick necessary column

# In[41]:


df.drop(df.index[df['DP_Status'] == 'X'], inplace = True)


# In[42]:


df=df[['Clean_Type','S2022_Client','Section_Heading','List_Heading', 'DP_Status', 'Show_Name','Show_name_index','Initial_Wave',
'F2021_Client','compare','col2pv']]


# In[43]:


#df.info()


# In[44]:


df.to_csv(r'C:\Users\saraswathy.rajaman\Documents\df_aftercleanup.csv',index=False,header=True,encoding='cp1252')


# In[45]:


#group data-df-(TVmedia file) based on cleantype into different dataframes
data={}
grouped = df.groupby('Clean_Type')
for group in grouped.groups.keys():
    #print(group)
    data[group] = grouped.get_group(group)


# In[46]:


#data['add_cabl']


# # Group Punchvalues

# In[47]:


#group punchvalues-pm-(Punchmap file) based on cleantype into different dataframes
PV={}
grouped = pm.groupby('Clean_Type')
for group in grouped.groups.keys():
    #print(group)
    PV[group] = grouped.get_group(group)


# In[48]:


#PV['Movie']


# In[49]:


# Function to find no of columns 1 or 2 pv in each cleantype in punch map dataframe


# In[50]:


def punchv(Punch):
    pm_ct=Punch['Clean_Type'].unique()
    #print (pm_ct,'-',Punch['Columns'].nunique())


# In[51]:



for i in grouped.groups.keys():
    Punch=PV[i]
    punchv(Punch)


# # add_cable

# In[52]:


Punch_variable=PV['add_cabl']['PunchValue']


# In[53]:


#data['add_cabl']['F2021_Client'].dtype


# In[54]:


datapv={}
add_cab=[]
for i in Punch_variable:
   
    datapv[i]=data['add_cabl'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    add_cab.append(datapv[i])


# In[55]:


add_cab=pd.concat(add_cab)


# In[56]:


df_merge_add_cab= pd.merge(add_cab, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[57]:


#df_merge_add_cab.shape


# In[58]:


#df_merge_add_cab.head(10)


# def PV_assign(PV):
#     Punch_variable=PV['add_cabl']['PunchValue']
#     print (Punch_variable)

# PV_assign(PV)

# In[59]:


df_merge_add_cab['LastDigit_PV']=df_merge_add_cab['S2022_Client'].str.strip().str[-1]


# In[60]:


df_merge_add_cab=df_merge_add_cab.sort_values(['LastDigit_PV'], na_position='last',
               ascending=[True])
						  
df_merge_add_cab['Tmpl']=df_merge_add_cab['Tmpl'].fillna(method='ffill')
df_merge_add_cab['Super']=df_merge_add_cab['Super'].fillna(method='ffill')
df_merge_add_cab['Detail3']=df_merge_add_cab['Detail3'].fillna(method='ffill')


# In[61]:


df_merge_add_cab=df_merge_add_cab.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_add_cab['Category']=df_merge_add_cab['Category'].fillna(method='ffill')
df_merge_add_cab['QLevel']=df_merge_add_cab['QLevel'].fillna(method='ffill')
df_merge_add_cab['Detail2']=df_merge_add_cab['Detail2'].fillna(method='ffill')


# In[62]:


df_merge_add_cab['ORD']=df_merge_add_cab['ORD'].astype(str)


# In[63]:


df_merge_add_cab=df_merge_add_cab.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[64]:


df_merge_add_cab['CCCC']=df_merge_add_cab["ORD"].str.slice(9,13,1)


# In[65]:


df_merge_add_cab['CCCC']=df_merge_add_cab['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[66]:


df_merge_add_cab['CCCC']=df_merge_add_cab['CCCC'].fillna(method='ffill')


# In[67]:


df_merge_add_cab['DDDD']=df_merge_add_cab.groupby('Show_name_index').ngroup()


# In[68]:


df_merge_add_cab['DDDD']=df_merge_add_cab['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[69]:


df_merge_add_cab=df_merge_add_cab.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[70]:


df_merge_add_cab['ORD']=df_merge_add_cab['ORD'].astype(str)


# In[71]:


df_merge_add_cab['AAAA']=df_merge_add_cab["ORD"].str.slice(0,4,1)


# In[72]:


df_merge_add_cab['BBBB']=df_merge_add_cab["ORD"].str.slice(4,9,1)


# In[73]:


df_merge_add_cab=df_merge_add_cab.sort_values(['Super','Category'],
               ascending=[True,True],na_position='last')


# In[74]:


df_merge_add_cab['AAAA']=df_merge_add_cab['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[75]:


df_merge_add_cab['AAAA']=df_merge_add_cab['AAAA'].replace(r'nan',np.nan, regex=True)


# In[76]:


df_merge_add_cab['AAAA']=df_merge_add_cab['AAAA'].fillna(method='ffill')


# In[77]:


df_merge_add_cab['BBBB']=df_merge_add_cab['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[78]:


df_merge_add_cab['BBBB']=df_merge_add_cab['BBBB'].replace(r'nan',np.nan, regex=True)


# In[79]:


df_merge_add_cab['BBBB']=df_merge_add_cab['BBBB'].fillna(method='ffill')


# In[80]:



df_merge_add_cab['ORD_new'] =df_merge_add_cab['AAAA']+df_merge_add_cab['BBBB']+df_merge_add_cab['CCCC']+df_merge_add_cab['DDDD']


# In[81]:


df_merge_add_cab.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_add_cab.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # Cable 

# In[82]:


Punch_variable=PV['cable']['PunchValue']


# In[83]:


#Punch_variable


# In[84]:


datapv={}
cable=[]
for i in Punch_variable:
   
    datapv[i]=data['cable'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    cable.append(datapv[i])


# In[85]:


cable=pd.concat(cable)


# In[86]:


cable.S2022_Client = cable.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[87]:


df_merge_cable= pd.merge(cable, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[88]:


df_merge_cable.shape


# In[89]:


df_merge_cable['LastDigit_PV']=df_merge_cable['S2022_Client'].str.strip().str[-1]


# In[90]:


df2=df_merge_cable['LastDigit_PV'].groupby(df_merge_cable['Detail3']).unique().apply(pd.Series)
#df['subreddit'].groupby(df['author']).unique().apply(pd.Series)


# In[91]:


df2=pd.DataFrame(df2)
df2.reset_index(inplace=True)


# In[92]:


#df2.loc([['Full Attention', 'Most Attention', 'Some Attention', 'Watched last 30 days', 'Watched last 7 days']])


# In[93]:


#df2


# In[94]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[95]:


df_merge_cable['ORD']=df_merge_cable['ORD'].astype(str)


# In[96]:


df_merge_cable['CCCC']=df_merge_cable["ORD"].str.slice(9,13,1)


# In[97]:


df3=df_merge_cable['Detail3'].groupby(df_merge_cable['CCCC']).unique().apply(pd.Series)


# In[98]:


df3=pd.DataFrame(df3)


# In[99]:


df3.reset_index(inplace=True)


# In[100]:


df3.dropna(inplace=True)


# In[101]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[102]:


df3.dropna(inplace=True)


# In[103]:


df_merge_cable=pd.merge(df_merge_cable,df2,on='LastDigit_PV',how='left')


# In[104]:


df_merge_cable.shape


# In[105]:


df_merge_cable.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_cable_detail3.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[106]:


#df_merge_cable.drop("Detail3_x",axis='columns',inplace=True)


# In[107]:


df_merge_cable=df_merge_cable.rename(columns={"Detail3_y":"Detail3"})


# In[108]:


df_merge_cable=pd.merge(df_merge_cable,df3,on='Detail3',how='left')


# In[109]:


df_merge_cable=df_merge_cable.rename(columns={"CCCC_y":"CCCC"})


# In[110]:



df_merge_cable=df_merge_cable.sort_values(['LastDigit_PV'],ascending=[True],na_position='last')


# In[111]:


df_merge_cable['Tmpl']=df_merge_cable['Tmpl'].fillna(method='ffill')
df_merge_cable['Super']=df_merge_cable['Super'].fillna(method='ffill')
#df_merge_cable['Detail3']=df_merge_cable['Detail3'].fillna(method='ffill')
df_merge_cable['Tmpl']=df_merge_cable['Tmpl'].fillna(2)
df_merge_cable['Super']=df_merge_cable['Super'].fillna('Media - Cable')


# In[112]:


df_merge_cable=df_merge_cable.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_cable['Category']=df_merge_cable['Category'].fillna(method='ffill')
df_merge_cable['QLevel']=df_merge_cable['QLevel'].fillna(method='ffill')
#df_merge_cable['Detail2']=df_merge_cable['Detail2'].fillna(method='ffill')


# In[113]:


df_merge_cable['ORD']=df_merge_cable['ORD'].astype(str)


# In[114]:



df_merge_cable=df_merge_cable.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[115]:



df_merge_cable['CCCC']=df_merge_cable['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[116]:


#df_merge_cable['CCCC']=df_merge_cable['CCCC'].fillna(method='ffill')


# In[117]:


df_merge_cable['DDDD']=df_merge_cable.groupby('Show_name_index').ngroup()


# In[118]:


df_merge_cable['DDDD']=df_merge_cable['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[119]:


df_merge_cable=df_merge_cable.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[120]:


df_merge_cable['ORD']=df_merge_cable['ORD'].astype(str)


# In[121]:


df_merge_cable['AAAA']=df_merge_cable["ORD"].str.slice(0,4,1)


# In[122]:


df_merge_cable['BBBB']=df_merge_cable["ORD"].str.slice(4,9,1)


# In[123]:


df_merge_cable=df_merge_cable.sort_values(['Super','Category'],
               ascending=[True,True],na_position='last')


# In[124]:


df_merge_cable['AAAA']=df_merge_cable['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[125]:


df_merge_cable['AAAA']=df_merge_cable['AAAA'].replace(r'nan',np.nan, regex=True)


# In[126]:


df_merge_cable['AAAA']=df_merge_cable['AAAA'].fillna('1098')


# In[127]:


df_merge_cable['BBBB']=df_merge_cable['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[128]:


df_merge_cable['BBBB']=df_merge_cable['BBBB'].replace(r'nan',np.nan, regex=True)


# In[129]:


df_merge_cable['BBBB']=df_merge_cable['BBBB'].fillna('00167')


# In[130]:


df_merge_cable['ORD_new'] =df_merge_cable['AAAA']+df_merge_cable['BBBB']+df_merge_cable['CCCC']+df_merge_cable['DDDD']


# In[131]:


df_merge_cable['Show_Name'] = df_merge_cable.apply(lambda x: x['List_Heading']+': '+x['Show_Name'] , axis=1)


# In[132]:



df_merge_cable.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_cable.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[133]:


from sqlalchemy import create_engine


# In[134]:


DB = {'server':'internalSQLdev.mridevops.com','database':'Codebook_Taxonomy','driver':'driver=SQL Server Native Client 11.0','pyodb_d':'SQL Server Native Client 11.0'}
#engine=create_engine('mssql+pyodbc://'+ DB['server']+'/'+ DB['database']+'?'+ DB['driver'])


# In[135]:



engine = create_engine('mssql+pyodbc://' + DB['server'] + '/' + DB['database'] + '?' + DB['driver'], fast_executemany = True)


# In[136]:


import pyodbc


# In[137]:


#conn = pyodbc.connect('Driver={'+DB['pyodb_d']+'}; Server='+DB['server']+';Database='+DB['database']+'; Trusted_Connection=yes;')


# In[138]:


#conn.commit()


# with engine.begin() as connection:
#     df_merge_add_cab.to_sql(name="tmp_EditedRecords_addcable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
# #df.to_sql('db_table2', engine, if_exists='replace')
# df_merge_cable.to_sql(name="tmp_EditedRecords_cable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
# #

# # Movies

# In[139]:


Punch_variable=PV['Movie']['PunchValue']

#Punch_variable=PV['add_cabl']['PunchValue']


# In[ ]:





# In[140]:


datapv={}
Movie=[]
for i in Punch_variable:
   
    datapv[i]=data['Movie'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    Movie.append(datapv[i])


# In[141]:


#PV['Movie']['PunchValue']


# In[142]:


#Movie


# In[143]:


Movie=pd.concat(Movie)


# In[144]:


#Movie


# In[145]:


Movie.F2021_Client = Movie.F2021_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[146]:


df_merge_Movie= pd.merge(Movie, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[147]:


df_merge_Movie['LastDigit_PV']=df_merge_Movie['S2022_Client'].str.strip().str[-1]


# In[148]:


#df_merge_Movie
df_merge_Movie.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_Movie_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[149]:


df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '1', df_merge_Movie['Detail3'].fillna('Saw at movie theater'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '2', df_merge_Movie['Detail3'].fillna('Rented movie and viewed on DVD or Blu-ray'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '3', df_merge_Movie['Detail3'].fillna('Purchased movie and viewed on DVD or Blu-ray'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3'] = np.where(df_merge_Movie['LastDigit_PV'] == '4', df_merge_Movie['Detail3'].fillna('Viewed with Video On Demand or PPV'),df_merge_Movie['Detail3'])
df_merge_Movie['Detail3']= np.where(df_merge_Movie['LastDigit_PV'] == '5', df_merge_Movie['Detail3'].fillna('Downloaded or Streamed from the Internet'),df_merge_Movie['Detail3'])


# In[150]:


#df_merge_Movie
#df_merge_Movie
df_merge_Movie.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_Movie_beforefill_arranged.txt", sep='\t', index=False,header=True,encoding='cp1252')


# df_merge_Movie=df_merge_Movie.sort_values(['LastDigit_PV','ORD'], 
#                ascending=[True,True],na_position='last')

# In[151]:


df_merge_Movie=df_merge_Movie.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_Movie['Tmpl']=df_merge_Movie['Tmpl'].fillna(method='ffill')
df_merge_Movie['Super']=df_merge_Movie['Super'].fillna(method='ffill')
#f_merge_Movie['Detail3']=df_merge_Movie['Detail3'].fillna(method='ffill')


# In[152]:


df_merge_Movie=df_merge_Movie.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_Movie['Category']=df_merge_Movie['Category'].fillna(method='ffill')
df_merge_Movie['QLevel']=df_merge_Movie['QLevel'].fillna(method='ffill')
df_merge_Movie['Detail2']=df_merge_Movie['Detail2'].fillna(method='ffill')


# In[153]:


df_merge_Movie['ORD']=df_merge_Movie['ORD'].astype(str)


# In[154]:


df_merge_Movie=df_merge_Movie.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[155]:


df_merge_Movie['CCCC']=df_merge_Movie["ORD"].str.slice(9,13,1)


# In[156]:


df_merge_Movie['CCCC']=df_merge_Movie['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[157]:


#df_merge_Movie['CCCC']=df_merge_Movie['CCCC'].fillna(method='ffill')


# In[158]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Downloaded or Streamed from the Internet', df_merge_Movie['CCCC'].fillna('0004'),df_merge_Movie['CCCC'])


# In[159]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Purchased movie and viewed on DVD or Blu-ray', df_merge_Movie['CCCC'].fillna('0002'),df_merge_Movie['CCCC'])


# In[160]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Rented movie and viewed on DVD or Blu-ray', df_merge_Movie['CCCC'].fillna('0001'),df_merge_Movie['CCCC'])


# In[161]:


df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Saw at movie theater', df_merge_Movie['CCCC'].fillna('0000'),df_merge_Movie['CCCC'])


# In[162]:



df_merge_Movie['CCCC'] = np.where(df_merge_Movie['Detail3'] == 'Viewed with Video On Demand or PPV', df_merge_Movie['CCCC'].fillna('0003'),df_merge_Movie['CCCC'])


# In[163]:


df_merge_Movie['DDDD']=df_merge_Movie.groupby('Show_name_index').ngroup()


# In[164]:



df_merge_Movie['DDDD']=df_merge_Movie['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[165]:



df_merge_Movie=df_merge_Movie.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[166]:


df_merge_Movie['ORD']=df_merge_Movie['ORD'].astype(str)


# In[167]:


df_merge_Movie['AAAA']=df_merge_Movie["ORD"].str.slice(0,4,1)


# In[168]:


df_merge_Movie['BBBB']=df_merge_Movie["ORD"].str.slice(4,9,1)


# In[169]:


df_merge_Movie=df_merge_Movie.sort_values(['Super','Category'],
               ascending=[True,True],na_position='last')


# In[170]:


df_merge_Movie['AAAA']=df_merge_Movie['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[171]:


df_merge_Movie['AAAA']=df_merge_Movie['AAAA'].replace(r'nan',np.nan, regex=True)


# In[172]:



df_merge_Movie['AAAA']=df_merge_Movie['AAAA'].fillna('1069')


# In[173]:


df_merge_Movie['BBBB']=df_merge_Movie['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[174]:


df_merge_Movie['BBBB']=df_merge_Movie['BBBB'].replace(r'nan',np.nan, regex=True)


# In[175]:


df_merge_Movie['BBBB']=df_merge_Movie['BBBB'].fillna('00157')


# In[176]:


df_merge_Movie['ORD_new'] =df_merge_Movie['AAAA']+df_merge_Movie['BBBB']+df_merge_Movie['CCCC']+df_merge_Movie['DDDD']


# In[177]:


df_merge_Movie.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_Movie.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[178]:


#df_merge_Movie.head(10)


# In[ ]:





# # SPTV1

# In[179]:


Punch_variable=PV['SPTV1']['PunchValue']


# In[180]:


datapv={}
SPTV1=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV1'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV1.append(datapv[i])


# In[181]:


#PV['SPTV1']['PunchValue']


# In[182]:


SPTV1=pd.concat(SPTV1)


# In[183]:


SPTV1.S2022_Client = SPTV1.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[184]:


df_merge_SPTV1= pd.merge(SPTV1, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[185]:


df_merge_SPTV1['LastDigit_PV']=df_merge_SPTV1['S2022_Client'].str.strip().str[-1]


# In[186]:


#df_merge_SPTV1.head(10)


# In[187]:


df_merge_SPTV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV1_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[188]:


df_merge_SPTV1['ORD']=df_merge_SPTV1['ORD'].astype(str)


# In[189]:


df_merge_SPTV1['AAAA']=df_merge_SPTV1["ORD"].str.slice(0,4,1)

df_merge_SPTV1['BBBB']=df_merge_SPTV1["ORD"].str.slice(4,9,1)


# In[190]:


df_merge_SPTV1['CCCC']=df_merge_SPTV1["ORD"].str.slice(9,13,1)


# In[191]:


df_merge_SPTV1=df_merge_SPTV1.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV1['Category']=df_merge_SPTV1['Category'].fillna(method='ffill')
df_merge_SPTV1['QLevel']=df_merge_SPTV1['QLevel'].fillna(method='ffill')
#df_merge_SPTV1['Detail2']=df_merge_SPTV1['Detail2'].fillna(method='ffill')


# In[192]:


df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'nan',np.nan, regex=True)


# In[193]:


df_merge_SPTV1['BBBB']=df_merge_SPTV1['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV1['BBBB']=df_merge_SPTV1['BBBB'].replace(r'nan',np.nan, regex=True)


# In[194]:


df2=df_merge_SPTV1['LastDigit_PV'].groupby(df_merge_SPTV1['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)


# In[195]:


type(df2)


# In[196]:


df2.reset_index(inplace=True)


# In[197]:



df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[198]:


df3=df_merge_SPTV1['Detail3'].groupby(df_merge_SPTV1['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[199]:


df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[200]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df2,on='LastDigit_PV',how='left')


# In[201]:


df_merge_SPTV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV1.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[202]:


df_merge_SPTV1.columns


# In[203]:


df_merge_SPTV1.drop("Detail3_x",axis='columns',inplace=True)


# In[204]:


df_merge_SPTV1=df_merge_SPTV1.rename(columns={"Detail3_y":"Detail3"})


# In[205]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df3,on='Detail3',how='left')


# In[206]:



df_merge_SPTV1=df_merge_SPTV1.rename(columns={"CCCC_y":"CCCC"})


# In[207]:


df_merge_SPTV1=df_merge_SPTV1.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV1['Tmpl']=df_merge_SPTV1['Tmpl'].fillna(method='ffill')
df_merge_SPTV1['Super']=df_merge_SPTV1['Super'].fillna(method='ffill')


# In[208]:



df_merge_SPTV1=df_merge_SPTV1.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV1['Category']=df_merge_SPTV1['Category'].fillna(method='ffill')
df_merge_SPTV1['QLevel']=df_merge_SPTV1['QLevel'].fillna(method='ffill')
#df_merge_SPTV1['Detail2']=df_merge_SPTV1['Detail2'].fillna(method='ffill')


# In[209]:


df_merge_SPTV1=df_merge_SPTV1.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[210]:


df_merge_SPTV1['CCCC']=df_merge_SPTV1['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[211]:


df_merge_SPTV1['DDDD']=df_merge_SPTV1.groupby('Show_name_index').ngroup()


# In[212]:


df_merge_SPTV1['DDDD']=df_merge_SPTV1['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[213]:



df_merge_SPTV1=df_merge_SPTV1.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[214]:


df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV1['AAAA']=df_merge_SPTV1['AAAA'].replace(r'nan',np.nan, regex=True)


# In[215]:


df4=df_merge_SPTV1['List_Heading'].groupby(df_merge_SPTV1['AAAA']).unique().apply(pd.Series)


# In[216]:


#df4


# In[217]:


df4.reset_index(inplace=True)


# In[218]:


df4=pd.DataFrame(df4)


# In[219]:


#df4.List_Heading


# In[220]:


df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[221]:


#df4['List_Heading']


# In[222]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df4,on='List_Heading',how='left')


# In[223]:


#df_merge_SPTV1.columns


# In[224]:


df_merge_SPTV1.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[225]:


df5=df_merge_SPTV1['List_Heading'].groupby(df_merge_SPTV1['BBBB']).unique().apply(pd.Series)


# In[226]:


#df5


# In[227]:


df5.reset_index(inplace=True)


# In[228]:


df5=pd.DataFrame(df5)


# In[229]:


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[230]:


df_merge_SPTV1=pd.merge(df_merge_SPTV1,df5,on='List_Heading',how='left')


# In[231]:


df_merge_SPTV1.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[232]:


#df_merge_SPTV1.columns


# In[233]:


df_merge_SPTV1['ORD_new'] =df_merge_SPTV1['AAAA']+df_merge_SPTV1['BBBB']+df_merge_SPTV1['CCCC']+df_merge_SPTV1['DDDD']


# In[234]:


df_merge_SPTV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV1_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


#  # SPTV2

# In[235]:


Punch_variable=PV['SPTV2']['PunchValue']


# In[236]:


datapv={}
SPTV2=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV2'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV2.append(datapv[i])


# In[237]:


SPTV2=pd.concat(SPTV2)


# In[238]:



SPTV2.S2022_Client = SPTV2.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[239]:


df_merge_SPTV2= pd.merge(SPTV2, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[240]:


df_merge_SPTV2['LastDigit_PV']=df_merge_SPTV2['S2022_Client'].str.strip().str[-1]


# In[241]:


df_merge_SPTV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV2_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[242]:


df_merge_SPTV2['ORD']=df_merge_SPTV2['ORD'].astype(str)


# In[243]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2["ORD"].str.slice(0,4,1)


# In[244]:



df_merge_SPTV2['BBBB']=df_merge_SPTV2["ORD"].str.slice(4,9,1)


# In[245]:


df_merge_SPTV2['CCCC']=df_merge_SPTV2["ORD"].str.slice(9,13,1)


# In[246]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV2['Tmpl']=df_merge_SPTV2['Tmpl'].fillna(method='ffill')
df_merge_SPTV2['Super']=df_merge_SPTV2['Super'].fillna(method='ffill')


# In[247]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV2['Category']=df_merge_SPTV2['Category'].fillna(method='ffill')
df_merge_SPTV2['QLevel']=df_merge_SPTV2['QLevel'].fillna(method='ffill')
df_merge_SPTV2['Detail2']=df_merge_SPTV2['Detail2'].fillna(method='ffill')


# In[248]:



df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)


# In[249]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[250]:


df_merge_SPTV2['BBBB']=df_merge_SPTV2['BBBB'].replace(r'^\s*$', np.nan, regex=True)


# In[251]:


df_merge_SPTV2['BBBB']=df_merge_SPTV2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[252]:


df2=df_merge_SPTV2['LastDigit_PV'].groupby(df_merge_SPTV2['Detail3']).unique().apply(pd.Series)


# In[253]:



df2=pd.DataFrame(df2)


# In[254]:


df2.reset_index(inplace=True)


# In[255]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[256]:


df3=df_merge_SPTV2['Detail3'].groupby(df_merge_SPTV2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[257]:


df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[258]:


df_merge_SPTV2=pd.merge(df_merge_SPTV2,df2,on='LastDigit_PV',how='left')


# In[259]:


df_merge_SPTV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[260]:


df_merge_SPTV2.drop("Detail3_x",axis='columns',inplace=True)


# In[261]:


df_merge_SPTV2=df_merge_SPTV2.rename(columns={"Detail3_y":"Detail3"})


# In[262]:


df_merge_SPTV2=pd.merge(df_merge_SPTV2,df3,on='Detail3',how='left')


# In[263]:


df_merge_SPTV2=df_merge_SPTV2.rename(columns={"CCCC_y":"CCCC"})


# In[264]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

#df_merge_SPTV2['CCCC']=df_merge_SPTV2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[265]:


df_merge_SPTV2['DDDD']=df_merge_SPTV2.groupby('Show_name_index').ngroup()

df_merge_SPTV2['DDDD']=df_merge_SPTV2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[266]:


df_merge_SPTV2=df_merge_SPTV2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[267]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[268]:


df4=df_merge_SPTV2['List_Heading'].groupby(df_merge_SPTV2['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)


# In[269]:


df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[270]:


df_merge_SPTV2=pd.merge(df_merge_SPTV2,df4,on='List_Heading',how='left')


# In[271]:


df_merge_SPTV2.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[272]:


df5=df_merge_SPTV2['List_Heading'].groupby(df_merge_SPTV2['BBBB']).unique().apply(pd.Series)


# In[273]:


df5.reset_index(inplace=True)


# In[274]:


df5=pd.DataFrame(df5)


# In[275]:


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[276]:



df_merge_SPTV2=pd.merge(df_merge_SPTV2,df5,on='List_Heading',how='left')

df_merge_SPTV2.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[277]:


#df_merge_SPTV2.columns


# In[278]:


df_merge_SPTV2['AAAA']=df_merge_SPTV2['AAAA'].astype(str)


# In[279]:


df_merge_SPTV2['BBBB']=df_merge_SPTV2['BBBB'].astype(str)


# In[280]:


df_merge_SPTV2['CCCC']=df_merge_SPTV2['CCCC'].astype(str)


# In[281]:


df_merge_SPTV2['DDDD']=df_merge_SPTV2['DDDD'].astype(str)


# In[282]:


df_merge_SPTV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV2_1.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[283]:


#df_merge_SPTV2.info()


# In[284]:


df_merge_SPTV2['ORD_new'] =df_merge_SPTV2['AAAA']+df_merge_SPTV2['BBBB']+df_merge_SPTV2['CCCC']+df_merge_SPTV2['DDDD']


# # SPTV3

# In[285]:


Punch_variable=PV['SPTV3']['PunchValue']


# In[286]:


datapv={}
SPTV3=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV3'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV3.append(datapv[i])


# In[287]:


SPTV3=pd.concat(SPTV3)


# In[288]:


SPTV3.S2022_Client = SPTV3.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[289]:


df_merge_SPTV3= pd.merge(SPTV3, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[290]:


#df_merge_SPTV3


# In[291]:


df_merge_SPTV3['LastDigit_PV']=df_merge_SPTV3['S2022_Client'].str.strip().str[-1]


# In[292]:


df_merge_SPTV3['ORD']=df_merge_SPTV3['ORD'].astype(str)


# In[293]:



df_merge_SPTV3['AAAA']=df_merge_SPTV3["ORD"].str.slice(0,4,1)

df_merge_SPTV3['BBBB']=df_merge_SPTV3["ORD"].str.slice(4,9,1)

df_merge_SPTV3['CCCC']=df_merge_SPTV3["ORD"].str.slice(9,13,1)


# In[294]:


df_merge_SPTV3=df_merge_SPTV3.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV3['Category']=df_merge_SPTV3['Category'].fillna(method='ffill')
df_merge_SPTV3['QLevel']=df_merge_SPTV3['QLevel'].fillna(method='ffill')
#df_merge_SPTV3['Detail2']=df_merge_SPTV3['Detail2'].fillna(method='ffill')


# In[295]:


df_merge_SPTV3=df_merge_SPTV3.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV3['Tmpl']=df_merge_SPTV3['Tmpl'].fillna(method='ffill')
df_merge_SPTV3['Super']=df_merge_SPTV3['Super'].fillna(method='ffill')


# In[296]:


df_merge_SPTV3['AAAA']=df_merge_SPTV3['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV3['AAAA']=df_merge_SPTV3['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV3['BBBB']=df_merge_SPTV3['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV3['BBBB']=df_merge_SPTV3['BBBB'].replace(r'nan',np.nan, regex=True)


# In[297]:


df2=df_merge_SPTV3['LastDigit_PV'].groupby(df_merge_SPTV3['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[298]:


#df2


# In[299]:


df_merge_SPTV3=pd.merge(df_merge_SPTV3,df2,on='LastDigit_PV',how='left')


# In[300]:


#df_merge_SPTV3


# In[301]:


df_merge_SPTV3.columns


# In[302]:


df_merge_SPTV3.drop("Detail3_x",axis='columns',inplace=True)

df_merge_SPTV3=df_merge_SPTV3.rename(columns={"Detail3_y":"Detail3"})


# In[303]:


df3=df_merge_SPTV3['Detail3'].groupby(df_merge_SPTV3['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[304]:


df3.drop(0,inplace=True)


# In[305]:


df3.dropna(axis=1,inplace=True)


# In[306]:


df_merge_SPTV3=pd.merge(df_merge_SPTV3,df3,on='Detail3',how='left')


# In[307]:


#df3


# In[308]:


df_merge_SPTV3=df_merge_SPTV3.rename(columns={"CCCC_y":"CCCC"})


# In[309]:


#df_merge_SPTV3


# In[310]:


df_merge_SPTV3=df_merge_SPTV3.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV3['CCCC']=df_merge_SPTV3['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[311]:


df_merge_SPTV3['DDDD']=df_merge_SPTV3.groupby('Show_name_index').ngroup()

df_merge_SPTV3['DDDD']=df_merge_SPTV3['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[312]:


#df_merge_SPTV3


# In[313]:


df4=df_merge_SPTV3['List_Heading'].groupby(df_merge_SPTV3['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV3=pd.merge(df_merge_SPTV3,df4,on='List_Heading',how='left')


# In[ ]:





# In[314]:


df_merge_SPTV3.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[315]:


df5=df_merge_SPTV3['List_Heading'].groupby(df_merge_SPTV3['BBBB']).unique().apply(pd.Series)


# In[316]:


df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV3=pd.merge(df_merge_SPTV3,df5,on='List_Heading',how='left')

df_merge_SPTV3.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[317]:


df_merge_SPTV3['ORD_new'] =df_merge_SPTV3['AAAA']+df_merge_SPTV3['BBBB']+df_merge_SPTV3['CCCC']+df_merge_SPTV3['DDDD']


# In[318]:


df_merge_SPTV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV3_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[319]:


#df_merge_SPTV3


# # SPTV4

# In[320]:


Punch_variable=PV['SPTV4']['PunchValue']


# In[321]:


#Punch_variable


# In[322]:


datapv={}
SPTV4=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV4'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV4.append(datapv[i])


# In[323]:


SPTV4=pd.concat(SPTV4)


# In[324]:


SPTV4.S2022_Client = SPTV4.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[325]:


df_merge_SPTV4= pd.merge(SPTV4, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[326]:


df_merge_SPTV4['LastDigit_PV']=df_merge_SPTV4['S2022_Client'].str.strip().str[-1]


# In[327]:


df_merge_SPTV4['ORD']=df_merge_SPTV4['ORD'].astype(str)

df_merge_SPTV4['AAAA']=df_merge_SPTV4["ORD"].str.slice(0,4,1)

df_merge_SPTV4['BBBB']=df_merge_SPTV4["ORD"].str.slice(4,9,1)

df_merge_SPTV4['CCCC']=df_merge_SPTV4["ORD"].str.slice(9,13,1)


# In[328]:


df_merge_SPTV4=df_merge_SPTV4.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV4['Category']=df_merge_SPTV4['Category'].fillna(method='ffill')
df_merge_SPTV4['QLevel']=df_merge_SPTV4['QLevel'].fillna(method='ffill')
df_merge_SPTV4['Detail2']=df_merge_SPTV4['Detail2'].fillna(method='ffill')


# In[329]:


df_merge_SPTV4=df_merge_SPTV4.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV4['Tmpl']=df_merge_SPTV4['Tmpl'].fillna(method='ffill')
df_merge_SPTV4['Super']=df_merge_SPTV4['Super'].fillna(method='ffill')


# In[330]:



df_merge_SPTV4['AAAA']=df_merge_SPTV4['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV4['AAAA']=df_merge_SPTV4['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV4['BBBB']=df_merge_SPTV4['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV4['BBBB']=df_merge_SPTV4['BBBB'].replace(r'nan',np.nan, regex=True)


# In[331]:



df2=df_merge_SPTV4['LastDigit_PV'].groupby(df_merge_SPTV4['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[332]:


df3=df_merge_SPTV4['Detail3'].groupby(df_merge_SPTV4['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)

df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df2,on='LastDigit_PV',how='left')


# In[333]:


df_merge_SPTV4.drop("Detail3_x",axis='columns',inplace=True)

df_merge_SPTV4=df_merge_SPTV4.rename(columns={"Detail3_y":"Detail3"})

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df3,on='Detail3',how='left')


# In[334]:


df_merge_SPTV4=df_merge_SPTV4.rename(columns={"CCCC_y":"CCCC"})


# In[335]:


f_merge_SPTV4=df_merge_SPTV4.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV4['CCCC']=df_merge_SPTV4['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[336]:


df_merge_SPTV4['DDDD']=df_merge_SPTV4.groupby('Show_name_index').ngroup()

df_merge_SPTV4['DDDD']=df_merge_SPTV4['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[337]:


df4=df_merge_SPTV4['List_Heading'].groupby(df_merge_SPTV4['AAAA']).unique().apply(pd.Series)
df4=pd.DataFrame(df4)


df4.reset_index(inplace=True)


df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df4,on='List_Heading',how='left')


# In[338]:


df_merge_SPTV4.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[339]:


df5=df_merge_SPTV4['List_Heading'].groupby(df_merge_SPTV4['BBBB']).unique().apply(pd.Series)


df5=pd.DataFrame(df5)

df5.reset_index(inplace=True)


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV4=pd.merge(df_merge_SPTV4,df5,on='List_Heading',how='left')

df_merge_SPTV4.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[340]:


df_merge_SPTV4['ORD_new'] =df_merge_SPTV4['AAAA']+df_merge_SPTV4['BBBB']+df_merge_SPTV4['CCCC']+df_merge_SPTV4['DDDD']


# In[341]:


df_merge_SPTV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV4_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # SPTV5

# In[342]:


Punch_variable=PV['SPTV5']['PunchValue']


# In[343]:


Punch_variable


# In[344]:


datapv={}
SPTV5=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV5'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV5.append(datapv[i])
	


# In[345]:


SPTV5=pd.concat(SPTV5)


# In[346]:


SPTV5.S2022_Client = SPTV5.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[347]:


df_merge_SPTV5= pd.merge(SPTV5, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')

df_merge_SPTV5['LastDigit_PV']=df_merge_SPTV5['S2022_Client'].str.strip().str[-1]


# In[348]:


df_merge_SPTV5['ORD']=df_merge_SPTV5['ORD'].astype(str)

df_merge_SPTV5['AAAA']=df_merge_SPTV5["ORD"].str.slice(0,4,1)

df_merge_SPTV5['BBBB']=df_merge_SPTV5["ORD"].str.slice(4,9,1)

df_merge_SPTV5['CCCC']=df_merge_SPTV5["ORD"].str.slice(9,13,1)


# In[349]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV5['Category']=df_merge_SPTV5['Category'].fillna(method='ffill')
df_merge_SPTV5['QLevel']=df_merge_SPTV5['QLevel'].fillna(method='ffill')
df_merge_SPTV5['Detail2']=df_merge_SPTV5['Detail2'].fillna(method='ffill')


# In[350]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV5['Tmpl']=df_merge_SPTV5['Tmpl'].fillna(method='ffill')
df_merge_SPTV5['Super']=df_merge_SPTV5['Super'].fillna(method='ffill')


# In[351]:


df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV5['BBBB']=df_merge_SPTV5['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV5['BBBB']=df_merge_SPTV5['BBBB'].replace(r'nan',np.nan, regex=True)


# In[352]:


df2=df_merge_SPTV5['LastDigit_PV'].groupby(df_merge_SPTV5['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df2,on='LastDigit_PV',how='left')


# In[353]:


df_merge_SPTV5.drop("Detail3_x",axis='columns',inplace=True)

df_merge_SPTV5=df_merge_SPTV5.rename(columns={"Detail3_y":"Detail3"})


# In[354]:


#df_merge_SPTV5.columns


# In[355]:


df3=df_merge_SPTV5['Detail3'].groupby(df_merge_SPTV5['CCCC']).unique().apply(pd.Series)
df3=pd.DataFrame(df3)
df3.reset_index(inplace=True)

df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df3,on='Detail3',how='left')

df_merge_SPTV5=df_merge_SPTV5.rename(columns={"CCCC_y":"CCCC"})


# In[356]:


#df_merge_SPTV5.columns


# In[357]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV5['CCCC']=df_merge_SPTV5['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[358]:




df_merge_SPTV5['DDDD']=df_merge_SPTV5.groupby('Show_name_index').ngroup()

df_merge_SPTV5['DDDD']=df_merge_SPTV5['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[359]:


df_merge_SPTV5=df_merge_SPTV5.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV5['AAAA']=df_merge_SPTV5['AAAA'].replace(r'nan',np.nan, regex=True)


# In[360]:


df4=df_merge_SPTV5['List_Heading'].groupby(df_merge_SPTV5['AAAA']).unique().apply(pd.Series)

df4=pd.DataFrame(df4)


df4.reset_index(inplace=True)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df4,on='List_Heading',how='left')

df_merge_SPTV5.rename(columns={'AAAA_x':'AAAA'},inplace=True)


# In[361]:


df5=df_merge_SPTV5['List_Heading'].groupby(df_merge_SPTV5['BBBB']).unique().apply(pd.Series)

df5=pd.DataFrame(df5)


df5.reset_index(inplace=True)


df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV5=pd.merge(df_merge_SPTV5,df5,on='List_Heading',how='left')

df_merge_SPTV5.rename(columns={'BBBB_x':'BBBB'},inplace=True)


# In[362]:



df_merge_SPTV5['ORD_new'] =df_merge_SPTV5['AAAA']+df_merge_SPTV5['BBBB']+df_merge_SPTV5['CCCC']+df_merge_SPTV5['DDDD']


# In[363]:


df_merge_SPTV5['Show_Name'] = df_merge_SPTV5.apply(lambda x: x['List_Heading']+': '+x['Show_Name'] , axis=1)


# In[364]:


df_merge_SPTV5.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV5_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # SPTV5.1

# In[365]:


Punch_variable=PV['SPTV5.1']['PunchValue']


# In[366]:


Punch_variable


# In[367]:


PV['SPTV5.1']['PunchValue']


# In[368]:


data['SPTV5.1'].shape


# In[369]:


datapv={}
SPTV51=[]
for i in Punch_variable:
   
    datapv[i]=data['SPTV5.1'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    SPTV51.append(datapv[i])


# In[370]:


SPTV51=pd.concat(SPTV51)


# In[371]:


SPTV51.S2022_Client = SPTV51.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[372]:


df_merge_SPTV51= pd.merge(SPTV51, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[373]:


df_merge_SPTV51['LastDigit_PV']=df_merge_SPTV51['S2022_Client'].str.strip().str[-1]


# In[374]:


#df_merge_SPTV51


# In[375]:


df_merge_SPTV51['ORD']=df_merge_SPTV51['ORD'].astype(str)

df_merge_SPTV51['AAAA']=df_merge_SPTV51["ORD"].str.slice(0,4,1)

df_merge_SPTV51['BBBB']=df_merge_SPTV51["ORD"].str.slice(4,9,1)

df_merge_SPTV51['CCCC']=df_merge_SPTV51["ORD"].str.slice(9,13,1)


# In[376]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_SPTV51['Category']=df_merge_SPTV51['Category'].fillna(method='ffill')
df_merge_SPTV51['QLevel']=df_merge_SPTV51['QLevel'].fillna(method='ffill')
df_merge_SPTV51['Detail2']=df_merge_SPTV51['Detail2'].fillna(method='ffill')


# In[377]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_SPTV51['Tmpl']=df_merge_SPTV51['Tmpl'].fillna(method='ffill')
df_merge_SPTV51['Super']=df_merge_SPTV51['Super'].fillna(method='ffill')


# In[378]:


#df_merge_SPTV51


# In[379]:


df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_SPTV51['BBBB']=df_merge_SPTV51['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_SPTV51['BBBB']=df_merge_SPTV51['BBBB'].replace(r'nan',np.nan, regex=True)


# In[380]:


df2=df_merge_SPTV51['LastDigit_PV'].groupby(df_merge_SPTV51['Detail3']).unique().apply(pd.Series)




# In[381]:


df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[382]:


df_merge_SPTV51=pd.merge(df_merge_SPTV51,df2,on='LastDigit_PV',how='left')


# In[383]:


#df2


# In[384]:


#df_merge_SPTV51


# In[385]:


#df_merge_SPTV51.drop("Detail3_x",axis='columns',inplace=True)
df_merge_SPTV51['Detail3_y']=df_merge_SPTV51['Detail3_y'].replace(np.nan,'UNKNOWN DETAIL',regex=True)

df_merge_SPTV51=df_merge_SPTV51.rename(columns={"Detail3_y":"Detail3"})


# In[386]:


df3=df_merge_SPTV51['Detail3'].groupby(df_merge_SPTV51['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[387]:


df3.dropna(axis=1,inplace=True)


# In[ ]:





# In[388]:


df3['CCCC']=df3['CCCC'].replace(r'^\s*$', np.nan, regex=True)

df3['CCCC']=df3['CCCC'].replace(r'nan',np.nan, regex=True)


# In[389]:


df3.dropna(inplace=True)


# In[390]:


#df3


# In[391]:


df_merge_SPTV51=pd.merge(df_merge_SPTV51,df3,on='Detail3',how='left')


# In[392]:


df_merge_SPTV51.shape


# In[393]:


df_merge_SPTV51['CCCC_y']=df_merge_SPTV51['CCCC_y'].replace(np.nan,'0001', regex=True)


# In[394]:


df_merge_SPTV51=df_merge_SPTV51.rename(columns={"CCCC_y":"CCCC"})


# In[395]:


#df_merge_SPTV51


# In[396]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_SPTV51['CCCC']=df_merge_SPTV51['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[397]:



df_merge_SPTV51['DDDD']=df_merge_SPTV51.groupby('Show_name_index').ngroup()

df_merge_SPTV51['DDDD']=df_merge_SPTV51['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[398]:


df_merge_SPTV51=df_merge_SPTV51.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_SPTV51['AAAA']=df_merge_SPTV51['AAAA'].replace(r'nan',np.nan, regex=True)


# In[399]:


df4=df_merge_SPTV51['List_Heading'].groupby(df_merge_SPTV51['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_SPTV51=pd.merge(df_merge_SPTV51,df4,on='List_Heading',how='left')

df_merge_SPTV51.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[400]:


df5=df_merge_SPTV51['List_Heading'].groupby(df_merge_SPTV51['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_SPTV51=pd.merge(df_merge_SPTV51,df5,on='List_Heading',how='left')

df_merge_SPTV51.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[401]:


df_merge_SPTV51['ORD_new'] =df_merge_SPTV51['AAAA']+df_merge_SPTV51['BBBB']+df_merge_SPTV51['CCCC']+df_merge_SPTV51['DDDD']


# In[402]:


df_merge_SPTV51.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_SPTV51_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[403]:


#df_merge_SPTV51


# # TV6

# In[404]:


Punch_variable=PV['TV6']['PunchValue']


# In[405]:


datapv={}
TV6=[]
for i in Punch_variable:
   
    datapv[i]=data['TV6'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV6.append(datapv[i])


# In[406]:



TV6=pd.concat(TV6)


# In[407]:


TV6.S2022_Client = TV6.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[408]:



df_merge_TV6= pd.merge(TV6, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[409]:


df_merge_TV6['LastDigit_PV']=df_merge_TV6['S2022_Client'].str.strip().str[-1]


# In[410]:


#df_merge_TV6.head(10)


# In[411]:



df_merge_TV6.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV6_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[412]:


df_merge_TV6['ORD']=df_merge_TV6['ORD'].astype(str)

df_merge_TV6['AAAA']=df_merge_TV6["ORD"].str.slice(0,4,1)

df_merge_TV6['BBBB']=df_merge_TV6["ORD"].str.slice(4,9,1)

df_merge_TV6['CCCC']=df_merge_TV6["ORD"].str.slice(9,13,1)


# In[413]:


df_merge_TV6=df_merge_TV6.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV6['Category']=df_merge_TV6['Category'].fillna(method='ffill')
df_merge_TV6['QLevel']=df_merge_TV6['QLevel'].fillna(method='ffill')
df_merge_TV6['Detail2']=df_merge_TV6['Detail2'].fillna(method='ffill')


# In[414]:


df_merge_TV6=df_merge_TV6.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV6['Tmpl']=df_merge_TV6['Tmpl'].fillna(method='ffill')
df_merge_TV6['Super']=df_merge_TV6['Super'].fillna(method='ffill')


# In[415]:


df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV6['BBBB']=df_merge_TV6['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV6['BBBB']=df_merge_TV6['BBBB'].replace(r'nan',np.nan, regex=True)


# In[416]:



df2=df_merge_TV6['LastDigit_PV'].groupby(df_merge_TV6['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df2,on='LastDigit_PV',how='left')


# In[417]:


#df2


# In[418]:


df_merge_TV6.drop("Detail3_x",axis='columns',inplace=True)

df_merge_TV6=df_merge_TV6.rename(columns={"Detail3_y":"Detail3"})


# In[419]:


df3=df_merge_TV6['Detail3'].groupby(df_merge_TV6['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[420]:


#df3.drop_duplicates(inplace=True)


# In[421]:


df3.drop(0,inplace=True)


# In[422]:


#df3


# In[423]:


df3.dropna(axis=1,inplace=True)


# In[424]:


#df3


# In[425]:



df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df3,on='Detail3',how='left')

df_merge_TV6=df_merge_TV6.rename(columns={"CCCC_y":"CCCC"})


# In[426]:


#df_merge_TV6.LastDigit_PV


# In[427]:


df_merge_TV6=df_merge_TV6.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV6['CCCC']=df_merge_TV6['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[428]:


df_merge_TV6['DDDD']=df_merge_TV6.groupby('Show_name_index').ngroup()

df_merge_TV6['DDDD']=df_merge_TV6['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[429]:


df_merge_TV6=df_merge_TV6.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV6['AAAA']=df_merge_TV6['AAAA'].replace(r'nan',np.nan, regex=True)


# In[430]:


df4=df_merge_TV6['List_Heading'].groupby(df_merge_TV6['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df4,on='List_Heading',how='left')

df_merge_TV6.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[431]:


#df4


# In[432]:


#df_merge_TV6


# In[433]:


df5=df_merge_TV6['List_Heading'].groupby(df_merge_TV6['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df_merge_TV6=pd.merge(df_merge_TV6,df5,on='List_Heading',how='left')

df_merge_TV6.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[434]:



df_merge_TV6['ORD_new'] =df_merge_TV6['AAAA']+df_merge_TV6['BBBB']+df_merge_TV6['CCCC']+df_merge_TV6['DDDD']


# In[435]:


df_merge_TV6.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV6_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV5

# In[436]:


Punch_variable=PV['TV5']['PunchValue']


# In[437]:


#Punch_variable


# In[438]:


datapv={}
TV5=[]
for i in Punch_variable:
   
    datapv[i]=data['TV5'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV5.append(datapv[i])


# In[439]:



TV5=pd.concat(TV5)


# In[440]:



TV5.S2022_Client = TV5.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[441]:


df_merge_TV5= pd.merge(TV5, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[442]:


df_merge_TV5['LastDigit_PV']=df_merge_TV5['S2022_Client'].str.strip().str[-1]


# In[443]:


#df_merge_TV5.head(10)


# In[444]:


df_merge_TV5.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV5_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[445]:


df_merge_TV5['ORD']=df_merge_TV5['ORD'].astype(str)

df_merge_TV5['AAAA']=df_merge_TV5["ORD"].str.slice(0,4,1)

df_merge_TV5['BBBB']=df_merge_TV5["ORD"].str.slice(4,9,1)

df_merge_TV5['CCCC']=df_merge_TV5["ORD"].str.slice(9,13,1)


# In[446]:



df_merge_TV5=df_merge_TV5.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV5['Category']=df_merge_TV5['Category'].fillna(method='ffill')
df_merge_TV5['QLevel']=df_merge_TV5['QLevel'].fillna(method='ffill')
df_merge_TV5['Detail2']=df_merge_TV5['Detail2'].fillna(method='ffill')


# In[447]:


df_merge_TV5=df_merge_TV5.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV5['Tmpl']=df_merge_TV5['Tmpl'].fillna(method='ffill')
df_merge_TV5['Super']=df_merge_TV5['Super'].fillna(method='ffill')


# In[448]:


df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV5['BBBB']=df_merge_TV5['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV5['BBBB']=df_merge_TV5['BBBB'].replace(r'nan',np.nan, regex=True)


# In[449]:


df2=df_merge_TV5['LastDigit_PV'].groupby(df_merge_TV5['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV5=pd.merge(df_merge_TV5,df2,on='LastDigit_PV',how='left')


# In[450]:


#df2


# In[451]:



df_merge_TV5.drop("Detail3_x",axis='columns',inplace=True)

df_merge_TV5=df_merge_TV5.rename(columns={"Detail3_y":"Detail3"})


# In[452]:


df3=df_merge_TV5['Detail3'].groupby(df_merge_TV5['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[453]:


#df3


# In[454]:


df3.drop(0,inplace=True)


# In[455]:


df3.dropna(axis=1,inplace=True)


# In[456]:


#df3


# In[457]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV5=pd.merge(df_merge_TV5,df3,on='Detail3',how='left')

df_merge_TV5=df_merge_TV5.rename(columns={"CCCC_y":"CCCC"})


# In[458]:


df_merge_TV5=df_merge_TV5.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV5['CCCC']=df_merge_TV5['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[459]:


df_merge_TV5['DDDD']=df_merge_TV5.groupby('Show_name_index').ngroup()

df_merge_TV5['DDDD']=df_merge_TV5['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[460]:



df_merge_TV5=df_merge_TV5.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].replace(r'nan',np.nan, regex=True)


# In[461]:


df4=df_merge_TV5['List_Heading'].groupby(df_merge_TV5['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)



# In[462]:


#df4


# In[463]:


#df_merge_TV5=pd.merge(df_merge_TV5,df4,on='List_Heading',how='left')

#df_merge_TV5.rename(columns={'AAAA_y':'AAAA'},inplace=True)
df_merge_TV5['AAAA']=df_merge_TV5['AAAA'].fillna('1237')


# In[464]:


df5=df_merge_TV5['List_Heading'].groupby(df_merge_TV5['BBBB']).unique().apply(pd.Series)

df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[465]:


#df5


# In[466]:


#df_merge_TV5=pd.merge(df_merge_TV5,df5,on='List_Heading',how='left')

#df_merge_TV5.rename(columns={'BBBB_y':'BBBB'},inplace=True)
df_merge_TV5['BBBB']=df_merge_TV5['BBBB'].fillna('00306')


# In[467]:


s=pd.value_counts(df_merge_TV5['Show_Name'])

s1 = pd.Series({'nunique': len(s), 'unique values': s.index.tolist()})
s.append(s1)


# In[468]:


s=pd.value_counts(data['TV5']['Show_Name'])

s1 = pd.Series({'nunique': len(s), 'unique values': s.index.tolist()})
s.append(s1)


# In[469]:



df_merge_TV5['ORD_new'] =df_merge_TV5['AAAA']+df_merge_TV5['BBBB']+df_merge_TV5['CCCC']+df_merge_TV5['DDDD']


# In[470]:


df_merge_TV5.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV5_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV1

# # seperate col1 and col2 Punch values

# In[471]:


pm_col2=pm.query('Columns==2')


# In[472]:


pm_col2=pd.DataFrame(pm_col2)


# In[473]:


pm_col1=pm.query('Columns==1')


# In[474]:


pm_col1=pd.DataFrame(pm_col1)


# In[475]:


PV1={}
grouped1 = pm_col1.groupby('Clean_Type')
for group1 in grouped1.groups.keys():
    PV1[group1] = grouped1.get_group(group1)


# In[476]:


PV2={}
grouped2 = pm_col2.groupby('Clean_Type')
for group2 in grouped2.groups.keys():
    PV2[group2] = grouped2.get_group(group2)


# In[ ]:





# In[477]:


Punch_variable=PV1['TV1']['PunchValue']


# In[478]:


#PV2['TV1']['PunchValue']


# In[479]:


datapv={}
TV1=[]
for i in Punch_variable:
   
    datapv[i]=data['TV1'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV1.append(datapv[i])


# In[480]:


TV1=pd.concat(TV1)


# In[481]:


TV1.S2022_Client = TV1.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[482]:


df_merge_TV1= pd.merge(TV1, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[483]:


df_merge_TV1['LastDigit_PV']=df_merge_TV1['S2022_Client'].str.strip().str[-1]


# In[484]:


#df_merge_TV1.head(10)


# In[485]:


df_merge_TV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[486]:


df_merge_TV1['ORD']=df_merge_TV1['ORD'].astype(str)

df_merge_TV1['AAAA']=df_merge_TV1["ORD"].str.slice(0,4,1)

df_merge_TV1['BBBB']=df_merge_TV1["ORD"].str.slice(4,9,1)

df_merge_TV1['CCCC']=df_merge_TV1["ORD"].str.slice(9,13,1)


# In[487]:


df_merge_TV1=df_merge_TV1.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV1['Category']=df_merge_TV1['Category'].fillna(method='ffill')
df_merge_TV1['QLevel']=df_merge_TV1['QLevel'].fillna(method='ffill')
df_merge_TV1['Detail2']=df_merge_TV1['Detail2'].fillna(method='ffill')


# In[488]:


df_merge_TV1=df_merge_TV1.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV1['Tmpl']=df_merge_TV1['Tmpl'].fillna(method='ffill')
df_merge_TV1['Super']=df_merge_TV1['Super'].fillna(method='ffill')


# In[489]:



df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].replace(r'nan',np.nan, regex=True)


# In[490]:


#df_merge_TV1


# In[491]:


df_merge_TV1=df_merge_TV1.rename(columns={"Detail3_y":"Detail3"})


# In[492]:


df2=df_merge_TV1['LastDigit_PV'].groupby(df_merge_TV1['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[493]:


#df_merge_TV1['LastDigit_PV'].unique()


# In[494]:



df_merge_TV1=pd.merge(df_merge_TV1,df2,on='LastDigit_PV',how='left')


# In[495]:


#df_merge_TV1


# In[496]:


df_merge_TV1 = df_merge_TV1.drop_duplicates(subset='S2022_Client',keep='first')


# In[497]:


df_merge_TV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_det3fill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[498]:


df2


# #Display settings
# pd.set_option('display.max_rows', None)
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width',None)
# pd.set_option('display.colheader_justify', 'center')
# pd.set_option('display.precision', 5)
# pd.set_option('display.max_colwidth', -1)

# In[499]:


#df_merge_TV1.drop("Detail3_x",axis='columns',inplace=True)

df_merge_TV1=df_merge_TV1.rename(columns={"Detail3_y":"Detail3"})


# In[500]:


df_1=df_merge_TV1.groupby('LastDigit_PV')['Detail3','List_Heading'].agg(['unique'])
#g = df.groupby('c')['l1'].unique()
#df_2=df_merge_TV1.groupby('LastDigit_PV')['Detail3','List_Heading'].apply(lambda x: list(np.unique(x)))


# In[501]:


df_1=pd.DataFrame(df_1)


# In[502]:


df_1.reset_index(inplace=True)


# In[503]:


df_1.columns


# In[504]:


df_1.rename(columns={(     'Detail3', 'unique'):('Detail3')},inplace=True)


# In[505]:


df_2=df_1.copy()


# In[506]:


df_2.columns


# #display(df_1)

# In[507]:


df_1.to_excel(r"C:\Users\saraswathy.rajaman\Documents\df_1.xlsx",  header=True,encoding='cp1252')


# In[508]:



df3=df_merge_TV1['Detail3'].groupby(df_merge_TV1['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[509]:


df3


# In[510]:


df3.drop(0,inplace=True)


# In[511]:


df3.dropna(axis=1,inplace=True)


# In[512]:


#df3


# In[513]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV1=pd.merge(df_merge_TV1,df3,on='Detail3',how='left')

#df_merge_TV1['CCCC_y']=df_merge_TV1['CCCC_y'].fillna('0001')


df_merge_TV1=df_merge_TV1.rename(columns={"CCCC_y":"CCCC"})


# In[514]:


#df_merge_TV1['LastDigit_PV'].unique()


# In[515]:


df_merge_TV1=df_merge_TV1.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV1['CCCC']=df_merge_TV1['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV1['DDDD']=df_merge_TV1.groupby('Show_name_index').ngroup()

df_merge_TV1['DDDD']=df_merge_TV1['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[516]:


df_merge_TV1=df_merge_TV1.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'nan',np.nan, regex=True)

df4=df_merge_TV1['List_Heading'].groupby(df_merge_TV1['AAAA']).unique().apply(pd.Series)


# In[517]:


df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[518]:


df4.dropna(axis=1,inplace=True)


# In[519]:


df_merge_TV1=pd.merge(df_merge_TV1,df4,on='List_Heading',how='left')

df_merge_TV1.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[520]:


#df_merge_TV1.LastDigit_PV.unique()


# In[521]:


df5=df_merge_TV1['List_Heading'].groupby(df_merge_TV1['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[522]:


#df5


# In[523]:


df_merge_TV1=pd.merge(df_merge_TV1,df5,on='List_Heading',how='left')


# In[524]:


df_merge_TV1.rename(columns={'BBBB_y':'BBBB'},inplace=True)
df_merge_TV1=df_merge_TV1.sort_values(['Super','Category'],
               ascending=[True,True],na_position='last')
df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'^\s*$', np.nan, regex=True)
df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].replace(r'nan',np.nan, regex=True)
df_merge_TV1['AAAA']=df_merge_TV1['AAAA'].fillna(method='ffill')
df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].replace(r'^\s*$', np.nan, regex=True)
df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].replace(r'nan',np.nan, regex=True)
df_merge_TV1['BBBB']=df_merge_TV1['BBBB'].fillna(method='ffill')

df_merge_TV1['ORD_new'] =df_merge_TV1['AAAA']+df_merge_TV1['BBBB']+df_merge_TV1['CCCC']+df_merge_TV1['DDDD']


# In[525]:


df_merge_TV1['Detail3_x']=df_merge_TV1['Detail3_x'].fillna(df_merge_TV1['Detail3'])


# # TV1 PV2

# In[526]:


Punch_variable=PV2['TV1']['PunchValue']


# In[527]:


Punch_variable


# In[528]:


data2=data['TV1'].copy()


# In[529]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[530]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[531]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[532]:


datapv={}
TV1=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV1.append(datapv[i])


# In[533]:



TV1=pd.concat(TV1)


# In[534]:


TV1.S2022_Client = TV1.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[535]:


df_merge_TV1_col2= pd.merge(TV1, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[536]:


df_merge_TV1_col2['Detail3'].unique()


# In[537]:


df_merge_TV1_col2['LastDigit_PV']=df_merge_TV1_col2['S2022_Client'].str.strip().str[-1]


# In[538]:


df_merge_TV1_col2['ORD']=df_merge_TV1_col2['ORD'].astype(str)

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2["ORD"].str.slice(0,4,1)

df_merge_TV1_col2['BBBB']=df_merge_TV1_col2["ORD"].str.slice(4,9,1)

df_merge_TV1_col2['CCCC']=df_merge_TV1_col2["ORD"].str.slice(9,13,1)


# In[539]:


df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV1_col2['Category']=df_merge_TV1_col2['Category'].fillna(method='ffill')
df_merge_TV1_col2['QLevel']=df_merge_TV1_col2['QLevel'].fillna(method='ffill')
df_merge_TV1_col2['Detail2']=df_merge_TV1_col2['Detail2'].fillna(method='ffill')


# In[540]:


#df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['col2pv','List_Heading','LastDigit_PV','Detail3'], 
               #ascending=[True,True,True,True],na_position='last')
						  

#df_merge_TV2['Super']=df_merge_TV2['Super'].fillna(method='ffill')
#df_merge_TV1_col2['Detail3']=df_merge_TV1_col2['Detail3'].fillna(method='ffill')


# In[541]:


df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV1_col2['Tmpl']=df_merge_TV1_col2['Tmpl'].fillna(method='ffill')
df_merge_TV1_col2['Super']=df_merge_TV1_col2['Super'].fillna(method='ffill')


# In[542]:


df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV1_col2['BBBB']=df_merge_TV1_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1_col2['BBBB']=df_merge_TV1_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[543]:


df2=df_merge_TV1_col2['LastDigit_PV'].groupby(df_merge_TV1_col2['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

#type(df2)


# In[544]:



df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df2,on='LastDigit_PV',how='left')


# In[545]:


df_merge_TV1_col2=df_merge_TV1_col2.rename(columns={"Detail3_y":"Detail3"})


# In[546]:


df_merge_TV1_col2 = df_merge_TV1_col2.drop_duplicates(subset='S2022_Client',keep='first')


# In[547]:


df3=df_merge_TV1_col2['Detail3'].groupby(df_merge_TV1_col2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[548]:


df3.drop(0,inplace=True)


# In[549]:


df3.dropna(axis=1,inplace=True)


# In[550]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df3,on='Detail3',how='left')

#df_merge_TV1_col2['CCCC_y']=df_merge_TV1_col2['CCCC_y'].replace(np.nan,'0001', regex=True)

df_merge_TV1_col2=df_merge_TV1_col2.rename(columns={"CCCC_y":"CCCC"})


# In[551]:


df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV1_col2['CCCC']=df_merge_TV1_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV1_col2['DDDD']=df_merge_TV1_col2.groupby('Show_name_index').ngroup()

df_merge_TV1_col2['DDDD']=df_merge_TV1_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[552]:


df_merge_TV1_col2=df_merge_TV1_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV1_col2['AAAA']=df_merge_TV1_col2['AAAA'].replace(r'nan',np.nan, regex=True)

df4=df_merge_TV1_col2['List_Heading'].groupby(df_merge_TV1_col2['AAAA']).unique().apply(pd.Series)


# In[ ]:





# In[553]:


df4.dropna(axis=1,inplace=True)


# In[554]:


df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[555]:


df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df4,on='List_Heading',how='left')

df_merge_TV1_col2.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[556]:


df5=df_merge_TV1_col2['List_Heading'].groupby(df_merge_TV1_col2['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[557]:


df5.dropna(axis=1,inplace=True)


# In[558]:


df_merge_TV1_col2=pd.merge(df_merge_TV1_col2,df5,on='List_Heading',how='left')


# In[559]:


df_merge_TV1_col2.rename(columns={'BBBB_y':'BBBB'},inplace=True)
df_merge_TV1_col2["AAAA"].fillna(df_merge_TV1_col2["AAAA_x"], inplace=True)
df_merge_TV1_col2["BBBB"].fillna(df_merge_TV1_col2["BBBB_x"], inplace=True)

df_merge_TV1_col2['ORD_new'] =df_merge_TV1_col2['AAAA']+df_merge_TV1_col2['BBBB']+df_merge_TV1_col2['CCCC']+df_merge_TV1_col2['DDDD']


# In[560]:



df_merge_TV1_col2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_col2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[561]:


#df_merge_TV1_col2


# In[562]:


df_merge_TV1=[df_merge_TV1,df_merge_TV1_col2]


# In[563]:


df_merge_TV1=pd.concat(df_merge_TV1)


# In[564]:


#df_merge_TV1


# In[565]:


df_merge_TV1['Detail3_x']=df_merge_TV1['Detail3_x'].fillna(df_merge_TV1['Detail3'])


# In[566]:


df_merge_TV1.drop("Detail3",axis='columns',inplace=True)


# In[567]:


df_merge_TV1=df_merge_TV1.rename(columns={"Detail3_x":"Detail3"})


# In[568]:



df_merge_TV1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV1_all.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV2 

# In[569]:


Punch_variable=PV1['TV2']['PunchValue']


# In[570]:


datapv={}
TV2=[]
for i in Punch_variable:
   
    datapv[i]=data['TV2'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV2.append(datapv[i])


# In[571]:


TV2=pd.concat(TV2)


# In[572]:


TV2.S2022_Client = TV2.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[573]:


df_merge_TV2= pd.merge(TV2, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[574]:


df_merge_TV2['LastDigit_PV']=df_merge_TV2['S2022_Client'].str.strip().str[-1]


# In[575]:


#df_merge_TV2


# In[576]:


df_merge_TV2['ORD']=df_merge_TV2['ORD'].astype(str)

df_merge_TV2['AAAA']=df_merge_TV2["ORD"].str.slice(0,4,1)

df_merge_TV2['BBBB']=df_merge_TV2["ORD"].str.slice(4,9,1)

df_merge_TV2['CCCC']=df_merge_TV2["ORD"].str.slice(9,13,1)


# In[577]:


#df_merge_TV2


# In[579]:



df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_b4fill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[580]:



df_merge_TV2=df_merge_TV2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV2['Category']=df_merge_TV2['Category'].fillna(method='ffill')
df_merge_TV2['QLevel']=df_merge_TV2['QLevel'].fillna(method='ffill')
#df_merge_TV2['Detail2']=df_merge_TV2['Detail2'].fillna(method='ffill')


# In[581]:


g=df_merge_TV2.groupby('List_Heading')


# In[582]:


i=0
n=0
df_merge_TV2_LH={}
for List_Heading, g_df in g:
    #print (List_Heading)
    df_merge_TV2_LH[i]=pd.DataFrame(g_df)
    i=i+1
    n=n+1


# In[583]:


Listheading=df_merge_TV2['List_Heading'].unique()


# In[584]:


n=0
for values in Listheading:
    df_merge_TV2_LH[n]=df_merge_TV2_LH[n].sort_values(['LastDigit_PV'], 
               ascending=[True
                        ])
    df_merge_TV2_LH[n]['Detail2']=df_merge_TV2_LH[n]['Detail2'].fillna(method='ffill')
    n=n+1
#for each values List heading FFill the Detail2 values in each  DF 


# In[585]:


df_merge_TV2_Frames=pd.DataFrame()
df_merge_TV2_Frames = df_merge_TV2_Frames.append([df_merge_TV2_LH[i] for i in range(n)])


# In[586]:


df_merge_TV2_Frames.shape


# In[587]:


df_merge_TV2=df_merge_TV2_Frames.copy()


# In[588]:


df_merge_TV2=df_merge_TV2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV2['Tmpl']=df_merge_TV2['Tmpl'].fillna(method='ffill')
df_merge_TV2['Super']=df_merge_TV2['Super'].fillna(method='ffill')


# In[589]:


df_merge_TV2.head(100)


# In[590]:



df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV2['BBBB']=df_merge_TV2['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2['BBBB']=df_merge_TV2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[591]:


df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_1", sep='\t', index=False,header=True,encoding='cp1252')


# In[592]:


df2=df_merge_TV2['LastDigit_PV'].groupby(df_merge_TV2['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)

df2.reset_index(inplace=True)


# In[593]:


df2


# In[594]:


df2.rename(columns={0:"LastDigit_PV"},inplace=True)


# In[595]:


df2_1=df_merge_TV2[['LastDigit_PV','Detail3','List_Heading']]


# In[596]:


#df2_1=df2_1.dropna(axis=0,inplace=True)


# In[597]:


df2_1


# In[598]:


df2_1.dropna(inplace=True)


# In[599]:


df2_1.drop_duplicates().reset_index(drop=True)


# In[600]:


df2_1.drop_duplicates().reset_index(drop=True)


# In[601]:


df_merge_TV2_copy=df_merge_TV2.copy()


# In[602]:


df_merge_TV2_copy_1=pd.merge(df_merge_TV2_copy,df2_1,on=['LastDigit_PV','Detail3','List_Heading'],how='left')


# In[603]:


df_merge_TV2_copy_1.head(100)


# In[604]:


#df_merge_TV2_copy_1
df_merge_TV2_copy_1.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_copy_1.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[605]:


#df2['Detail3']=df2['Detail3'].fillna('None')


# In[606]:


df2_1.drop_duplicates().reset_index(drop=True)


# In[607]:


#df2_1


# In[608]:


df2


# In[609]:





df_merge_TV2=pd.merge(df_merge_TV2,df2,on='LastDigit_PV',how='left')


# In[610]:


#df_merge_TV2


# In[611]:


#df_merge_TV2.drop("Detail3_x",axis='columns',inplace=True)

df_merge_TV2=df_merge_TV2.rename(columns={"Detail3_y":"Detail3"})


# In[612]:


df_merge_TV2 = df_merge_TV2.drop_duplicates(subset='S2022_Client',keep='first')


# In[613]:


df3=df_merge_TV2['Detail3'].groupby(df_merge_TV2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)

#df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)



# In[614]:


df3


# In[615]:


df3.drop(0,inplace=True)


# In[616]:


df3.dropna(axis=1,inplace=True)


# In[617]:


#df3


# In[618]:


df_merge_TV2=pd.merge(df_merge_TV2,df3,on='Detail3',how='left')

df_merge_TV2['CCCC_y']=df_merge_TV2['CCCC_y'].fillna(df_merge_TV2['CCCC_x'])

df_merge_TV2=df_merge_TV2.rename(columns={"CCCC_y":"CCCC"})


# In[619]:


df_merge_TV2=df_merge_TV2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV2['CCCC']=df_merge_TV2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2['DDDD']=df_merge_TV2.groupby('Show_name_index').ngroup()

df_merge_TV2['DDDD']=df_merge_TV2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))

df_merge_TV2=df_merge_TV2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2['AAAA']=df_merge_TV2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[620]:


df4=df_merge_TV2['List_Heading'].groupby(df_merge_TV2['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)

df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)


# In[621]:


#df4


# In[622]:


df4.dropna(axis=1,inplace=True)


# In[623]:


df_merge_TV2=pd.merge(df_merge_TV2,df4,on='List_Heading',how='left')

df_merge_TV2.rename(columns={'AAAA_y':'AAAA'},inplace=True)


# In[624]:


df5=df_merge_TV2['List_Heading'].groupby(df_merge_TV2['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)


# In[625]:


#df5


# In[626]:


df5.dropna(axis=1,inplace=True)


# In[627]:


df_merge_TV2=pd.merge(df_merge_TV2,df5,on='List_Heading',how='left')

df_merge_TV2.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[628]:


df_merge_TV2 = df_merge_TV2.drop_duplicates(subset='S2022_Client',keep='first')


# In[629]:


df_merge_TV2['ORD_new'] =df_merge_TV2['AAAA']+df_merge_TV2['BBBB']+df_merge_TV2['CCCC']+df_merge_TV2['DDDD']


# In[630]:



df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV2 col2

# In[631]:



Punch_variable=PV2['TV2']['PunchValue']
#Punch_variable=Punch_variable.lower()
data2=data['TV2'].copy()


# In[632]:


#Punch_variable


# In[633]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[634]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[635]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[636]:


datapv={}
TV2=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV2.append(datapv[i])


# In[637]:


TV2=pd.concat(TV2)


# In[638]:



TV2.S2022_Client = TV2.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[639]:


df_merge_TV2_col2= pd.merge(TV2, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[640]:


df_merge_TV2_col2['Detail3'].unique()


# In[641]:


df_merge_TV2_col2['LastDigit_PV']=df_merge_TV2_col2['S2022_Client'].str.strip().str[-1]


# In[642]:


df_merge_TV2_col2['ORD']=df_merge_TV2_col2['ORD'].astype(str)

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2["ORD"].str.slice(0,4,1)

df_merge_TV2_col2['BBBB']=df_merge_TV2_col2["ORD"].str.slice(4,9,1)

df_merge_TV2_col2['CCCC']=df_merge_TV2_col2["ORD"].str.slice(9,13,1)


# In[643]:


df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV2_col2['Category']=df_merge_TV2_col2['Category'].fillna(method='ffill')
df_merge_TV2_col2['QLevel']=df_merge_TV2_col2['QLevel'].fillna(method='ffill')
#df_merge_TV2_col2['Detail2']=df_merge_TV2_col2['Detail2'].fillna(method='ffill')


# In[644]:



Listheading=df_merge_TV2_col2['List_Heading'].unique()


# In[645]:


g=df_merge_TV2_col2.groupby('List_Heading')


# In[646]:


i=0
n=0
df_merge_TV2_LH={}
for List_Heading, g_df in g:
    #print (List_Heading)
    df_merge_TV2_LH[i]=pd.DataFrame(g_df)
    i=i+1
    n=n+1


# In[647]:


n=0

for values in Listheading:
    df_merge_TV2_LH[n]=df_merge_TV2_LH[n].sort_values(['LastDigit_PV'], 
               ascending=[True
                        ])
    df_merge_TV2_LH[n]['Detail2']=df_merge_TV2_LH[n]['Detail2'].fillna(method='ffill')
    n=n+1
#for each values List heading FFill the Detail2 values in each  DF 


# In[648]:



df_merge_TV2_Frames=pd.DataFrame()
df_merge_TV2_Frames = df_merge_TV2_Frames.append([df_merge_TV2_LH[i] for i in range(n)])


# In[649]:



df_merge_TV2_col2=df_merge_TV2_Frames.copy()


# In[650]:


df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV2_col2['Tmpl']=df_merge_TV2_col2['Tmpl'].fillna(method='ffill')
df_merge_TV2_col2['Super']=df_merge_TV2_col2['Super'].fillna(method='ffill')


# In[651]:


df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV2_col2['BBBB']=df_merge_TV2_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2_col2['BBBB']=df_merge_TV2_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[652]:


df2=df_merge_TV2_col2['LastDigit_PV'].groupby(df_merge_TV2_col2['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

#type(df2)


# In[653]:


df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df2,on='LastDigit_PV',how='left')

df_merge_TV2_col2=df_merge_TV2_col2.rename(columns={"Detail3_y":"Detail3"})


# In[654]:


df_merge_TV2_col2 = df_merge_TV2_col2.drop_duplicates(subset='S2022_Client',keep='first')

df3=df_merge_TV2_col2['Detail3'].groupby(df_merge_TV2_col2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[655]:


df3

df3.dropna(axis=1,inplace=True)

df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)



# In[656]:


#df3


# In[657]:


df3.drop(0,inplace=True)


# In[658]:


df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df3,on='Detail3',how='left')


# In[659]:


df_merge_TV2_col2=df_merge_TV2_col2.rename(columns={"CCCC_y":"CCCC"})

df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV2_col2['CCCC']=df_merge_TV2_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV2_col2['DDDD']=df_merge_TV2_col2.groupby('Show_name_index').ngroup()

df_merge_TV2_col2['DDDD']=df_merge_TV2_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[660]:


df_merge_TV2_col2=df_merge_TV2_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV2_col2['AAAA']=df_merge_TV2_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[661]:


df4=df_merge_TV2_col2['List_Heading'].groupby(df_merge_TV2_col2['AAAA']).unique().apply(pd.Series)

df4.dropna(axis=1,inplace=True)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)


# In[662]:


#df4


# In[663]:


df4.rename(columns={'AAAA':'AAAA',0:'List_Heading'},inplace=True)

df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df4,on='List_Heading',how='left')


# In[664]:



df_merge_TV2_col2.rename(columns={'AAAA_y':'AAAA'},inplace=True)

df5=df_merge_TV2_col2['List_Heading'].groupby(df_merge_TV2_col2['BBBB']).unique().apply(pd.Series)


# In[665]:


df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)

df5.rename(columns={'BBBB':'BBBB',0:'List_Heading'},inplace=True)

df5.dropna(axis=1,inplace=True)

df_merge_TV2_col2=pd.merge(df_merge_TV2_col2,df5,on='List_Heading',how='left')


# In[666]:


df_merge_TV2_col2.rename(columns={'BBBB_y':'BBBB'},inplace=True)


# In[667]:


df_merge_TV2_col2["AAAA"].fillna(df_merge_TV2_col2["AAAA_x"], inplace=True)
df_merge_TV2_col2["BBBB"].fillna(df_merge_TV2_col2["BBBB_x"], inplace=True)


# In[668]:


df_merge_TV2_col2['ORD_new'] =df_merge_TV2_col2['AAAA']+df_merge_TV2_col2['BBBB']+df_merge_TV2_col2['CCCC']+df_merge_TV2_col2['DDDD']


# In[669]:


df_merge_TV2=[df_merge_TV2,df_merge_TV2_col2]


# In[670]:


type(df_merge_TV2)


# In[671]:


#df_merge_TV2=pd.DataFrame(df_merge_TV2)


# In[672]:


df_merge_TV2=pd.concat(df_merge_TV2)


# In[673]:


df_merge_TV2['Detail3_x']=df_merge_TV2['Detail3_x'].fillna(df_merge_TV2['Detail3'])


# In[674]:


df_merge_TV2.drop("Detail3",axis='columns',inplace=True)


# In[675]:


df_merge_TV2=df_merge_TV2.rename(columns={"Detail3_x":"Detail3"})


# In[676]:


df_merge_TV2 = df_merge_TV2.drop_duplicates(subset='S2022_Client',keep='first')


# In[677]:


df_merge_TV2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV3

# In[678]:


Punch_variable=PV1['TV3']['PunchValue']


# In[679]:


datapv={}
TV3=[]
for i in Punch_variable:
   
    datapv[i]=data['TV3'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV3.append(datapv[i])


# In[680]:



TV3=pd.concat(TV3)


# In[681]:


TV3.S2022_Client = TV3.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[682]:


df_merge_TV3= pd.merge(TV3, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')

df_merge_TV3['LastDigit_PV']=df_merge_TV3['S2022_Client'].str.strip().str[-1]


# In[683]:



df_merge_TV3.head(10)


# In[684]:


df_merge_TV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[685]:


df_merge_TV3['ORD']=df_merge_TV3['ORD'].astype(str)

df_merge_TV3['AAAA']=df_merge_TV3["ORD"].str.slice(0,4,1)

df_merge_TV3['BBBB']=df_merge_TV3["ORD"].str.slice(4,9,1)

df_merge_TV3['CCCC']=df_merge_TV3["ORD"].str.slice(9,13,1)


# In[686]:



df_merge_TV3=df_merge_TV3.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV3['Category']=df_merge_TV3['Category'].fillna(method='ffill')
df_merge_TV3['QLevel']=df_merge_TV3['QLevel'].fillna(method='ffill')
df_merge_TV3['Detail2']=df_merge_TV3['Detail2'].fillna(method='ffill')


# In[687]:


df_merge_TV3=df_merge_TV3.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV3['Tmpl']=df_merge_TV3['Tmpl'].fillna(method='ffill')
df_merge_TV3['Super']=df_merge_TV3['Super'].fillna(method='ffill')


# In[688]:


df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].replace(r'nan',np.nan, regex=True)


# In[689]:



df2=df_merge_TV3['LastDigit_PV'].groupby(df_merge_TV3['Detail3']).unique().apply(pd.Series)


# In[690]:


df2=pd.DataFrame(df2)


# In[691]:


df2


# In[692]:


df2.reset_index(inplace=True)


# In[693]:



df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[694]:


df_merge_TV3=pd.merge(df_merge_TV3,df2,on='LastDigit_PV',how='left')


# In[695]:


df_merge_TV3=df_merge_TV3.rename(columns={"Detail3_y":"Detail3"})


# In[696]:


df3=df_merge_TV3['Detail3'].groupby(df_merge_TV3['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[697]:



#df3.dropna()
df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[698]:


df3


# In[699]:


df_merge_TV3=pd.merge(df_merge_TV3,df3,on='Detail3',how='left')


# In[700]:


df_merge_TV3=df_merge_TV3.rename(columns={"CCCC_y":"CCCC"})


# In[701]:


df_merge_TV3=df_merge_TV3.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV3['CCCC']=df_merge_TV3['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[702]:



df_merge_TV3['DDDD']=df_merge_TV3.groupby('Show_name_index').ngroup()

df_merge_TV3['DDDD']=df_merge_TV3['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[703]:


df_merge_TV3=df_merge_TV3.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].replace(r'nan',np.nan, regex=True)


# In[704]:


df4=df_merge_TV3['List_Heading'].groupby(df_merge_TV3['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)


# In[705]:


df4


# In[706]:


df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].fillna('1246')


# In[707]:


df5=df_merge_TV3['List_Heading'].groupby(df_merge_TV3['BBBB']).unique().apply(pd.Series)



df5.reset_index(inplace=True)

df5=pd.DataFrame(df5)


# In[708]:


df5


# In[709]:


df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].fillna('00315')


# In[ ]:





# In[710]:


df_merge_TV3['ORD_new'] =df_merge_TV3['AAAA']+df_merge_TV3['BBBB']+df_merge_TV3['CCCC']+df_merge_TV3['DDDD']


# In[711]:


df_merge_TV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV3 Col2

# In[712]:


Punch_variable=PV2['TV3']['PunchValue']
data2=data['TV3'].copy()


# In[713]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[714]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[715]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[716]:


datapv={}
TV3=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV3.append(datapv[i])


# In[717]:


TV3=pd.concat(TV3)


# In[718]:


TV3.S2022_Client = TV3.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[719]:


df_merge_TV3_col2= pd.merge(TV3, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[720]:



df_merge_TV3_col2['Detail3'].unique()


# In[721]:


df_merge_TV3_col2['LastDigit_PV']=df_merge_TV3_col2['S2022_Client'].str.strip().str[-1]


# In[722]:


df_merge_TV3_col2['ORD']=df_merge_TV3_col2['ORD'].astype(str)


# In[723]:



df_merge_TV3_col2['AAAA']=df_merge_TV3_col2["ORD"].str.slice(0,4,1)

df_merge_TV3_col2['BBBB']=df_merge_TV3_col2["ORD"].str.slice(4,9,1)

df_merge_TV3_col2['CCCC']=df_merge_TV3_col2["ORD"].str.slice(9,13,1)


# In[724]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV3_col2['Category']=df_merge_TV3_col2['Category'].fillna(method='ffill')
df_merge_TV3_col2['QLevel']=df_merge_TV3_col2['QLevel'].fillna(method='ffill')
df_merge_TV3_col2['Detail2']=df_merge_TV3_col2['Detail2'].fillna(method='ffill')


# In[725]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV3_col2['Tmpl']=df_merge_TV3_col2['Tmpl'].fillna(method='ffill')
df_merge_TV3_col2['Super']=df_merge_TV3_col2['Super'].fillna(method='ffill')


# In[726]:


df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[727]:


df_merge_TV3_col2['BBBB']=df_merge_TV3_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3_col2['BBBB']=df_merge_TV3_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[728]:


df2=df_merge_TV3_col2['LastDigit_PV'].groupby(df_merge_TV3_col2['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)

type(df2)


# In[729]:


df2


# In[730]:



df2.reset_index(inplace=True)


# In[731]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)

df_merge_TV3_col2=pd.merge(df_merge_TV3_col2,df2,on='LastDigit_PV',how='left')


# In[732]:



df_merge_TV3_col2=df_merge_TV3_col2.rename(columns={"Detail3_y":"Detail3"})


# In[733]:


#df_merge_TV3_col2 = df_merge_TV3_col2.drop_duplicates(subset='S2022_Client',keep='first')


# In[734]:


df3=df_merge_TV3_col2['Detail3'].groupby(df_merge_TV3_col2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[735]:


df3


# In[736]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[737]:



df_merge_TV3_col2=pd.merge(df_merge_TV3_col2,df3,on='Detail3',how='left')


# In[738]:


df_merge_TV3_col2=df_merge_TV3_col2.rename(columns={"CCCC_y":"CCCC"})


# In[739]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[740]:


df_merge_TV3_col2['CCCC']=df_merge_TV3_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[741]:


df_merge_TV3_col2['DDDD']=df_merge_TV3_col2.groupby('Show_name_index').ngroup()

df_merge_TV3_col2['DDDD']=df_merge_TV3_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[742]:


df_merge_TV3_col2=df_merge_TV3_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[743]:


df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV3_col2['AAAA']=df_merge_TV3_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[744]:


df4=df_merge_TV3_col2['List_Heading'].groupby(df_merge_TV3_col2['AAAA']).unique().apply(pd.Series)


# In[745]:


df4


# In[746]:


df_merge_TV3['AAAA']=df_merge_TV3['AAAA'].fillna('1247')


# In[747]:


df5=df_merge_TV3_col2['List_Heading'].groupby(df_merge_TV3_col2['BBBB']).unique().apply(pd.Series)


# In[748]:


df_merge_TV3['BBBB']=df_merge_TV3['BBBB'].fillna('00316')


# In[749]:


df_merge_TV3_col2['ORD_new'] =df_merge_TV3_col2['AAAA']+df_merge_TV3_col2['BBBB']+df_merge_TV3_col2['CCCC']+df_merge_TV3_col2['DDDD']


# In[750]:


df_merge_TV3_col2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3_col2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[751]:


df_merge_TV3=[df_merge_TV3,df_merge_TV3_col2]

df_merge_TV3=pd.concat(df_merge_TV3)


# In[752]:


df_merge_TV3['Detail3_x']=df_merge_TV3['Detail3_x'].fillna(df_merge_TV3['Detail3'])


# In[753]:


df_merge_TV3.drop("Detail3",axis='columns',inplace=True)


# In[754]:


df_merge_TV3=df_merge_TV3.rename(columns={"Detail3_x":"Detail3"})


# In[755]:


df_merge_TV3.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV3.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[756]:


df_merge_TV3.head(100)


# # TV4 

# In[757]:


Punch_variable=PV1['TV4']['PunchValue']


# In[758]:


datapv={}
TV4=[]
for i in Punch_variable:
   
    datapv[i]=data['TV4'].copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    TV4.append(datapv[i])


# In[759]:



TV4=pd.concat(TV4)


# In[760]:



TV4.S2022_Client = TV4.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[761]:



df_merge_TV4= pd.merge(TV4, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[762]:


df_merge_TV4['LastDigit_PV']=df_merge_TV4['S2022_Client'].str.strip().str[-1]


# In[763]:


df_merge_TV4.head(10)


# In[764]:


df_merge_TV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4_beforefill.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[765]:


df_merge_TV4['ORD']=df_merge_TV4['ORD'].astype(str)

df_merge_TV4['AAAA']=df_merge_TV4["ORD"].str.slice(0,4,1)

df_merge_TV4['BBBB']=df_merge_TV4["ORD"].str.slice(4,9,1)

df_merge_TV4['CCCC']=df_merge_TV4["ORD"].str.slice(9,13,1)


# In[766]:


df_merge_TV4=df_merge_TV4.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV4['Category']=df_merge_TV4['Category'].fillna(method='ffill')
df_merge_TV4['QLevel']=df_merge_TV4['QLevel'].fillna(method='ffill')
df_merge_TV4['Detail2']=df_merge_TV4['Detail2'].fillna(method='ffill')


# In[767]:


df_merge_TV4=df_merge_TV4.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV4['Tmpl']=df_merge_TV4['Tmpl'].fillna(method='ffill')
df_merge_TV4['Super']=df_merge_TV4['Super'].fillna(method='ffill')


# In[768]:


df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV4['BBBB']=df_merge_TV4['BBBB'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV4['BBBB']=df_merge_TV4['BBBB'].replace(r'nan',np.nan, regex=True)


# In[769]:


df2=df_merge_TV4['LastDigit_PV'].groupby(df_merge_TV4['Detail3']).unique().apply(pd.Series)

df2=pd.DataFrame(df2)


# In[770]:


df2


# In[771]:


df2.reset_index(inplace=True)


# In[772]:


df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[773]:


df_merge_TV4=pd.merge(df_merge_TV4,df2,on='LastDigit_PV',how='left')


# In[774]:



df_merge_TV4=df_merge_TV4.rename(columns={"Detail3_y":"Detail3"})


# In[775]:


df3=df_merge_TV4['Detail3'].groupby(df_merge_TV4['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[776]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)


# In[777]:


#df3


# In[778]:


#df3.dropna(axis=1,inplace=True)


# In[779]:



df_merge_TV4=pd.merge(df_merge_TV4,df3,on='Detail3',how='left')

df_merge_TV4=df_merge_TV4.rename(columns={"CCCC_y":"CCCC"})


# In[780]:


df_merge_TV4=df_merge_TV4.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')

df_merge_TV4['CCCC']=df_merge_TV4['CCCC'].replace(r'^\s*$', np.nan, regex=True)


# In[781]:


df_merge_TV4['DDDD']=df_merge_TV4.groupby('Show_name_index').ngroup()

df_merge_TV4['DDDD']=df_merge_TV4['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


# In[782]:


df_merge_TV4=df_merge_TV4.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')

df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].replace(r'nan',np.nan, regex=True)


# In[783]:


df4=df_merge_TV4['List_Heading'].groupby(df_merge_TV4['AAAA']).unique().apply(pd.Series)

df4.reset_index(inplace=True)

df4=pd.DataFrame(df4)


# In[784]:


#df4


# In[785]:


df_merge_TV4['AAAA']=df_merge_TV4['AAAA'].fillna('1235')


# In[786]:


df5=df_merge_TV4['List_Heading'].groupby(df_merge_TV4['BBBB']).unique().apply(pd.Series)


# In[787]:


#df5


# In[788]:


df_merge_TV4['BBBB']=df_merge_TV4['BBBB'].fillna('00304')


# In[789]:


df_merge_TV4['ORD_new'] =df_merge_TV4['AAAA']+df_merge_TV4['BBBB']+df_merge_TV4['CCCC']+df_merge_TV4['DDDD']


# In[790]:


df_merge_TV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4_ordfillcheck.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # TV4 col2

# In[791]:


Punch_variable=PV2['TV4']['PunchValue']
data2=data['TV4'].copy()


# In[792]:


data2['S2022_Client']=data2['S2022_Client'].apply(pd.to_numeric)
data2['F2021_Client']=data2['F2021_Client'].apply(pd.to_numeric)


# In[793]:


data2['S2022_Client']=data2['S2022_Client']+1
data2['F2021_Client']=data2['F2021_Client']+1


# In[794]:


data2['S2022_Client']=data2['S2022_Client'].astype(str)
data2['F2021_Client']=data2['F2021_Client'].astype(str)


# In[795]:


datapv={}
TV4=[]
for i in Punch_variable:
   
    datapv[i]=data2.copy()

    datapv[i]['S2022_Client']=datapv[i].apply(lambda x:str(x['S2022_Client'])+str(i), axis=1)
    datapv[i]['F2021_Client']=datapv[i].apply(lambda x:str(x['F2021_Client'])+str(i), axis=1)
    
    datapv[i]['col2pv']='yes'
    
    TV4.append(datapv[i])


# In[796]:


TV4=pd.concat(TV4)


# In[797]:


TV4.S2022_Client = TV4.S2022_Client.str.strip()
df_dict.CCP = df_dict.CCP.str.strip()


# In[798]:


df_merge_TV4_col2= pd.merge(TV4, df_dict, left_on=['S2022_Client'], right_on=['CCP'],how='left')


# In[799]:


#df_merge_TV4_col2['Detail3'].unique()


# In[800]:


df_merge_TV4_col2['LastDigit_PV']=df_merge_TV4_col2['S2022_Client'].str.strip().str[-1]


# In[801]:


df_merge_TV4_col2['ORD']=df_merge_TV4_col2['ORD'].astype(str)

df_merge_TV4_col2['AAAA']=df_merge_TV4_col2["ORD"].str.slice(0,4,1)

df_merge_TV4_col2['BBBB']=df_merge_TV4_col2["ORD"].str.slice(4,9,1)

df_merge_TV4_col2['CCCC']=df_merge_TV4_col2["ORD"].str.slice(9,13,1)


# In[802]:


df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['List_Heading'], 
               ascending=[True],na_position='last')
df_merge_TV4_col2['Category']=df_merge_TV4_col2['Category'].fillna(method='ffill')
df_merge_TV4_col2['QLevel']=df_merge_TV4_col2['QLevel'].fillna(method='ffill')
df_merge_TV4_col2['Detail2']=df_merge_TV4_col2['Detail2'].fillna(method='ffill')


# In[803]:



df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['LastDigit_PV','ORD'],ascending=[True,True],na_position='last')
df_merge_TV4_col2['Tmpl']=df_merge_TV4_col2['Tmpl'].fillna(method='ffill')
df_merge_TV4_col2['Super']=df_merge_TV4_col2['Super'].fillna(method='ffill')


# In[804]:


df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'nan',np.nan, regex=True)

df_merge_TV4_col2['BBBB']=df_merge_TV4_col2['BBBB'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4_col2['BBBB']=df_merge_TV4_col2['BBBB'].replace(r'nan',np.nan, regex=True)


# In[805]:


df2=df_merge_TV4_col2['LastDigit_PV'].groupby(df_merge_TV4_col2['Detail3']).unique().apply(pd.Series)


# In[806]:


#df2


# In[807]:



df2=pd.DataFrame(df2)


# In[808]:



df2.reset_index(inplace=True)

df2.rename(columns={"Detail3":"Detail3",0:"LastDigit_PV"},inplace=True)


# In[809]:



df_merge_TV4_col2=pd.merge(df_merge_TV4_col2,df2,on='LastDigit_PV',how='left')

df_merge_TV4_col2=df_merge_TV4_col2.rename(columns={"Detail3_y":"Detail3"})


# In[810]:


df3=df_merge_TV4_col2['Detail3'].groupby(df_merge_TV4_col2['CCCC']).unique().apply(pd.Series)
df3.reset_index(inplace=True)


# In[811]:


#df3


# In[812]:


df3.rename(columns={"CCCC":"CCCC",0:"Detail3"},inplace=True)

df_merge_TV4_col2=pd.merge(df_merge_TV4_col2,df3,on='Detail3',how='left')


# In[813]:



df_merge_TV4_col2=df_merge_TV4_col2.rename(columns={"CCCC_y":"CCCC"})

df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['Category','Detail3','Show_name_index','LastDigit_PV'],ascending=[True,True,True,True],na_position='last')


# In[814]:



df_merge_TV4_col2['CCCC']=df_merge_TV4_col2['CCCC'].replace(r'^\s*$', np.nan, regex=True)


df_merge_TV4_col2['DDDD']=df_merge_TV4_col2.groupby('Show_name_index').ngroup()


# In[815]:


df_merge_TV4_col2['DDDD']=df_merge_TV4_col2['DDDD'].apply(lambda x: '{0:0>7}'.format(x))


df_merge_TV4_col2=df_merge_TV4_col2.sort_values(['Category','Show_Name'],ascending=[True,True],na_position='last')


# In[816]:


df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'^\s*$', np.nan, regex=True)

df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].replace(r'nan',np.nan, regex=True)


# In[817]:



df4=df_merge_TV4_col2['List_Heading'].groupby(df_merge_TV4_col2['AAAA']).unique().apply(pd.Series)

#df4.dropna(axis=1,inplace=True)

df4.reset_index(inplace=True)


# In[818]:


#df4


# In[819]:


df_merge_TV4_col2['AAAA']=df_merge_TV4_col2['AAAA'].fillna('1236')


# In[820]:


df5=df_merge_TV4_col2['List_Heading'].groupby(df_merge_TV4_col2['BBBB']).unique().apply(pd.Series)


# In[821]:


#df5


# In[822]:


df_merge_TV4_col2['BBBB']=df_merge_TV4_col2['BBBB'].fillna('00305')


# In[823]:



df_merge_TV4_col2['ORD_new'] =df_merge_TV4_col2['AAAA']+df_merge_TV4_col2['BBBB']+df_merge_TV4_col2['CCCC']+df_merge_TV4_col2['DDDD']


# In[824]:


df_merge_TV4_col2.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4_col2.txt", sep='\t', index=False,header=True,encoding='cp1252')


# In[825]:



df_merge_TV4=[df_merge_TV4,df_merge_TV4_col2]


# In[826]:


df_merge_TV4=pd.concat(df_merge_TV4)


# In[827]:



df_merge_TV4['Detail3_x']=df_merge_TV4['Detail3_x'].fillna(df_merge_TV4['Detail3'])


# In[828]:


df_merge_TV4.drop("Detail3",axis='columns',inplace=True)


# In[829]:


df_merge_TV4=df_merge_TV4.rename(columns={"Detail3_x":"Detail3"})


# In[830]:


df_merge_TV4.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_merge_TV4.txt", sep='\t', index=False,header=True,encoding='cp1252')


# # concat DF

# In[852]:


df_all=[df_merge_TV1,
        df_merge_TV2,
        df_merge_TV3,
        df_merge_TV4,
        df_merge_TV5,
        df_merge_TV6,
        df_merge_SPTV1,
        df_merge_SPTV2,
        df_merge_SPTV3,
        df_merge_SPTV4,
        df_merge_SPTV5,
        df_merge_SPTV51,
        df_merge_Movie,
        df_merge_cable,      
        df_merge_add_cab]


# In[853]:


df_all=pd.concat(df_all)


# In[854]:


#df_all.columns


# In[855]:


df_all.drop(['StatisticID', 'CatSynID','NoteID','statusid'], axis=1, inplace=True)


# In[856]:


df_all['EditedBy']='codebookcreator'
df_all['EditedDate']=pd.to_datetime('today')
df_all['StudyEntryID']=451
df_all['VersionID']=0
df_all['SID']=1950


# In[857]:


df_all['Status']='Add'


# In[858]:


df_all['Definition'] = df_all.apply(lambda x: 0  if x['compare']==False else x['Definition'], axis=1)


# In[859]:


df_all['Definition']=df_all['Definition'].fillna('0')


# In[860]:


df_all['UCode']=df_all['UCode'].fillna('U0')
df_all['QuestionID']=df_all['QuestionID'].fillna(0)
df_all['QUESTID']=df_all['QUESTID'].fillna(0)
df_all['SDID']=df_all['SDID'].fillna(0)


# In[861]:


df_all['Initial_wave']=df_all['Initial_Wave'].replace(r'nan',np.nan, regex=True)
df_all['Initial_wave']=df_all['Initial_Wave'].fillna(0)


# In[862]:


df_all['StudyAnswerID']=0
df_all['Full_Label']=''


# In[863]:


df_all['AnswerID']=df_all['AnswerID'].fillna(0)


# In[864]:


df_all['Imported']=''
df_all['Min']=''
df_all['Max']=''


# In[865]:


df_all=df_all.sort_values(['Category','Detail1','Detail3'],ascending=[True,True,True])


# In[866]:


#df_all = df_all.astype( {"QLevel":'int32', "QUESTID":'int32', "AnswerID":'int32',"QuestionID":'int32',"SID":'int64', "SDID":'int32', "VersionID":'int32', "Wave":'int32', "Min":'float',"Max":'float', "StudyEntryID":'int64',"Imported":'bool'} )


# In[867]:


#import difflib as dl


# In[868]:


#list1=df_all['Show_Name']


# In[869]:


#list2=df_all['Detail1']


# In[870]:


#list1=list(list1)
#list2=list(list2)


# In[871]:


#dl.context_diff(list1,list2)


# In[872]:


#for diff in dl.context_diff(list1,list2):
   # print(diff)


# In[911]:


df_all['Diff_Detail1'] = df_all.apply(lambda x: 'Same'  if x['Show_Name']==x['Detail1'] else 'Not_Same', axis=1)


# In[917]:


df_all_with_detail1_diff=df_all[['Show_Name','Detail1','Diff_Detail1','Clean_Type']]


# In[918]:


type(df_all_with_detail1_diff)


# In[921]:


df_all_with_detail1_diff.drop_duplicates(inplace=True)
#dataframe.where(dataframe.ID=='1')


# In[922]:


df_all_with_detail1_diff


# In[878]:


df_all_with_detail1_diff.shape


# In[879]:


df_all.shape


# In[873]:


df_all['Show_Name'] = df_all.apply(lambda x: '#'+ x['Show_Name']  if x['DP_Status']=='#' else x['Show_Name'], axis=1)
#df_TV_Movie['Shows_Name'] = df_TV_Movie.apply(lambda x: '#'+x['Shows_Name']  if x['OneWave_Suppress']=='#' else x['Shows_Name'], axis=1)


# In[875]:


df_all.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_all.csv",index=False,header=True,encoding='cp1252')


# In[ ]:


df_all_with_detail1_diff.to_csv(r"C:\Users\saraswathy.rajaman\Documents\df_all_with_detail1_diff.csv",index=False,header=True,encoding='cp1252')


# In[ ]:


df_all.columns


# In[ ]:





# In[ ]:


df_all.drop(['Clean_Type', 'Detail1','F2021_Client','compare','col2pv','LastDigit_PV','QID','CCP','Wave','ORD'], axis=1, inplace=True)


# In[ ]:


df_all.rename(columns={'S2022_Client':'CCP','Show_Name':'Detail1','VersionID':'Version','Initial_Wave':'Wave','ORD_new':'ORD'},inplace=True)


# In[ ]:


df_all=df_all[["StudyEntryID","SID","Version","Category","Super","Tmpl","Time Period","Detail1","Detail2",
"Detail3","Detail4","UCode","Definition","CCP","ORD","Wave","Status","Full_Label","QLevel","QUESTID","AnswerID","EditedBy","EditedDate","SDID",
"StudyAnswerID","QuestionID","Imported","Min","Max"]]


# In[ ]:


df_all['Max'] =df_all['Max'].apply(pd.to_numeric)
df_all['Min'] =df_all['Min'].apply(pd.to_numeric)


# In[ ]:


df_all['Wave']=df_all['Wave'].fillna(0)
df_all['Wave']=df_all['Wave'].replace(r'W', '', regex=True)


# In[ ]:





# In[ ]:


df_all.to_csv('C:\\Users\\saraswathy.rajaman\\Documents\\Spring-2022_needupdates.csv',index=False,header=True,encoding='cp1252')


# In[ ]:


#df_all.isnull().sum ()


# In[ ]:


df_all = df_all.astype( {"QLevel":'int32',
                         "QUESTID":'int32',
                         "AnswerID":'int32',
                         "QuestionID":'int64', 
                         "SDID":'int32', 
                         "Version":'int32', 
                         "Min":'float',
                         "Max":'float', 
                         "StudyEntryID":'int64',
                         "Imported":'bool',
                         "Tmpl":'int64',
                         "Wave":'int32',
                         "StudyAnswerID":'int32'} )


# In[ ]:


with engine.begin() as connection:
    df_merge_add_cab.to_sql(name="tmp_EditedRecords_addcable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
#df.to_sql('db_table2', engine, if_exists='replace')
df_merge_cable.to_sql(name="tmp_EditedRecords_cable_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_Movie.to_sql(name="tmp_EditedRecords_Movie_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_SPTV1.to_sql(name="tmp_EditedRecords_SPTV1_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_SPTV2.to_sql(name="tmp_EditedRecords_SPTV2_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_SPTV3.to_sql(name="tmp_EditedRecords_SPTV3_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_SPTV4.to_sql(name="tmp_EditedRecords_SPTV4_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_SPTV5.to_sql(name="tmp_EditedRecords_SPTV5_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_SPTV51.to_sql(name="tmp_EditedRecords_SPTV51_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_TV6.to_sql(name="tmp_EditedRecords_TV6_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_TV5.to_sql(name="tmp_EditedRecords_TV5_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_TV1.to_sql(name="tmp_EditedRecords_TV1_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_TV2.to_sql(name="tmp_EditedRecords_TV2_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_TV3.to_sql(name="tmp_EditedRecords_TV3_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_merge_TV4.to_sql(name="tmp_EditedRecords_TV4_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_all.to_sql(name="tmp_EditedRecords_dfall_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)
df_all_with_detail1_diff.to_sql(name="tmp_EditedRecords_dfall_detail1_diff_test",con=engine,schema="dbo",if_exists='replace', chunksize=1000,index=False)


# In[ ]:




